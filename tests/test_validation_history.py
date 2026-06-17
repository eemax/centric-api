from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from centric_api.cli import main


def test_validate_history_groups_latest_run_and_writes_outputs(
    tmp_path: Path,
    capsys,
) -> None:
    runs_dir = tmp_path / "runs"
    output_dir = tmp_path / "history"
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="run-old",
        started_at="2026-05-25T09:00:00Z",
        value=10.0,
        brand="CRAFT",
        trend="up",
    )
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="run-new",
        started_at="2026-06-03T09:00:00Z",
        value=20.0,
        brand="CRAFT",
        trend="up",
    )
    _write_history_run(
        runs_dir,
        validator="dpp-readiness",
        run_id="dpp-run",
        started_at="2026-06-03T09:00:00Z",
        value=30.0,
        brand="CRAFT",
    )

    assert (
        main(
            [
                "validate",
                "history",
                "--runs-dir",
                str(runs_dir),
                "--output-dir",
                str(output_dir),
                "--group",
                "week",
                "--validator",
                "style-readiness",
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["run_count"] == 2
    assert payload["raw_metric_count"] == 2
    assert payload["point_count"] == 2
    assert payload["validators"] == ["style-readiness"]
    history_payload = json.loads((output_dir / "history.json").read_text(encoding="utf-8"))
    assert len(history_payload["points"]) == 2
    latest_point = history_payload["points"][-1]
    assert latest_point["run_id"] == "run-new"
    assert latest_point["value"] == 20.0
    assert latest_point["trend"] == "up"
    workbook = load_workbook(output_dir / "history.xlsx", read_only=True)
    assert workbook.sheetnames == ["History", "Latest", "Runs"]
    latest_rows = list(workbook["Latest"].iter_rows(values_only=True))
    latest_headers = {header: index for index, header in enumerate(latest_rows[0])}
    assert latest_rows[1][latest_headers["Previous Value"]] == 10
    assert latest_rows[1][latest_headers["Change"]] == 10
    assert latest_rows[1][latest_headers["Change Percent"]] == 100
    assert latest_rows[1][latest_headers["Movement"]] == "good"
    html = (output_dir / "history.html").read_text(encoding="utf-8")
    assert "__HISTORY_JSON__" not in html
    assert "const historyPayload =" in html
    assert 'select id="brandSelect" multiple' in html
    assert 'id="allBrands"' in html
    assert 'select id="seasonType"' in html
    assert 'select id="seasonSelect" multiple' in html
    assert "brandFilter" not in html


def test_validate_history_counts_same_run_id_per_validator_and_skips_bad_schema(
    tmp_path: Path,
    capsys,
) -> None:
    runs_dir = tmp_path / "runs"
    output_dir = tmp_path / "history"
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="shared-run-id",
        started_at="2026-06-03T09:00:00Z",
        value=20.0,
        brand="CRAFT",
    )
    _write_history_run(
        runs_dir,
        validator="dpp-readiness",
        run_id="shared-run-id",
        started_at="2026-06-03T10:00:00Z",
        value=30.0,
        brand="CRAFT",
    )
    bad_history_dir = runs_dir / "bad-validator" / "bad-run"
    bad_history_dir.mkdir(parents=True)
    (bad_history_dir / "history.json").write_text(
        json.dumps({"schema_version": 1, "metrics": []}),
        encoding="utf-8",
    )

    assert (
        main(
            [
                "validate",
                "history",
                "--runs-dir",
                str(runs_dir),
                "--output-dir",
                str(output_dir),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["run_count"] == 2
    assert payload["raw_metric_count"] == 2
    history_payload = json.loads((output_dir / "history.json").read_text(encoding="utf-8"))
    assert history_payload["validators"] == ["dpp-readiness", "style-readiness"]


def test_validate_history_marks_downward_metric_improvement(
    tmp_path: Path,
    capsys,
) -> None:
    runs_dir = tmp_path / "runs"
    output_dir = tmp_path / "history"
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="run-old",
        started_at="2026-06-01T09:00:00Z",
        metric="Styles With Blocking Issues",
        value=8,
        unit="count",
        brand="CRAFT",
        trend="down",
    )
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="run-new",
        started_at="2026-06-08T09:00:00Z",
        metric="Styles With Blocking Issues",
        value=3,
        unit="count",
        brand="CRAFT",
        trend="down",
    )

    assert (
        main(
            [
                "validate",
                "history",
                "--runs-dir",
                str(runs_dir),
                "--output-dir",
                str(output_dir),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["point_count"] == 2
    workbook = load_workbook(output_dir / "history.xlsx", read_only=True)
    latest_rows = list(workbook["Latest"].iter_rows(values_only=True))
    latest_headers = {header: index for index, header in enumerate(latest_rows[0])}
    assert latest_rows[1][latest_headers["Trend"]] == "down"
    assert latest_rows[1][latest_headers["Previous Value"]] == 8
    assert latest_rows[1][latest_headers["Change"]] == -5
    assert latest_rows[1][latest_headers["Change Percent"]] == -62.5
    assert latest_rows[1][latest_headers["Movement"]] == "good"


def test_validate_history_keeps_season_dimensions_as_separate_series(
    tmp_path: Path,
    capsys,
) -> None:
    runs_dir = tmp_path / "runs"
    output_dir = tmp_path / "history"
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="run-one",
        started_at="2026-06-01T09:00:00Z",
        value=25,
        brand="CRAFT",
        trend="up",
        scope="brand_season",
        dimensions={
            "season_type": "cycle",
            "season_year": "2026",
            "season_slot": "1C",
            "season_label": "1C26",
        },
    )
    _write_history_run(
        runs_dir,
        validator="style-readiness",
        run_id="run-two",
        started_at="2026-06-01T10:00:00Z",
        value=75,
        brand="CRAFT",
        trend="up",
        scope="brand_season",
        dimensions={
            "season_type": "cycle",
            "season_year": "2026",
            "season_slot": "2C",
            "season_label": "2C26",
        },
    )

    assert (
        main(
            [
                "validate",
                "history",
                "--runs-dir",
                str(runs_dir),
                "--output-dir",
                str(output_dir),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["point_count"] == 2
    history_payload = json.loads((output_dir / "history.json").read_text(encoding="utf-8"))
    assert {point["dimensions"]["season_label"] for point in history_payload["points"]} == {
        "1C26",
        "2C26",
    }
    workbook = load_workbook(output_dir / "history.xlsx", read_only=True)
    history_rows = list(workbook["History"].iter_rows(values_only=True))
    history_headers = {header: index for index, header in enumerate(history_rows[0])}
    assert {row[history_headers["Season Label"]] for row in history_rows[1:]} == {
        "1C26",
        "2C26",
    }


def _write_history_run(
    runs_dir: Path,
    *,
    validator: str,
    run_id: str,
    started_at: str,
    value: float,
    brand: str,
    metric: str = "Style Completion %",
    unit: str = "percent",
    trend: str = "neutral",
    scope: str = "brand",
    dimensions: dict[str, str] | None = None,
) -> None:
    run_dir = runs_dir / validator / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "history.json").write_text(
        json.dumps(
            {
                "schema_version": 2,
                "validator": validator,
                "title": validator.replace("-", " ").title(),
                "run_id": run_id,
                "status": "failed",
                "started_at": started_at,
                "finished_at": started_at,
                "metrics": [
                    {
                        "scope": scope,
                        "brand": brand,
                        "metric": metric,
                        "value": value,
                        "unit": unit,
                        "trend": trend,
                        "numerator": value,
                        "denominator": 100,
                        "dimensions": dimensions or {},
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
