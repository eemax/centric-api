from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from centric_api.cli import main
from centric_api.config import ConfigError
from centric_api.store import connect
from centric_api.validation import ValidationFinding, ValidationFindingTotals
from centric_api.validation.registry import discover_validators


def test_validate_list_is_empty_without_private_validators(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    monkeypatch.setenv("CENTRIC_API_HOME", str(tmp_path))

    assert main(["validate", "list"]) == 0

    output = capsys.readouterr().out
    assert "Validators: 0" in output


def test_validate_cli_runs_private_validator_and_writes_artifacts(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    output_root = tmp_path / "validation-runs"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["run_status"] == "completed"
    assert payload["validation_outcome"] == "failed"
    assert payload["status"] == "failed"
    assert payload["summary"]["styles"] == 2
    assert payload["summary"]["styles_missing_name"] == 1
    assert payload["findings"] == 1
    assert payload["blocking"] == 1
    assert payload["finding_samples"][0]["code"] == "STYLE_NAME_MISSING"

    report_path = Path(payload["report_path"])
    assert report_path.is_file()
    assert Path(payload["summary_path"]).is_file()
    assert Path(payload["findings_path"]).is_file()
    findings_payload = json.loads(Path(payload["findings_path"]).read_text(encoding="utf-8"))
    assert findings_payload["total_findings"] == 1
    assert findings_payload["exported_findings"] == 1
    assert findings_payload["truncated"] is False
    assert findings_payload["findings"][0]["code"] == "STYLE_NAME_MISSING"

    workbook = load_workbook(report_path, read_only=False)
    assert workbook.sheetnames[:3] == ["Summary", "Styles", "Findings"]
    assert workbook["Styles"].auto_filter.ref == "A1:C3"
    assert not workbook["Styles"].tables
    rows = list(workbook["Styles"].iter_rows(values_only=True))
    assert rows[0] == ("Style Id", "Style Name", "Status")
    assert rows[1] == ("S1", "Style One", "ok")
    assert rows[2] == ("S2", None, "error")


def test_validate_human_output_distinguishes_run_from_outcome(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    output_root = tmp_path / "validation-runs"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
            ]
        )
        == 0
    )

    output = capsys.readouterr().out
    assert "Run:      completed" in output
    assert "Outcome:  failed" in output
    assert "ID:       " in output
    assert "Findings\nTotal:" in output
    assert "Blocking: 1" in output
    assert "BLOCKING STYLE_NAME_MISSING" in output
    assert "ERROR STYLE_NAME_MISSING" not in output


def test_validate_cli_show_private_validator(tmp_path: Path, capsys) -> None:
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")

    assert (
        main(
            [
                "validate",
                "show",
                "style-name-check",
                "--validators-dir",
                str(validators_dir),
            ]
        )
        == 0
    )

    output = capsys.readouterr().out
    assert "Validator: style-name-check" in output
    assert "Needs: styles" in output


def test_validation_finding_totals_can_count_findings() -> None:
    totals = ValidationFindingTotals.from_findings(
        (
            ValidationFinding(
                severity="error",
                code="ERROR",
                message="Error.",
            ),
            ValidationFinding(
                severity="warning",
                code="WARNING",
                message="Warning.",
            ),
            ValidationFinding(
                severity="info",
                code="INFO",
                message="Info.",
            ),
        )
    )

    assert totals == ValidationFindingTotals(findings=3, errors=1, warnings=1, info=1)


def test_validation_artifacts_can_cap_raw_finding_exports(tmp_path: Path) -> None:
    from centric_api.validation.artifacts import write_validation_artifacts
    from centric_api.validation.contracts import ValidationFinding, ValidationResult

    findings = tuple(
        ValidationFinding(
            severity="error",
            code="STYLE_NAME_MISSING",
            message=f"Style {index} is missing node_name.",
            endpoint="styles",
            record_id=f"S{index}",
        )
        for index in range(3)
    )
    result = ValidationResult(
        summary={"styles": 3},
        findings=findings,
        findings_export_limit=1,
    )

    report_path, summary_path, findings_path = write_validation_artifacts(
        tmp_path,
        result,
        run_record={
            "run_id": "run-1",
            "validator": "style-name-check",
            "title": "Style Name Check",
            "status": "failed",
            "started_at": "2026-01-01T00:00:00Z",
            "finished_at": "2026-01-01T00:00:01Z",
            "findings": 3,
            "errors": 3,
            "warnings": 0,
            "info": 0,
        },
    )

    assert report_path.is_file()
    assert summary_path.is_file()
    payload = json.loads(findings_path.read_text(encoding="utf-8"))
    assert payload["truncated"] is True
    assert payload["total_findings"] == 3
    assert payload["errors"] == 3
    assert payload["warnings"] == 0
    assert payload["info"] == 0
    assert payload["exported_findings"] == 1
    assert payload["findings"][0]["record_id"] == "S0"

    workbook = load_workbook(report_path, read_only=False)
    assert workbook["Findings"].auto_filter.ref == "A1:N2"
    assert not workbook["Findings"].tables
    rows = list(workbook["Findings"].iter_rows(values_only=True))
    assert len(rows) == 2
    assert rows[1][8] == "S0"


def test_validation_artifacts_can_use_custom_report_workbook(tmp_path: Path) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    from centric_api.validation.artifacts import write_validation_artifacts
    from centric_api.validation.contracts import ValidationResult

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP Summary"
    sheet.append(["Metric", "ALL"])
    sheet.append(["Ready Styles", 0])
    buffer = BytesIO()
    workbook.save(buffer)

    result = ValidationResult(
        summary={"styles": 1},
        report_workbook=buffer.getvalue(),
    )

    report_path, summary_path, findings_path = write_validation_artifacts(
        tmp_path,
        result,
        run_record={
            "run_id": "run-1",
            "validator": "dpp-readiness",
            "title": "DPP Readiness",
            "status": "failed",
            "started_at": "2026-01-01T00:00:00Z",
            "finished_at": "2026-01-01T00:00:01Z",
            "findings": 0,
            "errors": 0,
            "warnings": 0,
            "info": 0,
        },
    )

    assert summary_path.is_file()
    assert findings_path.is_file()
    report = load_workbook(report_path, read_only=True)
    assert report.sheetnames == ["DPP Summary"]
    rows = list(report["DPP Summary"].iter_rows(values_only=True))
    assert rows == [("Metric", "ALL"), ("Ready Styles", 0)]


def test_validation_run_can_use_explicit_samples_and_totals(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    output_root = tmp_path / "validation-runs"
    validators_dir.mkdir()
    _write_sampled_validator(validators_dir / "sampled.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "sampled-check",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["findings"] == 100
    assert payload["errors"] == 90
    assert payload["warnings"] == 10
    assert payload["finding_samples"] == [
        {
            "brand": None,
            "code": "SAMPLE",
            "details": {},
            "endpoint": "styles",
            "message": "Sample finding.",
            "record_id": "S1",
            "record_name": None,
            "season": None,
            "severity": "error",
            "source_endpoint": None,
            "source_field": None,
            "source_record_id": None,
            "style_id": None,
            "style_name": None,
        }
    ]
    findings_payload = json.loads(Path(payload["findings_path"]).read_text(encoding="utf-8"))
    assert findings_payload["total_findings"] == 100
    assert findings_payload["exported_findings"] == 1
    assert findings_payload["truncated"] is True


def test_validation_run_rejects_inconsistent_explicit_totals(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_invalid_totals_validator(validators_dir / "invalid_totals.py")
    _seed_styles_cache(db_path)
    output_root = tmp_path / "validation-runs"

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "invalid-totals",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "finding_totals cannot be smaller than exported samples" in captured.err
    assert not (output_root / "invalid-totals").exists()


def test_validation_run_rejects_invalid_custom_report_workbook(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_invalid_report_workbook_validator(validators_dir / "invalid_report_workbook.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "invalid-report-workbook",
                "--db",
                str(db_path),
                "--output-dir",
                str(tmp_path / "validation-runs"),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "report_workbook must be bytes" in captured.err


def test_validate_run_all_requires_private_validators(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    monkeypatch.setenv("CENTRIC_API_HOME", str(tmp_path))

    assert main(["validate", "run", "all"]) == 1

    captured = capsys.readouterr()
    assert captured.err == "Error: No validators found.\n"


def test_private_validator_required_endpoints_are_preflighted(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")
    with connect(db_path):
        pass

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "Missing cached endpoint; fetch one of: styles." in captured.err


def test_validator_registry_rejects_duplicate_private_names(tmp_path: Path) -> None:
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "one.py")
    _write_style_name_validator(validators_dir / "two.py")

    try:
        discover_validators(validators_dir)
    except ConfigError as exc:
        assert "Duplicate private validator name" in str(exc)
    else:  # pragma: no cover - keeps the assertion message crisp
        raise AssertionError("expected duplicate validator error")


def _write_style_name_validator(path: Path) -> None:
    path.write_text(
        """
from centric_api.validation import (
    ValidationDefinition,
    ValidationFinding,
    ValidationResult,
    ValidationSheet,
)


class StyleNameValidator:
    definition = ValidationDefinition(
        name="style-name-check",
        title="Style Name Check",
        required_endpoints=("styles",),
        description="Checks that cached styles have names.",
    )

    def run(self, ctx):
        findings = []
        rows = []
        for style in ctx.records("styles"):
            style_id = style.get("id")
            style_name = style.get("node_name")
            if not style_name:
                findings.append(
                    ValidationFinding(
                        severity="error",
                        code="STYLE_NAME_MISSING",
                        message="Style is missing node_name.",
                        endpoint="styles",
                        record_id=style_id,
                    )
                )
            rows.append(
                {
                    "style_id": style_id,
                    "style_name": style_name,
                    "status": "ok" if style_name else "error",
                }
            )
        return ValidationResult(
            summary={
                "styles": len(rows),
                "styles_missing_name": len(findings),
            },
            findings=tuple(findings),
            sheets=(ValidationSheet("Styles", tuple(rows)),),
        )


VALIDATOR = StyleNameValidator()
""",
        encoding="utf-8",
    )


def _write_sampled_validator(path: Path) -> None:
    path.write_text(
        """
from centric_api.validation import (
    ValidationDefinition,
    ValidationFinding,
    ValidationFindingTotals,
    ValidationResult,
)


class SampledValidator:
    definition = ValidationDefinition(
        name="sampled-check",
        title="Sampled Check",
        required_endpoints=("styles",),
    )

    def run(self, ctx):
        return ValidationResult(
            summary={"styles": len(ctx.records("styles"))},
            finding_samples=(
                ValidationFinding(
                    severity="error",
                    code="SAMPLE",
                    message="Sample finding.",
                    endpoint="styles",
                    record_id="S1",
                ),
            ),
            finding_totals=ValidationFindingTotals(
                findings=100,
                errors=90,
                warnings=10,
            ),
        )


VALIDATOR = SampledValidator()
""",
        encoding="utf-8",
    )


def _write_invalid_totals_validator(path: Path) -> None:
    path.write_text(
        """
from centric_api.validation import (
    ValidationDefinition,
    ValidationFinding,
    ValidationFindingTotals,
    ValidationResult,
)


class InvalidTotalsValidator:
    definition = ValidationDefinition(
        name="invalid-totals",
        title="Invalid Totals",
        required_endpoints=("styles",),
    )

    def run(self, ctx):
        return ValidationResult(
            summary={},
            finding_samples=(
                ValidationFinding(
                    severity="error",
                    code="SAMPLE",
                    message="Sample finding.",
                    endpoint="styles",
                    record_id="S1",
                ),
            ),
            finding_totals=ValidationFindingTotals(findings=1, errors=0),
        )


VALIDATOR = InvalidTotalsValidator()
""",
        encoding="utf-8",
    )


def _write_invalid_report_workbook_validator(path: Path) -> None:
    path.write_text(
        """
from centric_api.validation import (
    ValidationDefinition,
    ValidationResult,
)


class InvalidReportWorkbookValidator:
    definition = ValidationDefinition(
        name="invalid-report-workbook",
        title="Invalid Report Workbook",
        required_endpoints=("styles",),
    )

    def run(self, ctx):
        return ValidationResult(
            summary={},
            report_workbook="not bytes",
        )


VALIDATOR = InvalidReportWorkbookValidator()
""",
        encoding="utf-8",
    )


def _seed_styles_cache(db_path: Path) -> None:
    with connect(db_path) as conn:
        _insert_record(conn, "styles", "S1", {"id": "S1", "node_name": "Style One"})
        _insert_record(conn, "styles", "S2", {"id": "S2"})


def _insert_record(
    conn: sqlite3.Connection,
    endpoint: str,
    record_id: str,
    payload: dict[str, object],
) -> None:
    conn.execute(
        """
        INSERT INTO endpoint_records (
            endpoint, record_id, payload_json, payload_sha256, modified_at,
            source_file, source_run_id, ingested_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            endpoint,
            record_id,
            json.dumps(payload, sort_keys=True),
            f"hash-{endpoint}-{record_id}",
            None,
            f"{endpoint}.jsonl",
            "run-1",
            "2026-01-01T00:00:00Z",
        ],
    )
