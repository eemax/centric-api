from __future__ import annotations

import re
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from centric_api.store import connect
from centric_api.view_config import load_view_config, select_view
from centric_api.view_export import export_view
from tests.helpers_view import _insert_record, _seed_bom_line_view_records, _view_config


def test_view_export_writes_flat_xlsx_from_joined_cache_records(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    _seed_bom_line_view_records(db_path)
    config = load_view_config(_view_config(tmp_path))
    view = select_view(config, "bom-lines")

    result = export_view(
        db_path,
        config,
        view,
        export_format="xlsx",
        output_path=tmp_path / "bom-lines.xlsx",
    )

    assert result.row_count == 2
    assert result.column_count == 6
    assert result.output_path.is_file()
    workbook = load_workbook(result.output_path)
    sheet = workbook.active
    assert sheet.title == "BOM Lines"
    assert sheet.freeze_panes == "A2"
    assert len(sheet.tables) == 1
    assert sheet.auto_filter.ref is None
    assert sheet["A1"].font.bold is True
    assert [cell.value for cell in sheet[1]] == [
        "BOM Line ID",
        "Style",
        "Colorway",
        "Season",
        "Supplier",
        "Quantity",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "BL1",
        "Linen Shirt",
        "Ivory",
        "SS26",
        "Acme Mills",
        2.5,
    ]
    assert [cell.value for cell in sheet[3]] == [
        "BL2",
        "Linen Shirt",
        "Ivory",
        "SS26",
        "Acme Mills",
        1,
    ]


def test_view_export_default_output_path_uses_shared_artifact_name(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("CENTRIC_API_HOME", str(tmp_path / "home"))
    db_path = tmp_path / "centric.db"
    _seed_bom_line_view_records(db_path)
    config = load_view_config(_view_config(tmp_path))
    view = select_view(config, "bom-lines")

    first = export_view(db_path, config, view, export_format="xlsx")
    second = export_view(db_path, config, view, export_format="xlsx")

    assert first.output_path.parent == config.output_dir
    assert re.fullmatch(r"bom-lines-\d{4}-\d{2}-\d{2}-\d{4}\.xlsx", first.output_path.name)
    assert re.fullmatch(r"bom-lines-\d{4}-\d{2}-\d{2}-\d{4}-2\.xlsx", second.output_path.name)


def test_view_export_leaves_blank_xlsx_cells_untyped(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="bom_lines",
            record_id="BL1",
            payload={"id": "BL1", "style": "MISSING"},
        )
    config = load_view_config(_view_config(tmp_path))
    result = export_view(
        db_path,
        config,
        select_view(config, "bom-lines"),
        export_format="xlsx",
        output_path=tmp_path / "blank-cells.xlsx",
    )

    with ZipFile(result.output_path) as workbook:
        sheet_xml = workbook.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert 'r="B2"' not in sheet_xml


def test_view_export_escapes_formula_like_xlsx_text(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={"id": "S1", "node_name": "=SUM(1,1)", "code": ""},
        )
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: formula-text
    root:
      endpoint: styles
      as: style
    columns:
      - header: Code
        path: style.code
      - header: Name
        path: style.node_name
""",
        encoding="utf-8",
    )
    config = load_view_config(config_path)
    result = export_view(
        db_path,
        config,
        select_view(config, "formula-text"),
        export_format="xlsx",
        output_path=tmp_path / "formula-text.xlsx",
    )

    sheet = load_workbook(result.output_path).active

    assert sheet["A2"].value is None
    assert sheet["B2"].value == "'=SUM(1,1)"
    assert sheet["B2"].data_type == "s"


def test_view_integer_columns_do_not_truncate_decimals(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={"id": "S1", "whole": "2.0", "fractional": "1.9"},
        )
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: integers
    root:
      endpoint: styles
      as: style
    columns:
      - header: Whole
        path: style.whole
        type: integer
      - header: Fractional
        path: style.fractional
        type: integer
""",
        encoding="utf-8",
    )
    config = load_view_config(config_path)
    result = export_view(
        db_path,
        config,
        select_view(config, "integers"),
        export_format="xlsx",
        output_path=tmp_path / "integers.xlsx",
    )

    sheet = load_workbook(result.output_path).active

    assert sheet["A2"].value == 2
    assert sheet["B2"].value == "1.9"


def test_view_xlsx_export_preserves_existing_file_when_write_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={"id": "S1", "node_name": "Linen Shirt"},
        )
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: styles
    root:
      endpoint: styles
      as: style
    columns:
      - header: Style
        path: style.node_name
""",
        encoding="utf-8",
    )
    output_path = tmp_path / "styles.xlsx"
    output_path.write_bytes(b"existing")

    def fail_save(_workbook, filename) -> None:
        Path(filename).write_bytes(b"partial")
        raise RuntimeError("xlsx write failed")

    monkeypatch.setattr("openpyxl.workbook.workbook.Workbook.save", fail_save)

    config = load_view_config(config_path)
    with pytest.raises(RuntimeError, match="xlsx write failed"):
        export_view(
            db_path,
            config,
            select_view(config, "styles"),
            export_format="xlsx",
            output_path=output_path,
        )

    assert output_path.read_bytes() == b"existing"
    assert not (tmp_path / ".styles.xlsx.tmp").exists()
