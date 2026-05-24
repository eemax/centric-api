from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from centric_api.cli import main
from centric_api.config import ConfigError
from centric_api.store import connect
from centric_api.view_config import load_view_config, select_view
from centric_api.view_export import export_view, materialize_view


def test_default_view_config_loads_style_colorways_demo() -> None:
    config = load_view_config(Path("config/views.yml"))

    view = select_view(config, "style-colorways-demo")

    assert view.root.endpoint == "styles"
    assert view.columns


def test_view_export_writes_flat_xlsx_from_joined_cache_records(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    _seed_bom_line_view_records(db_path)
    config = load_view_config(_view_config(tmp_path))
    view = select_view(config, "bom-lines")

    result = export_view(
        db_path,
        config,
        view,
        export_format="xlsx",
        output_path=tmp_path / "bom-lines.xlsx",
    )

    assert result.row_count == 2
    assert result.column_count == 6
    assert result.output_path.is_file()
    workbook = load_workbook(result.output_path)
    sheet = workbook.active
    assert sheet.title == "BOM Lines"
    assert sheet.freeze_panes == "A2"
    assert len(sheet.tables) == 1
    assert sheet.auto_filter.ref is None
    assert sheet.row_dimensions[1].height == 18
    assert sheet["A1"].fill.fgColor.rgb == "001F4E78"
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].font.sz == 11
    assert sheet["A1"].font.color.rgb == "00FFFFFF"
    assert sheet["A1"].border.bottom.style == "thin"
    assert sheet["A1"].alignment.wrap_text is None
    assert sheet["B2"].alignment.vertical == "top"
    assert sheet["B2"].alignment.wrap_text is None
    assert sheet["F2"].alignment.wrap_text is None
    assert [cell.value for cell in sheet[1]] == [
        "BOM Line ID",
        "Style",
        "Colorway",
        "Season",
        "Supplier",
        "Quantity",
    ]
    assert [cell.value for cell in sheet[2]] == [
        "BL1",
        "Linen Shirt",
        "Ivory",
        "SS26",
        "Acme Mills",
        2.5,
    ]
    assert [cell.value for cell in sheet[3]] == [
        "BL2",
        "Linen Shirt",
        "Ivory",
        "SS26",
        "Acme Mills",
        1,
    ]


def test_view_many_concat_preserves_row_grain_and_csv_output(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={
                "id": "S1",
                "node_name": "Linen Shirt",
                "active": True,
                "documents": ["D1", "D2", "D3"],
            },
        )
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S2",
            payload={
                "id": "S2",
                "node_name": "Inactive Style",
                "active": False,
                "documents": ["D1"],
            },
        )
        _insert_record(
            conn,
            endpoint="documents",
            record_id="D1",
            payload={"id": "D1", "node_name": "Spec.pdf", "active": True},
        )
        _insert_record(
            conn,
            endpoint="documents",
            record_id="D2",
            payload={"id": "D2", "node_name": "Artwork.ai", "active": True},
        )
        _insert_record(
            conn,
            endpoint="documents",
            record_id="D3",
            payload={"id": "D3", "node_name": "Old.pdf", "active": False},
        )

    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
output_dir: exports
views:
  - name: style-docs
    title: Style Documents
    root:
      endpoint: styles
      as: style
    joins:
      - as: document
        endpoint: documents
        from: style.documents
        to: id
        relationship: many_concat
        separator: " | "
        filters:
          - path: document.active
            equals: true
          - path: document.node_name
            matches: '\\.pdf$'
    filters:
      - path: style.active
        equals: true
      - path: document.id
        exists: true
    columns:
      - header: Style
        path: style.node_name
      - header: Documents
        path: document.node_name
""",
        encoding="utf-8",
    )
    config = load_view_config(config_path)
    view = select_view(config, "style-docs")

    result = export_view(
        db_path,
        config,
        view,
        export_format="csv",
        output_path=tmp_path / "style-docs.csv",
    )

    assert result.row_count == 1
    with result.output_path.open("r", encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))
    assert rows == [["Style", "Documents"], ["Linen Shirt", "Spec.pdf"]]


def test_view_rejects_independent_expansion_chains(tmp_path: Path) -> None:
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: invalid
    root:
      endpoint: styles
      as: style
    joins:
      - as: colorway
        endpoint: colorways
        from: style.id
        to: style
        relationship: many_expand
      - as: document
        endpoint: documents
        from: style.documents
        to: id
        relationship: many_expand
    columns:
      - header: Style
        path: style.node_name
""",
        encoding="utf-8",
    )

    with pytest.raises(ConfigError, match="multiple independent many_expand"):
        load_view_config(config_path)


def test_view_allows_one_join_inside_expansion_chain(tmp_path: Path) -> None:
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: nested
    root:
      endpoint: styles
      as: style
    joins:
      - as: colorway
        endpoint: colorways
        from: style.id
        to: style
        relationship: many_expand
      - as: bom
        endpoint: boms
        from: colorway.bom
        to: id
        relationship: one
      - as: bom_line
        endpoint: bom_lines
        from: bom.id
        to: bom
        relationship: many_expand
    columns:
      - header: Style
        path: style.node_name
      - header: BOM Line
        path: bom_line.node_name
""",
        encoding="utf-8",
    )

    config = load_view_config(config_path)

    assert select_view(config, "nested").joins[-1].alias == "bom_line"


def test_view_cli_list_show_and_export_json(tmp_path: Path, capsys) -> None:
    db_path = tmp_path / "centric.db"
    _seed_bom_line_view_records(db_path)
    config_path = _view_config(tmp_path)
    output_path = tmp_path / "export.csv"

    assert main(["view", "list", "--view-config", str(config_path)]) == 0
    list_output = capsys.readouterr().out
    assert "Configured Views" in list_output
    assert "bom-lines" in list_output

    assert main(["view", "show", "bom-lines", "--view-config", str(config_path)]) == 0
    show_output = capsys.readouterr().out
    assert "View: bom-lines" in show_output
    assert "many_expand" not in show_output

    assert (
        main(
            [
                "view",
                "check",
                "bom-lines",
                "--view-config",
                str(config_path),
                "--db",
                str(db_path),
                "--json",
            ]
        )
        == 0
    )
    check_payload = json.loads(capsys.readouterr().out)
    assert check_payload["view"] == "bom-lines"
    assert check_payload["ok"] is True
    assert check_payload["rows_scanned"] == 2
    assert check_payload["rows_projected"] == 2
    assert check_payload["missing_join_details"] == []

    assert (
        main(
            [
                "view",
                "export",
                "bom-lines",
                "--view-config",
                str(config_path),
                "--db",
                str(db_path),
                "--output",
                str(output_path),
                "--json",
            ]
        )
        == 0
    )
    payload = json.loads(capsys.readouterr().out)
    assert payload["view"] == "bom-lines"
    assert payload["format"] == "csv"
    assert payload["rows"] == 2
    assert payload["missing_join_details"] == []
    assert output_path.is_file()


def test_view_missing_join_defaults_to_blank(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        for index in range(12):
            _insert_record(
                conn,
                endpoint="bom_lines",
                record_id=f"BL{index:02d}",
                payload={"id": f"BL{index:02d}", "style": f"MISSING{index:02d}"},
            )
    config = load_view_config(_view_config(tmp_path))
    view = select_view(config, "bom-lines")

    materialized = materialize_view(db_path, view)

    assert materialized.missing_join_count == 48
    assert [
        (item.alias, item.endpoint, item.missing_count)
        for item in materialized.missing_join_details
    ] == [
        ("style", "styles", 12),
        ("colorway", "colorways", 12),
        ("season", "seasons", 12),
        ("supplier", "suppliers", 12),
    ]
    assert materialized.missing_join_details[0].missing_ref_count == 12
    assert materialized.missing_join_details[0].missing_endpoint is True
    assert materialized.missing_join_details[0].sample_keys == tuple(
        f"MISSING{index:02d}" for index in range(10)
    )
    assert materialized.missing_join_details[1].missing_source_count == 12
    assert materialized.rows[0][1:5] == (None, None, None, None)


def test_view_export_leaves_blank_xlsx_cells_untyped(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="bom_lines",
            record_id="BL1",
            payload={"id": "BL1", "style": "MISSING"},
        )
    config = load_view_config(_view_config(tmp_path))
    result = export_view(
        db_path,
        config,
        select_view(config, "bom-lines"),
        export_format="xlsx",
        output_path=tmp_path / "blank-cells.xlsx",
    )

    with ZipFile(result.output_path) as workbook:
        sheet_xml = workbook.read("xl/worksheets/sheet1.xml").decode("utf-8")

    assert 'r="B2"' not in sheet_xml


def test_view_export_escapes_formula_like_xlsx_text(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={"id": "S1", "node_name": "=SUM(1,1)", "code": ""},
        )
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: formula-text
    root:
      endpoint: styles
      as: style
    columns:
      - header: Code
        path: style.code
      - header: Name
        path: style.node_name
""",
        encoding="utf-8",
    )
    config = load_view_config(config_path)
    result = export_view(
        db_path,
        config,
        select_view(config, "formula-text"),
        export_format="xlsx",
        output_path=tmp_path / "formula-text.xlsx",
    )

    sheet = load_workbook(result.output_path).active

    assert sheet["A2"].value is None
    assert sheet["B2"].value == "'=SUM(1,1)"
    assert sheet["B2"].data_type == "s"


def test_view_root_filter_runs_before_join_missing_counts(tmp_path: Path) -> None:
    db_path = tmp_path / "centric.db"
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="bom_lines",
            record_id="BL1",
            payload={"id": "BL1", "style": "S1", "active": True},
        )
        _insert_record(
            conn,
            endpoint="bom_lines",
            record_id="BL2",
            payload={"id": "BL2", "style": "MISSING", "active": False},
        )
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={"id": "S1", "node_name": "Linen Shirt"},
        )
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
views:
  - name: filtered
    root:
      endpoint: bom_lines
      as: bom_line
    joins:
      - as: style
        endpoint: styles
        from: bom_line.style
        to: id
        relationship: one
    filters:
      - path: bom_line.active
        equals: true
    columns:
      - header: BOM Line
        path: bom_line.id
      - header: Style
        path: style.node_name
""",
        encoding="utf-8",
    )
    config = load_view_config(config_path)

    materialized = materialize_view(db_path, select_view(config, "filtered"))

    assert materialized.missing_join_count == 0
    assert materialized.rows == (("BL1", "Linen Shirt"),)


def _view_config(tmp_path: Path) -> Path:
    config_path = tmp_path / "views.yml"
    config_path.write_text(
        """
version: 1
output_dir: exports
options:
  sheet_name: BOM Lines
views:
  - name: bom-lines
    title: BOM Lines
    root:
      endpoint: bom_lines
      as: bom_line
    joins:
      - as: style
        endpoint: styles
        from: bom_line.style
        to: id
        relationship: one
      - as: colorway
        endpoint: colorways
        from: bom_line.colorway
        to: id
        relationship: one
      - as: season
        endpoint: seasons
        from: style.season
        to: id
        relationship: one
      - as: supplier
        endpoint: suppliers
        from: style.supplier
        to: id
        relationship: one
    columns:
      - header: BOM Line ID
        path: bom_line.id
      - header: Style
        path: style.node_name
      - header: Colorway
        path: colorway.node_name
      - header: Season
        path: season.node_name
      - header: Supplier
        path: supplier.node_name
      - header: Quantity
        path: bom_line.quantity
        type: number
        number_format: "0.00"
""",
        encoding="utf-8",
    )
    return config_path


def _seed_bom_line_view_records(db_path: Path) -> None:
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="bom_lines",
            record_id="BL1",
            payload={"id": "BL1", "style": "S1", "colorway": "C1", "quantity": "2.5"},
        )
        _insert_record(
            conn,
            endpoint="bom_lines",
            record_id="BL2",
            payload={"id": "BL2", "style": "S1", "colorway": "C1", "quantity": 1},
        )
        _insert_record(
            conn,
            endpoint="styles",
            record_id="S1",
            payload={
                "id": "S1",
                "node_name": "Linen Shirt",
                "season": "SE1",
                "supplier": "SUP1",
            },
        )
        _insert_record(
            conn,
            endpoint="colorways",
            record_id="C1",
            payload={"id": "C1", "node_name": "Ivory", "style": "S1"},
        )
        _insert_record(
            conn, endpoint="seasons", record_id="SE1", payload={"id": "SE1", "node_name": "SS26"}
        )
        _insert_record(
            conn,
            endpoint="suppliers",
            record_id="SUP1",
            payload={"id": "SUP1", "node_name": "Acme Mills"},
        )


def _insert_record(
    conn: sqlite3.Connection,
    *,
    endpoint: str,
    record_id: str,
    payload: dict,
) -> None:
    payload_json = json.dumps(payload, sort_keys=True)
    conn.execute(
        """
        INSERT INTO endpoint_records (
            endpoint, record_id, payload_json, payload_sha256, modified_at,
            source_file, source_run_id, ingested_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            endpoint,
            record_id,
            payload_json,
            f"hash-{endpoint}-{record_id}",
            payload.get("_modified_at"),
            f"{endpoint}.jsonl",
            "test-run",
            "2026-01-01T00:00:00Z",
        ],
    )
