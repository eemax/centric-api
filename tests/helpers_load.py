from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _write_material_workbook(
    path: Path,
    *,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Materials"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)

def _insert_record(
    conn: sqlite3.Connection,
    *,
    endpoint: str,
    record_id: str,
    payload: dict[str, object],
) -> None:
    payload_json = json.dumps(payload, sort_keys=True)
    conn.execute(
        """
        INSERT INTO endpoint_records (
            endpoint, record_id, payload_json, payload_sha256, modified_at,
            source_file, source_run_id, ingested_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            endpoint,
            record_id,
            payload_json,
            f"hash-{endpoint}-{record_id}",
            None,
            f"{endpoint}.jsonl",
            "test-run",
            "2026-01-01T00:00:00Z",
        ],
    )

def _review_row(path: Path, row_number: int = 2) -> dict[str, object]:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[row_number]]
        return dict(zip(headers, values, strict=False))
    finally:
        workbook.close()

class _FakeAuthContext:
    base_url = "https://example.test"

    def request(self, method: str, url: str, *, json_body: object) -> object:
        assert method == "POST"
        assert url == "https://example.test/api/v2/materials"
        assert isinstance(json_body, dict)
        return _FakeResponse()

class _FakeResponse:
    status_code = 201
    text = '{"id":"created"}'

    def json(self) -> dict[str, str]:
        return {"id": "created"}

class _FailingAuthContext:
    base_url = "https://example.test"

    def request(self, method: str, url: str, *, json_body: object) -> object:
        raise RuntimeError("connection dropped")

class _MixedAuthContext:
    base_url = "https://example.test"

    def __init__(self) -> None:
        self.calls = 0

    def request(self, method: str, url: str, *, json_body: object) -> object:
        self.calls += 1
        if self.calls == 1:
            return _FakeResponse()
        return _FailedResponse()

class _FailedResponse:
    status_code = 400
    text = '{"message":"bad row"}'

    def json(self) -> dict[str, str]:
        return {"message": "bad row"}
