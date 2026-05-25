from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook, load_workbook

from centric_api.cli import main
from centric_api.load import materialize_load, run_load
from centric_api.load_config import load_load_config, parse_load_config, select_load_job
from centric_api.store import connect


def test_load_check_resolves_material_create_refs_and_alias_headers(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Material Code", "Material", "Type", "Desc"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Fabric", "Main body fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT0",
            payload={"id": "MT0", "node_name": "FABRIC", "available": False},
        )

    config = load_load_config()
    result = materialize_load(
        db_path,
        select_load_job(config, "material-create"),
        workbook_path,
    )

    assert result.valid_rows == 1
    assert result.issues == ()
    assert result.requests[0].body == {
        "code": "MAT-001",
        "node_name": "Cotton Rib 240 GSM",
        "product_type": "MT1",
        "description": "Main body fabric",
    }


def test_load_cli_dry_run_writes_request_artifacts(tmp_path, monkeypatch, capsys) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )

    assert (
        main(
            [
                "load",
                "run",
                "material-create",
                str(workbook_path),
                "--db",
                str(db_path),
                "--dry-run",
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["dry_run"] is True
    assert payload["requests"] == 1
    assert payload["request_samples"][0]["body"]["product_type"] == "MT1"
    requests_path = Path(payload["run_dir"]) / "requests.jsonl"
    assert requests_path.is_file()
    assert payload["review_workbook"] is None
    request_record = json.loads(requests_path.read_text(encoding="utf-8").splitlines()[0])
    assert request_record["method"] == "POST"
    assert request_record["path"] == "/v2/materials"


def test_load_cli_reports_human_progress(tmp_path, monkeypatch, capsys) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )

    assert (
        main(
            [
                "load",
                "run",
                "material-create",
                str(workbook_path),
                "--db",
                str(db_path),
                "--dry-run",
            ]
        )
        == 0
    )

    captured = capsys.readouterr()
    assert "[load] planning: job=material-create mode=dry-run" in captured.err
    assert "[load] headers: matched=3/4 required=3/3" in captured.err
    assert "[load] refs: material_types" in captured.err
    assert "[load] validate: scanned=1 valid=1 errors=0" in captured.err
    assert "[load] artifacts:" in captured.err
    assert "Load dry run: material-create" in captured.out


def test_load_cli_json_suppresses_human_progress(tmp_path, monkeypatch, capsys) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )

    assert (
        main(
            [
                "load",
                "check",
                "material-create",
                str(workbook_path),
                "--db",
                str(db_path),
                "--json",
            ]
        )
        == 0
    )

    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert payload["ok"] is True
    assert captured.err == ""


def test_load_run_emits_send_progress(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()
    events: list[dict[str, object]] = []

    result = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=_FakeAuthContext(),
        progress_callback=events.append,
    )

    assert result.success_count == 1
    assert result.review_path is not None
    send_events = [event for event in events if event["event"] == "load_send"]
    assert send_events[0]["index"] == 1
    assert send_events[0]["total"] == 1
    assert send_events[0]["status_code"] == 201
    review_row = _review_row(result.review_path)
    assert review_row["_cent_load_status"] == "success"
    assert review_row["_cent_load_status_code"] == 201
    assert review_row["_cent_load_response_id"] == "created"


def test_load_run_records_request_exceptions_as_failures(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()

    result = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=_FailingAuthContext(),
    )

    assert result.success_count == 0
    assert result.failure_count == 1
    assert result.responses[0].status_code == 0
    assert result.responses[0].body == {
        "error": "connection dropped",
        "type": "RuntimeError",
    }
    assert result.review_path is not None
    review_row = _review_row(result.review_path)
    assert review_row["_cent_load_status"] == "failed"
    assert review_row["_cent_load_status_code"] == 0
    assert review_row["_cent_load_message"] == "connection dropped"
    assert (result.run_dir / "responses.jsonl").is_file()
    assert (result.run_dir / "summary.json").is_file()


def test_load_retry_processes_failed_review_rows(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[
            ["MAT-001", "Cotton Rib 240 GSM", "Fabric"],
            ["MAT-002", "Cotton Jersey 180 GSM", "Fabric"],
        ],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()
    first = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=_MixedAuthContext(),
    )
    assert first.review_path is not None

    retry = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        first.review_path,
        sheet=None,
        limit=None,
        dry_run=True,
        yes=False,
        retry_statuses={"failed"},
    )

    assert retry.rows_scanned == 1
    assert retry.requests[0].row == 3
    assert retry.requests[0].body["code"] == "MAT-002"


def test_load_retry_review_clears_unprocessed_old_statuses(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[
            ["MAT-001", "Cotton Rib 240 GSM", "Fabric"],
            ["MAT-002", "Cotton Jersey 180 GSM", "Fabric"],
        ],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()
    first = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=_MixedAuthContext(),
    )
    assert first.review_path is not None

    retry = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        first.review_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        retry_statuses={"failed"},
        auth_ctx=_FakeAuthContext(),
    )

    assert retry.review_path is not None
    assert _review_row(retry.review_path, row_number=2)["_cent_load_status"] is None
    assert _review_row(retry.review_path, row_number=3)["_cent_load_status"] == "success"


def test_load_run_writes_validation_error_review(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Missing Type"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()

    result = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=_FakeAuthContext(),
    )

    assert result.issues
    assert result.success_count == 0
    assert result.failure_count == 0
    assert result.review_path is not None
    review_row = _review_row(result.review_path)
    assert review_row["_cent_load_status"] == "validation_error"
    assert "was not found" in review_row["_cent_load_message"]


def test_load_run_with_only_validation_errors_does_not_require_auth(
    tmp_path,
    monkeypatch,
) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Missing Type"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()

    result = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=None,
    )

    assert result.requests == ()
    assert result.issues
    assert result.review_path is not None


def test_load_run_header_errors_do_not_write_empty_review_file(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name"],
        rows=[["MAT-001", "Cotton Rib 240 GSM"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()

    result = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=None,
    )

    assert result.requests == ()
    assert [issue.code for issue in result.issues] == ["missing_required_header"]
    assert result.review_path is None


def test_load_run_sends_valid_rows_when_other_rows_have_validation_errors(
    tmp_path,
    monkeypatch,
) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Material Type"],
        rows=[
            ["MAT-001", "Cotton Rib 240 GSM", "Missing Type"],
            ["MAT-002", "Cotton Jersey 180 GSM", "Fabric"],
        ],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
    config = load_load_config()

    result = run_load(
        db_path,
        config,
        select_load_job(config, "material-create"),
        workbook_path,
        sheet=None,
        limit=None,
        dry_run=False,
        yes=True,
        auth_ctx=_FakeAuthContext(),
    )

    assert result.success_count == 1
    assert result.failure_count == 0
    assert len(result.issues) == 1
    assert result.review_path is not None
    assert _review_row(result.review_path, row_number=2)["_cent_load_status"] == "validation_error"
    assert _review_row(result.review_path, row_number=3)["_cent_load_status"] == "success"


def test_ref_indexes_keep_same_ref_with_different_filters(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Active Type", "Inactive Type"],
        rows=[["MAT-001", "Fabric", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT0",
            payload={"id": "MT0", "node_name": "Fabric", "available": False},
        )
    config = parse_load_config(
        {
            "version": 1,
            "jobs": [
                {
                    "name": "filtered-refs",
                    "method": "POST",
                    "path": "/v2/materials",
                    "columns": {
                        "code": {"header": "Code", "required": True},
                        "active_type": {
                            "header": "Active Type",
                            "type": "ref",
                            "required": True,
                            "resolve": {
                                "endpoint": "material_types",
                                "match": "node_name",
                                "output": "id",
                                "filters": {"available": True},
                            },
                        },
                        "inactive_type": {
                            "header": "Inactive Type",
                            "type": "ref",
                            "required": True,
                            "resolve": {
                                "endpoint": "material_types",
                                "match": "node_name",
                                "output": "id",
                                "filters": {"available": False},
                            },
                        },
                    },
                    "body": {
                        "code": "code",
                        "active": "active_type",
                        "inactive": "inactive_type",
                    },
                }
            ],
        },
        path=tmp_path / "load.yml",
    )

    result = materialize_load(
        db_path,
        select_load_job(config, "filtered-refs"),
        workbook_path,
    )

    assert result.issues == ()
    assert result.requests[0].body == {
        "code": "MAT-001",
        "active": "MT1",
        "inactive": "MT0",
    }


def test_load_check_fails_ambiguous_alias_headers(tmp_path, monkeypatch) -> None:
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setenv("CENTRIC_API_HOME", str(home))
    db_path = tmp_path / "centric.db"
    workbook_path = tmp_path / "materials.xlsx"
    _write_material_workbook(
        workbook_path,
        headers=["Code", "Material Name", "Name", "Material Type"],
        rows=[["MAT-001", "Cotton Rib 240 GSM", "Duplicate", "Fabric"]],
    )
    with connect(db_path) as conn:
        _insert_record(
            conn,
            endpoint="material_types",
            record_id="MT1",
            payload={"id": "MT1", "node_name": "Fabric", "available": True},
        )

    config = load_load_config()
    result = materialize_load(
        db_path,
        select_load_job(config, "material-create"),
        workbook_path,
    )

    assert [issue.code for issue in result.issues] == ["ambiguous_header"]
    assert result.issues[0].column == "node_name"


def _write_material_workbook(
    path: Path,
    *,
    headers: list[str],
    rows: list[list[object]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Materials"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _insert_record(
    conn: sqlite3.Connection,
    *,
    endpoint: str,
    record_id: str,
    payload: dict[str, object],
) -> None:
    payload_json = json.dumps(payload, sort_keys=True)
    conn.execute(
        """
        INSERT INTO endpoint_records (
            endpoint, record_id, payload_json, payload_sha256, modified_at,
            source_file, source_run_id, ingested_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            endpoint,
            record_id,
            payload_json,
            f"hash-{endpoint}-{record_id}",
            None,
            f"{endpoint}.jsonl",
            "test-run",
            "2026-01-01T00:00:00Z",
        ],
    )


def _review_row(path: Path, row_number: int = 2) -> dict[str, object]:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[row_number]]
        return dict(zip(headers, values, strict=False))
    finally:
        workbook.close()


class _FakeAuthContext:
    base_url = "https://example.test"

    def request(self, method: str, url: str, *, json_body: object) -> object:
        assert method == "POST"
        assert url == "https://example.test/api/v2/materials"
        assert isinstance(json_body, dict)
        return _FakeResponse()


class _FakeResponse:
    status_code = 201
    text = '{"id":"created"}'

    def json(self) -> dict[str, str]:
        return {"id": "created"}


class _FailingAuthContext:
    base_url = "https://example.test"

    def request(self, method: str, url: str, *, json_body: object) -> object:
        raise RuntimeError("connection dropped")


class _MixedAuthContext:
    base_url = "https://example.test"

    def __init__(self) -> None:
        self.calls = 0

    def request(self, method: str, url: str, *, json_body: object) -> object:
        self.calls += 1
        if self.calls == 1:
            return _FakeResponse()
        return _FailedResponse()


class _FailedResponse:
    status_code = 400
    text = '{"message":"bad row"}'

    def json(self) -> dict[str, str]:
        return {"message": "bad row"}
