from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from centric_api.cli import main
from centric_api.validation import (
    ValidationFinding,
    ValidationFindingTotals,
    ValidationHistoryMetric,
)
from tests.helpers_validation import (
    _seed_styles_cache,
    _write_invalid_history_metric_trend_validator,
    _write_invalid_history_metric_validator,
    _write_invalid_report_workbook_validator,
    _write_invalid_totals_validator,
    _write_sampled_validator,
)


def test_validation_finding_totals_can_count_findings() -> None:
    totals = ValidationFindingTotals.from_findings(
        (
            ValidationFinding(
                severity="error",
                code="ERROR",
                message="Error.",
            ),
            ValidationFinding(
                severity="warning",
                code="WARNING",
                message="Warning.",
            ),
            ValidationFinding(
                severity="info",
                code="INFO",
                message="Info.",
            ),
        )
    )

    assert totals == ValidationFindingTotals(findings=3, errors=1, warnings=1, info=1)


def test_validation_artifacts_can_cap_raw_finding_exports(tmp_path: Path) -> None:
    from centric_api.validation.artifacts import write_validation_artifacts
    from centric_api.validation.contracts import ValidationFinding, ValidationResult

    findings = tuple(
        ValidationFinding(
            severity="error",
            code="STYLE_NAME_MISSING",
            message=f"Style {index} is missing node_name.",
            endpoint="styles",
            record_id=f"S{index}",
        )
        for index in range(3)
    )
    result = ValidationResult(
        summary={"styles": 3},
        findings=findings,
        findings_export_limit=1,
    )

    report_path, summary_path, findings_path, history_path = write_validation_artifacts(
        tmp_path,
        result,
        run_record={
            "run_id": "run-1",
            "validator": "style-name-check",
            "title": "Style Name Check",
            "status": "failed",
            "started_at": "2026-01-01T00:00:00Z",
            "finished_at": "2026-01-01T00:00:01Z",
            "findings": 3,
            "errors": 3,
            "warnings": 0,
            "info": 0,
        },
    )

    assert report_path.is_file()
    assert report_path.name == "report_26-01-01-0000.xlsx"
    assert summary_path.is_file()
    assert history_path.is_file()
    payload = json.loads(findings_path.read_text(encoding="utf-8"))
    assert payload["truncated"] is True
    assert payload["total_findings"] == 3
    assert payload["errors"] == 3
    assert payload["warnings"] == 0
    assert payload["info"] == 0
    assert payload["exported_findings"] == 1
    assert payload["findings"][0]["record_id"] == "S0"

    workbook = load_workbook(report_path, read_only=False)
    assert workbook["Findings"].auto_filter.ref == "A1:N2"
    assert not workbook["Findings"].tables
    rows = list(workbook["Findings"].iter_rows(values_only=True))
    assert len(rows) == 2
    assert rows[1][8] == "S0"


def test_validation_artifacts_can_use_custom_report_workbook(tmp_path: Path) -> None:
    from io import BytesIO

    from openpyxl import Workbook

    from centric_api.validation.artifacts import write_validation_artifacts
    from centric_api.validation.contracts import ValidationResult

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DPP Summary"
    sheet.append(["Metric", "ALL"])
    sheet.append(["Ready Styles", 0])
    buffer = BytesIO()
    workbook.save(buffer)

    result = ValidationResult(
        summary={"styles": 1},
        report_workbook=buffer.getvalue(),
    )

    report_path, summary_path, findings_path, history_path = write_validation_artifacts(
        tmp_path,
        result,
        run_record={
            "run_id": "run-1",
            "validator": "dpp-readiness",
            "title": "DPP Readiness",
            "status": "failed",
            "started_at": "2026-01-01T00:00:00Z",
            "finished_at": "2026-01-01T00:00:01Z",
            "findings": 0,
            "errors": 0,
            "warnings": 0,
            "info": 0,
        },
    )

    assert summary_path.is_file()
    assert findings_path.is_file()
    assert history_path.is_file()
    assert report_path.name == "report_26-01-01-0000.xlsx"
    report = load_workbook(report_path, read_only=True)
    assert report.sheetnames == ["DPP Summary"]
    rows = list(report["DPP Summary"].iter_rows(values_only=True))
    assert rows == [("Metric", "ALL"), ("Ready Styles", 0)]


def test_validation_artifacts_write_history_metrics(tmp_path: Path) -> None:
    from centric_api.validation.artifacts import write_validation_artifacts
    from centric_api.validation.contracts import ValidationResult

    result = ValidationResult(
        summary={"styles": 10},
        history_metrics=(
            ValidationHistoryMetric(
                metric="Style Completion %",
                value=40.0,
                unit="percent",
                trend="up",
                scope="overall",
                numerator=4,
                denominator=10,
            ),
        ),
    )

    report_path, _summary_path, _findings_path, history_path = write_validation_artifacts(
        tmp_path,
        result,
        run_record={
            "run_id": "run-1",
            "validator": "style-readiness",
            "title": "Style Readiness",
            "status": "failed",
            "started_at": "2026-01-01T00:00:00Z",
            "finished_at": "2026-01-01T00:00:01Z",
            "findings": 0,
            "errors": 0,
            "warnings": 0,
            "info": 0,
        },
    )

    payload = json.loads(history_path.read_text(encoding="utf-8"))
    assert payload["schema_version"] == 2
    assert payload["validator"] == "style-readiness"
    assert payload["artifact_timestamp"] == "26-01-01-0000"
    assert payload["report_path"] == str(report_path)
    assert payload["metrics"] == [
        {
            "brand": None,
            "denominator": 10,
            "dimensions": {},
            "metric": "Style Completion %",
            "numerator": 4,
            "scope": "overall",
            "trend": "up",
            "unit": "percent",
            "value": 40.0,
        }
    ]


def test_validation_run_can_use_explicit_samples_and_totals(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    output_root = tmp_path / "validation-runs"
    validators_dir.mkdir()
    _write_sampled_validator(validators_dir / "sampled.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "sampled-check",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["findings"] == 100
    assert payload["errors"] == 90
    assert payload["warnings"] == 10
    assert payload["finding_samples"] == [
        {
            "brand": None,
            "code": "SAMPLE",
            "details": {},
            "endpoint": "styles",
            "message": "Sample finding.",
            "record_id": "S1",
            "record_name": None,
            "season": None,
            "severity": "error",
            "source_endpoint": None,
            "source_field": None,
            "source_record_id": None,
            "style_id": None,
            "style_name": None,
        }
    ]
    findings_payload = json.loads(Path(payload["findings_path"]).read_text(encoding="utf-8"))
    assert findings_payload["total_findings"] == 100
    assert findings_payload["exported_findings"] == 1
    assert findings_payload["truncated"] is True


def test_validation_run_rejects_inconsistent_explicit_totals(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_invalid_totals_validator(validators_dir / "invalid_totals.py")
    _seed_styles_cache(db_path)
    output_root = tmp_path / "validation-runs"

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "invalid-totals",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "finding_totals cannot be smaller than exported samples" in captured.err
    assert not (output_root / "invalid-totals").exists()


def test_validation_run_rejects_invalid_custom_report_workbook(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_invalid_report_workbook_validator(validators_dir / "invalid_report_workbook.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "invalid-report-workbook",
                "--db",
                str(db_path),
                "--output-dir",
                str(tmp_path / "validation-runs"),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "report_workbook must be bytes" in captured.err


def test_validation_run_rejects_invalid_history_metric(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_invalid_history_metric_validator(validators_dir / "invalid_history_metric.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "invalid-history-metric",
                "--db",
                str(db_path),
                "--output-dir",
                str(tmp_path / "validation-runs"),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "history metric Style Completion % value must be finite" in captured.err


def test_validation_run_rejects_invalid_history_metric_trend(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_invalid_history_metric_trend_validator(validators_dir / "invalid_history_trend.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "invalid-history-trend",
                "--db",
                str(db_path),
                "--output-dir",
                str(tmp_path / "validation-runs"),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "history metric Active Styles trend is invalid" in captured.err
