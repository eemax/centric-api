from __future__ import annotations

import json
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from centric_api.cli import main
from centric_api.config import ConfigError
from centric_api.store import connect
from centric_api.validation.registry import discover_validators
from centric_api.validation.runner import _allocate_output_dir
from tests.helpers_validation import _seed_styles_cache, _write_style_name_validator


@pytest.fixture(autouse=True)
def _isolate_runtime_home(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("CENTRIC_API_HOME", str(tmp_path / "home"))


def test_validate_list_is_empty_without_private_validators(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    monkeypatch.setenv("CENTRIC_API_HOME", str(tmp_path))

    assert main(["validate", "list"]) == 0

    output = capsys.readouterr().out
    assert "Validators: 0" in output


def test_validate_cli_runs_private_validator_and_writes_artifacts(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    output_root = tmp_path / "validation-runs"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
                "--json",
            ]
        )
        == 0
    )

    captured = capsys.readouterr()
    payload = json.loads(captured.out)
    assert captured.err == ""
    assert payload["run_status"] == "completed"
    assert payload["validation_outcome"] == "failed"
    assert payload["status"] == "failed"
    assert payload["summary"]["styles"] == 2
    assert payload["summary"]["styles_missing_name"] == 1
    assert re.fullmatch(r"\d{2}-\d{2}-\d{2}-\d{4}", payload["summary"]["artifact_timestamp"])
    assert re.fullmatch(r"style-name-check-\d{4}-\d{2}-\d{2}-\d{4}(?:-\d+)?", payload["run_id"])
    assert payload["findings"] == 1
    assert payload["blocking"] == 1
    assert payload["finding_samples"][0]["code"] == "STYLE_NAME_MISSING"

    report_path = Path(payload["report_path"])
    assert report_path.is_file()
    assert report_path.parent.name == payload["run_id"]
    assert re.fullmatch(r"report_\d{2}-\d{2}-\d{2}-\d{4}\.xlsx", report_path.name)
    assert Path(payload["summary_path"]).is_file()
    assert Path(payload["findings_path"]).is_file()
    assert Path(payload["history_path"]).is_file()
    findings_payload = json.loads(Path(payload["findings_path"]).read_text(encoding="utf-8"))
    assert findings_payload["total_findings"] == 1
    assert findings_payload["exported_findings"] == 1
    assert findings_payload["truncated"] is False
    assert findings_payload["findings"][0]["code"] == "STYLE_NAME_MISSING"

    workbook = load_workbook(report_path, read_only=False)
    assert workbook.sheetnames[:3] == ["Summary", "Styles", "Findings"]
    assert workbook["Styles"].auto_filter.ref == "A1:C3"
    assert not workbook["Styles"].tables
    rows = list(workbook["Styles"].iter_rows(values_only=True))
    assert rows[0] == ("Style Id", "Style Name", "Status")
    assert rows[1] == ("S1", "Style One", "ok")
    assert rows[2] == ("S2", None, "error")


def test_validation_output_dir_suffixes_same_minute_collisions(tmp_path: Path) -> None:
    first_run_id, first_output_dir = _allocate_output_dir(
        "style name check",
        started_at="2026-06-19T12:34:56+00:00",
        output_root=tmp_path,
    )
    second_run_id, second_output_dir = _allocate_output_dir(
        "style name check",
        started_at="2026-06-19T12:34:59+00:00",
        output_root=tmp_path,
    )

    assert first_run_id == "style-name-check-2026-06-19-1234"
    assert second_run_id == "style-name-check-2026-06-19-1234-2"
    assert first_output_dir.parent.name == "style-name-check"
    assert first_output_dir.is_dir()
    assert second_output_dir.is_dir()


def test_validate_run_passes_mode_and_input_file_to_context(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    input_file = tmp_path / "input.xlsx"
    validators_dir.mkdir()
    input_file.write_text("placeholder", encoding="utf-8")
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--mode",
                "excel",
                "--input-file",
                str(input_file),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["summary"]["mode"] == "excel"
    assert payload["summary"]["input_file"] == str(input_file)


def test_validate_run_input_file_implies_excel_mode(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    input_file = tmp_path / "input.xlsx"
    validators_dir.mkdir()
    input_file.write_text("placeholder", encoding="utf-8")
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--input-file",
                str(input_file),
                "--json",
            ]
        )
        == 0
    )

    payload = json.loads(capsys.readouterr().out)
    assert payload["summary"]["mode"] == "excel"
    assert payload["summary"]["input_file"] == str(input_file)


def test_validate_run_rejects_input_file_with_cache_mode(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    input_file = tmp_path / "input.xlsx"
    validators_dir.mkdir()
    input_file.write_text("placeholder", encoding="utf-8")
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--mode",
                "cache",
                "--input-file",
                str(input_file),
                "--json",
            ]
        )
        == 1
    )

    assert "--input-file requires --mode excel or no --mode." in capsys.readouterr().err


def test_validate_human_output_distinguishes_run_from_outcome(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    output_root = tmp_path / "validation-runs"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")
    _seed_styles_cache(db_path)

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
                "--output-dir",
                str(output_root),
            ]
        )
        == 0
    )

    captured = capsys.readouterr()
    output = captured.out
    progress = captured.err
    assert "Validation run: validators=1 mode=cache db=" in progress
    assert "[style-name-check] START  1/1" in progress
    assert "[style-name-check] DONE   status=failed findings=1 elapsed=" in progress
    assert "validation=done validators=1 findings=1 elapsed=" in progress
    assert "Run:      completed" in output
    assert "Outcome:  failed" in output
    assert "ID:       " in output
    assert "Findings\nTotal:" in output
    assert "Blocking: 1" in output
    assert "Value Sets:" in output
    assert "Product Group: enabled (104 values)" in output
    assert "'Product Group':" not in output
    assert "BLOCKING STYLE_NAME_MISSING" in output
    assert "ERROR STYLE_NAME_MISSING" not in output


def test_validate_cli_show_private_validator(tmp_path: Path, capsys) -> None:
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")

    assert (
        main(
            [
                "validate",
                "show",
                "style-name-check",
                "--validators-dir",
                str(validators_dir),
            ]
        )
        == 0
    )

    output = capsys.readouterr().out
    assert "Validator: style-name-check" in output
    assert "Needs: styles" in output


def test_validate_run_all_requires_private_validators(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    monkeypatch.setenv("CENTRIC_API_HOME", str(tmp_path))

    assert main(["validate", "run", "all"]) == 1

    captured = capsys.readouterr()
    assert captured.err == "Error: No validators found.\n"


def test_private_validator_required_endpoints_are_preflighted(
    tmp_path: Path,
    capsys,
) -> None:
    db_path = tmp_path / "centric.db"
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "style_names.py")
    with connect(db_path):
        pass

    assert (
        main(
            [
                "validate",
                "--validators-dir",
                str(validators_dir),
                "run",
                "style-name-check",
                "--db",
                str(db_path),
            ]
        )
        == 1
    )

    captured = capsys.readouterr()
    assert "Missing cached endpoint; fetch one of: styles." in captured.err


def test_validator_registry_rejects_duplicate_private_names(tmp_path: Path) -> None:
    validators_dir = tmp_path / "validators"
    validators_dir.mkdir()
    _write_style_name_validator(validators_dir / "one.py")
    _write_style_name_validator(validators_dir / "two.py")

    try:
        discover_validators(validators_dir)
    except ConfigError as exc:
        assert "Duplicate private validator name" in str(exc)
    else:  # pragma: no cover - keeps the assertion message crisp
        raise AssertionError("expected duplicate validator error")
