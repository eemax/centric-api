from __future__ import annotations

import json
import re
import sqlite3
import time
import unicodedata
import uuid
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .auth import AuthContext
from .config import ConfigError, runtime_path
from .load_config import LoadColumn, LoadConfig, LoadJob, LoadResolve, LoadScope
from .store import connect_readonly, endpoint_has_cache_evidence, table_exists

LOAD_RUNS_DIR = Path("load/runs")
LOAD_VALUE_SETS_DIR = Path("load/value-sets")
MAX_SAMPLES = 3
REVIEW_WORKBOOK_NAME = "review.xlsx"
REVIEW_COLUMN_HEADERS = (
    "_cent_load_run_id",
    "_cent_load_status",
    "_cent_load_status_code",
    "_cent_load_message",
    "_cent_load_request_path",
    "_cent_load_response_id",
    "_cent_load_processed_at",
)
RETRY_STATUSES = {"failed", "validation_error"}
REVIEW_STATUSES = {"success", "failed", "validation_error"}

LoadProgressCallback = Callable[[dict[str, Any]], None]


@dataclass(frozen=True)
class LoadIssue:
    row: int | None
    code: str
    message: str
    column: str | None = None
    sample: Any = None


@dataclass(frozen=True)
class LoadRequest:
    row: int
    method: str
    path: str
    body: Any


@dataclass(frozen=True)
class LoadMaterialized:
    job_name: str
    title: str
    workbook_path: Path
    sheet: str
    header_row: int
    rows_scanned: int
    valid_rows: int
    error_rows: int
    issues: tuple[LoadIssue, ...]
    requests: tuple[LoadRequest, ...]


@dataclass(frozen=True)
class LoadResponse:
    row: int
    status_code: int
    ok: bool
    body: Any


@dataclass(frozen=True)
class LoadRunResult:
    run_id: str
    job_name: str
    title: str
    mode: str
    dry_run: bool
    workbook_path: Path
    sheet: str
    rows_scanned: int
    valid_rows: int
    error_rows: int
    request_count: int
    success_count: int
    failure_count: int
    issues: tuple[LoadIssue, ...]
    requests: tuple[LoadRequest, ...]
    responses: tuple[LoadResponse, ...]
    run_dir: Path
    review_path: Path | None
    started_at: str
    finished_at: str


@dataclass(frozen=True)
class StyleBomRow:
    row: int
    values: dict[str, Any]


@dataclass(frozen=True)
class LoadValueSetIndex:
    name: str
    path: Path
    values: tuple[str, ...]
    exact: dict[str, str]
    normalized: dict[str, str]
    loose: dict[str, str]


def materialize_load(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None = None,
    limit: int | None = None,
    mode: str = "check",
    retry_statuses: set[str] | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadMaterialized:
    workbook_path = workbook_path.expanduser()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = _select_sheet(workbook, sheet)
        _emit_progress(
            progress_callback,
            {
                "event": "load_planning",
                "job": job.name,
                "mode": mode,
                "workbook": str(workbook_path),
                "sheet": worksheet.title,
            },
        )
        header_map, header_issues, header_stats = _map_headers(job, worksheet)
        retry_status_index = _retry_status_index(
            worksheet,
            job.input.header_row,
            retry_statuses=retry_statuses,
        )
        _emit_progress(
            progress_callback,
            {
                "event": "load_headers",
                "matched": header_stats["matched"],
                "columns": len(job.columns),
                "required_matched": header_stats["required_matched"],
                "required": header_stats["required"],
                "aliases": header_stats["aliases"],
                "issues": len(header_issues),
            },
        )
        reference_indexes = _build_reference_indexes(
            db_path,
            job,
            progress_callback=progress_callback,
        )
        value_set_indexes = _build_value_set_indexes(
            job,
            progress_callback=progress_callback,
        )
        requests: list[LoadRequest] = []
        issues: list[LoadIssue] = list(header_issues)
        rows_scanned = 0
        error_rows = 0
        if not header_issues:
            for row_number, row_values in _iter_data_rows(worksheet, job.input.header_row):
                if not _include_retry_row(row_values, retry_status_index, retry_statuses):
                    continue
                if limit is not None and rows_scanned >= limit:
                    break
                rows_scanned += 1
                values, row_issues = _row_values(
                    job,
                    row_number=row_number,
                    row_values=row_values,
                    header_map=header_map,
                    reference_indexes=reference_indexes,
                    value_set_indexes=value_set_indexes,
                )
                if row_issues:
                    error_rows += 1
                    issues.extend(row_issues)
                    continue
                path = _request_path(job, values, row_number=row_number)
                if isinstance(path, LoadIssue):
                    error_rows += 1
                    issues.append(path)
                    continue
                requests.append(
                    LoadRequest(
                        row=row_number,
                        method=job.method,
                        path=path,
                        body=_request_body(job, values),
                    )
                )
        _emit_progress(
            progress_callback,
            {
                "event": "load_validate",
                "scanned": rows_scanned,
                "valid": len(requests),
                "errors": error_rows,
            },
        )
        return LoadMaterialized(
            job_name=job.name,
            title=job.title,
            workbook_path=workbook_path,
            sheet=worksheet.title,
            header_row=job.input.header_row,
            rows_scanned=rows_scanned,
            valid_rows=len(requests),
            error_rows=error_rows,
            issues=tuple(issues),
            requests=tuple(requests),
        )
    finally:
        workbook.close()


def run_load(
    db_path: Path,
    config: LoadConfig,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    dry_run: bool,
    yes: bool,
    retry_statuses: set[str] | None = None,
    materialized: LoadMaterialized | None = None,
    auth_ctx: AuthContext | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadRunResult:
    mode = (
        "retry-dry-run"
        if retry_statuses and dry_run
        else ("retry" if retry_statuses else ("dry-run" if dry_run else "run"))
    )
    if materialized is None:
        materialized = materialize_load(
            db_path,
            job,
            workbook_path,
            sheet=sheet,
            limit=limit,
            mode=mode,
            retry_statuses=retry_statuses,
            progress_callback=progress_callback,
        )
    if not dry_run and not yes:
        raise ConfigError("Non-dry-run load requires --yes.")
    started_at = _utc_iso()
    run_id = _run_id(job.name)
    run_dir = runtime_path(LOAD_RUNS_DIR / run_id)
    run_dir.mkdir(parents=True, exist_ok=True)
    _write_requests(run_dir / "requests.jsonl", materialized.requests)
    _emit_progress(
        progress_callback,
        {
            "event": "load_artifacts",
            "run_dir": str(run_dir),
            "requests": len(materialized.requests),
        },
    )
    responses: tuple[LoadResponse, ...] = ()
    if not dry_run and materialized.requests:
        if auth_ctx is None:
            raise ConfigError("Load run requires an auth context.")
        responses = tuple(
            _execute_requests(
                auth_ctx,
                materialized.requests,
                progress_callback=progress_callback,
            )
        )
        _write_responses(run_dir / "responses.jsonl", responses)
    finished_at = _utc_iso()
    review_path = None
    if responses or _has_row_issues(materialized.issues):
        review_path = _write_review_workbook(
            materialized,
            responses=responses,
            run_id=run_id,
            processed_at=finished_at,
            output_path=run_dir / REVIEW_WORKBOOK_NAME,
        )
    result = LoadRunResult(
        run_id=run_id,
        job_name=job.name,
        title=job.title,
        mode=mode,
        dry_run=dry_run,
        workbook_path=materialized.workbook_path,
        sheet=materialized.sheet,
        rows_scanned=materialized.rows_scanned,
        valid_rows=materialized.valid_rows,
        error_rows=materialized.error_rows,
        request_count=len(materialized.requests),
        success_count=sum(1 for response in responses if response.ok),
        failure_count=sum(1 for response in responses if not response.ok),
        issues=materialized.issues,
        requests=materialized.requests,
        responses=responses,
        run_dir=run_dir,
        review_path=review_path,
        started_at=started_at,
        finished_at=finished_at,
    )
    _write_summary(run_dir / "summary.json", result, config)
    return result


def materialize_style_bom_workflow(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None = None,
    limit: int | None = None,
    mode: str = "check",
    retry_statuses: set[str] | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadMaterialized:
    parsed = _style_bom_rows(
        db_path,
        job,
        workbook_path,
        sheet=sheet,
        limit=limit,
        mode=mode,
        retry_statuses=retry_statuses,
        progress_callback=progress_callback,
    )
    requests = _style_bom_planned_requests(parsed["rows"])
    _emit_progress(
        progress_callback,
        {
            "event": "load_validate",
            "scanned": parsed["rows_scanned"],
            "valid": len(parsed["rows"]),
            "errors": parsed["error_rows"],
        },
    )
    return LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        issues=tuple(parsed["issues"]),
        requests=tuple(requests),
    )


def run_style_bom_workflow(
    db_path: Path,
    config: LoadConfig,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    dry_run: bool,
    yes: bool,
    retry_statuses: set[str] | None = None,
    materialized: LoadMaterialized | None = None,
    auth_ctx: AuthContext | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadRunResult:
    if not dry_run and not yes:
        raise ConfigError("Non-dry-run load requires --yes.")
    mode = (
        "retry-dry-run"
        if retry_statuses and dry_run
        else ("retry" if retry_statuses else ("dry-run" if dry_run else "run"))
    )
    parsed = _style_bom_rows(
        db_path,
        job,
        workbook_path,
        sheet=sheet,
        limit=limit,
        mode=mode,
        retry_statuses=retry_statuses,
        progress_callback=progress_callback,
    )
    started_at = _utc_iso()
    run_id = _run_id(job.name)
    run_dir = runtime_path(LOAD_RUNS_DIR / run_id)
    run_dir.mkdir(parents=True, exist_ok=True)

    materialized = materialized or LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        issues=tuple(parsed["issues"]),
        requests=tuple(_style_bom_planned_requests(parsed["rows"])),
    )
    requests = list(materialized.requests)
    responses: list[LoadResponse] = []
    issues = list(materialized.issues)
    workflow_issues: list[LoadIssue] = []
    if dry_run:
        _write_requests(run_dir / "requests.jsonl", tuple(requests))
    else:
        if parsed["rows"] and auth_ctx is None:
            raise ConfigError("Load run requires an auth context.")
        requests = []
        responses = []
        if parsed["rows"]:
            workflow_issues = _execute_style_bom_workflow(
                auth_ctx,
                parsed["rows"],
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            issues.extend(workflow_issues)
        _write_requests(run_dir / "requests.jsonl", tuple(requests))
        _write_responses(run_dir / "responses.jsonl", tuple(responses))
    _emit_progress(
        progress_callback,
        {
            "event": "load_artifacts",
            "run_dir": str(run_dir),
            "requests": len(requests),
        },
    )

    finished_at = _utc_iso()
    materialized = LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]) + len(workflow_issues if not dry_run else ()),
        issues=tuple(issues),
        requests=tuple(requests),
    )
    review_path = None
    if responses or _has_row_issues(materialized.issues):
        review_path = _write_review_workbook(
            materialized,
            responses=tuple(responses),
            run_id=run_id,
            processed_at=finished_at,
            output_path=run_dir / REVIEW_WORKBOOK_NAME,
        )
    result = LoadRunResult(
        run_id=run_id,
        job_name=job.name,
        title=job.title,
        mode=mode,
        dry_run=dry_run,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        request_count=len(requests),
        success_count=sum(1 for response in responses if response.ok),
        failure_count=sum(1 for response in responses if not response.ok),
        issues=tuple(issues),
        requests=tuple(requests),
        responses=tuple(responses),
        run_dir=run_dir,
        review_path=review_path,
        started_at=started_at,
        finished_at=finished_at,
    )
    _write_summary(run_dir / "summary.json", result, config)
    return result


def _style_bom_rows(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    mode: str,
    retry_statuses: set[str] | None,
    progress_callback: LoadProgressCallback | None,
) -> dict[str, Any]:
    _require_style_bom_columns(job)
    workbook_path = workbook_path.expanduser()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    section_names = _style_bom_section_names(db_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = _select_sheet(workbook, sheet)
        _emit_progress(
            progress_callback,
            {
                "event": "load_planning",
                "job": job.name,
                "mode": mode,
                "workbook": str(workbook_path),
                "sheet": worksheet.title,
            },
        )
        header_map, header_issues, header_stats = _map_headers(job, worksheet)
        retry_status_index = _retry_status_index(
            worksheet,
            job.input.header_row,
            retry_statuses=retry_statuses,
        )
        _emit_progress(
            progress_callback,
            {
                "event": "load_headers",
                "matched": header_stats["matched"],
                "columns": len(job.columns),
                "required_matched": header_stats["required_matched"],
                "required": header_stats["required"],
                "aliases": header_stats["aliases"],
                "issues": len(header_issues),
            },
        )
        reference_indexes = _build_reference_indexes(
            db_path,
            job,
            progress_callback=progress_callback,
        )
        value_set_indexes = _build_value_set_indexes(
            job,
            progress_callback=progress_callback,
        )
        rows: list[StyleBomRow] = []
        issues: list[LoadIssue] = list(header_issues)
        rows_scanned = 0
        error_rows = 0
        if not header_issues:
            for row_number, row_values in _iter_data_rows(worksheet, job.input.header_row):
                if not _include_retry_row(row_values, retry_status_index, retry_statuses):
                    continue
                if limit is not None and rows_scanned >= limit:
                    break
                rows_scanned += 1
                values, row_issues = _row_values(
                    job,
                    row_number=row_number,
                    row_values=row_values,
                    header_map=header_map,
                    reference_indexes=reference_indexes,
                    value_set_indexes=value_set_indexes,
                )
                section_name = values.get("section")
                if not _is_blank(section_name) and str(section_name) not in section_names:
                    row_issues.append(
                        LoadIssue(
                            row=row_number,
                            code="bom_section_not_found",
                            column="section",
                            message=(
                                f"Section {section_name!r} was not found exactly in "
                                "bom_sections.node_name."
                            ),
                        )
                    )
                if row_issues:
                    error_rows += 1
                    issues.extend(row_issues)
                    continue
                rows.append(StyleBomRow(row=row_number, values=values))
        return {
            "sheet": worksheet.title,
            "rows_scanned": rows_scanned,
            "error_rows": error_rows,
            "issues": issues,
            "rows": tuple(rows),
        }
    finally:
        workbook.close()


def _require_style_bom_columns(job: LoadJob) -> None:
    required = {
        "season",
        "style",
        "node_name",
        "description",
        "subtype",
        "section",
        "pm_id",
        "qty_default",
        "actual",
    }
    present = {column.key for column in job.columns}
    missing = sorted(required - present)
    if missing:
        raise ConfigError(
            f"load job[{job.name}] workflow style_bom is missing columns: "
            f"{', '.join(missing)}."
        )


def _style_bom_section_names(db_path: Path) -> set[str]:
    with connect_readonly(db_path) as conn:
        if not table_exists(conn, "endpoint_records"):
            raise ConfigError(
                "BOM load requires endpoint_records. Run fetch first."
            )
        if not endpoint_has_cache_evidence(conn, "bom_sections"):
            raise ConfigError(
                "BOM load requires cached endpoint records for: bom_sections. "
                "Run centric-api fetch --endpoint bom_sections first."
            )
        rows = conn.execute(
            """
            SELECT payload_json
            FROM endpoint_records
            WHERE endpoint = 'bom_sections'
            ORDER BY record_id
            """
        ).fetchall()
    names: set[str] = set()
    for row in rows:
        payload = _json_dict(row["payload_json"])
        if payload.get("active") is not True or payload.get("ad_hoc") is not False:
            continue
        value = payload.get("node_name")
        if not _is_blank(value):
            names.add(str(value))
    return names


def _style_bom_groups(rows: tuple[StyleBomRow, ...]) -> list[list[StyleBomRow]]:
    groups: dict[tuple[str, str, str, str], list[StyleBomRow]] = {}
    for row in rows:
        values = row.values
        key = (
            str(values["style"]),
            str(values["node_name"]),
            str(values.get("description") or ""),
            str(values["subtype"]),
        )
        groups.setdefault(key, []).append(row)
    return list(groups.values())


def _style_bom_planned_requests(rows: tuple[StyleBomRow, ...]) -> tuple[LoadRequest, ...]:
    requests: list[LoadRequest] = []
    for group in _style_bom_groups(rows):
        first = group[0]
        requests.append(_style_bom_header_request(first))
        for section_name in _style_bom_unique_sections(group):
            requests.append(
                LoadRequest(
                    row=_style_bom_first_section_row(group, section_name),
                    method="POST",
                    path=(
                        "/v2/apparel_bom_revisions/DRY-RUN-REVISION/"
                        "owned_sections/bom_section_definition"
                    ),
                    body={"node_name": section_name},
                )
            )
        for row in group:
            requests.append(
                _style_bom_line_request(
                    row,
                    revision_id="DRY-RUN-REVISION",
                    section_id=f"DRY-RUN-SECTION-{_slug(str(row.values['section']))}",
                )
            )
    return tuple(requests)


def _execute_style_bom_workflow(
    auth_ctx: AuthContext,
    rows: tuple[StyleBomRow, ...],
    *,
    requests: list[LoadRequest],
    responses: list[LoadResponse],
    progress_callback: LoadProgressCallback | None,
) -> list[LoadIssue]:
    issues: list[LoadIssue] = []
    groups = _style_bom_groups(rows)
    total = sum(1 + len(_style_bom_unique_sections(group)) + len(group) for group in groups)
    index = 0
    for group in groups:
        first = group[0]
        header_request = _style_bom_header_request(first)
        index += 1
        header_response = _execute_style_bom_request(
            auth_ctx,
            header_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        if not header_response.ok:
            issues.extend(
                _style_bom_group_issues(group, "bom_header_failed", "BOM header request failed.")
            )
            continue
        revision_id = _response_field(header_response, "latest_revision") or _response_field(
            header_response,
            "current_revision",
        )
        if _is_blank(revision_id):
            issues.extend(
                _style_bom_group_issues(
                    group,
                    "bom_revision_missing",
                    "BOM header response did not include latest_revision or current_revision.",
                )
            )
            continue

        section_ids: dict[str, str] = {}
        for section_name in _style_bom_unique_sections(group):
            section_request = LoadRequest(
                row=_style_bom_first_section_row(group, section_name),
                method="POST",
                path=(
                    f"/v2/apparel_bom_revisions/{revision_id}/"
                    "owned_sections/bom_section_definition"
                ),
                body={"node_name": section_name},
            )
            index += 1
            section_response = _execute_style_bom_request(
                auth_ctx,
                section_request,
                index=index,
                total=total,
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            section_id = _response_field(section_response, "id")
            if not section_response.ok or _is_blank(section_id):
                issues.extend(
                    LoadIssue(
                        row=row.row,
                        code="bom_section_create_failed",
                        column="section",
                        message=f"Could not create BOM section {section_name!r}.",
                    )
                    for row in group
                    if str(row.values["section"]) == section_name
                )
                continue
            section_ids[section_name] = str(section_id)

        for row in group:
            section_id = section_ids.get(str(row.values["section"]))
            if section_id is None:
                continue
            line_request = _style_bom_line_request(
                row,
                revision_id=str(revision_id),
                section_id=section_id,
            )
            index += 1
            line_response = _execute_style_bom_request(
                auth_ctx,
                line_request,
                index=index,
                total=total,
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            if not line_response.ok:
                issues.append(
                    LoadIssue(
                        row=row.row,
                        code="bom_line_create_failed",
                        message="BOM line request failed.",
                    )
                )
    return issues


def _style_bom_header_request(row: StyleBomRow) -> LoadRequest:
    values = row.values
    body = {
        "node_name": values["node_name"],
        "subtype": values["subtype"],
    }
    if not _is_blank(values.get("description")):
        body["description"] = values["description"]
    return LoadRequest(
        row=row.row,
        method="POST",
        path=f"/v2/styles/{values['style']}/data_sheets/apparel_boms",
        body=body,
    )


def _style_bom_line_request(
    row: StyleBomRow,
    *,
    revision_id: str,
    section_id: str,
) -> LoadRequest:
    values = row.values
    return LoadRequest(
        row=row.row,
        method="POST",
        path=f"/v2/apparel_bom_revisions/{revision_id}/items/part_materials",
        body={
            "ds_section": section_id,
            "pm_id": values["pm_id"],
            "qty_default": values["qty_default"],
            "actual": values["actual"],
        },
    )


def _execute_style_bom_request(
    auth_ctx: AuthContext,
    request: LoadRequest,
    *,
    index: int,
    total: int,
    requests: list[LoadRequest],
    responses: list[LoadResponse],
    progress_callback: LoadProgressCallback | None,
) -> LoadResponse:
    requests.append(request)
    started = time.time()
    try:
        response = auth_ctx.request(
            request.method,
            _request_url(auth_ctx, request.path),
            json_body=request.body,
        )
        status_code = response.status_code
        body = _response_body(response)
    except Exception as exc:
        status_code = 0
        body = _exception_body(exc)
    _emit_progress(
        progress_callback,
        {
            "event": "load_send",
            "index": index,
            "total": total,
            "row": request.row,
            "method": request.method,
            "path": request.path,
            "status_code": status_code,
            "elapsed_seconds": time.time() - started,
        },
    )
    load_response = LoadResponse(
        row=request.row,
        status_code=status_code,
        ok=0 < status_code < 400,
        body=body,
    )
    responses.append(load_response)
    return load_response


def _style_bom_unique_sections(group: list[StyleBomRow]) -> list[str]:
    seen: set[str] = set()
    sections: list[str] = []
    for row in group:
        section = str(row.values["section"])
        if section in seen:
            continue
        seen.add(section)
        sections.append(section)
    return sections


def _style_bom_first_section_row(group: list[StyleBomRow], section_name: str) -> int:
    for row in group:
        if str(row.values["section"]) == section_name:
            return row.row
    return group[0].row


def _style_bom_group_issues(
    group: list[StyleBomRow],
    code: str,
    message: str,
) -> list[LoadIssue]:
    return [LoadIssue(row=row.row, code=code, message=message) for row in group]


def _response_field(response: LoadResponse, key: str) -> str | None:
    if not isinstance(response.body, dict):
        return None
    value = response.body.get(key)
    return None if _is_blank(value) else str(value)


def _slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-")
    return slug or "section"


def _has_row_issues(issues: tuple[LoadIssue, ...]) -> bool:
    return any(issue.row is not None for issue in issues)


def _map_headers(
    job: LoadJob,
    worksheet: Any,
) -> tuple[dict[str, int], list[LoadIssue], dict[str, int]]:
    header_cells = next(
        worksheet.iter_rows(
            min_row=job.input.header_row,
            max_row=job.input.header_row,
            values_only=True,
        ),
        (),
    )
    actual_headers: dict[str, list[tuple[int, str]]] = {}
    for index, value in enumerate(header_cells):
        if value is None or str(value).strip() == "":
            continue
        text = str(value).strip()
        actual_headers.setdefault(_lookup_key(text), []).append((index, text))
    mapping: dict[str, int] = {}
    aliases = 0
    issues: list[LoadIssue] = []
    for column in job.columns:
        matches: list[tuple[int, str]] = []
        for accepted in column.accepted_headers:
            matches.extend(actual_headers.get(_lookup_key(accepted), []))
        unique_matches = sorted(set(matches))
        if len(unique_matches) > 1:
            issues.append(
                LoadIssue(
                    row=None,
                    code="ambiguous_header",
                    column=column.key,
                    message=(
                        f"Column {column.key!r} matched multiple headers: "
                        + ", ".join(header for _index, header in unique_matches)
                    ),
                )
            )
        elif not unique_matches and column.required:
            issues.append(
                LoadIssue(
                    row=None,
                    code="missing_required_header",
                    column=column.key,
                    message=f"Missing required header: {column.header}",
                )
            )
        elif unique_matches:
            index, header = unique_matches[0]
            mapping[column.key] = index
            if _lookup_key(header) != _lookup_key(column.header):
                aliases += 1
    stats = {
        "matched": len(mapping),
        "required_matched": sum(
            1 for column in job.columns if column.required and column.key in mapping
        ),
        "required": sum(1 for column in job.columns if column.required),
        "aliases": aliases,
    }
    return mapping, issues, stats


def _retry_status_index(
    worksheet: Any,
    header_row: int,
    *,
    retry_statuses: set[str] | None,
) -> int | None:
    if retry_statuses is None:
        return None
    header_cells = next(
        worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
        (),
    )
    for index, value in enumerate(header_cells):
        if _lookup_key(str(value or "")) == _lookup_key("_cent_load_status"):
            return index
    raise ConfigError("Retry workbook is missing _cent_load_status.")


def _include_retry_row(
    row_values: tuple[Any, ...],
    retry_status_index: int | None,
    retry_statuses: set[str] | None,
) -> bool:
    if retry_statuses is None:
        return True
    value = _cell_value(row_values, retry_status_index)
    return _lookup_key(str(value or "")) in retry_statuses


def _row_values(
    job: LoadJob,
    *,
    row_number: int,
    row_values: tuple[Any, ...],
    header_map: dict[str, int],
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
    value_set_indexes: dict[str, LoadValueSetIndex],
) -> tuple[dict[str, Any], list[LoadIssue]]:
    values: dict[str, Any] = {}
    issues: list[LoadIssue] = []
    for column in job.columns:
        raw_value = _cell_value(row_values, header_map.get(column.key))
        if _is_blank(raw_value):
            if column.required:
                issues.append(
                    LoadIssue(
                        row=row_number,
                        code="missing_required_value",
                        column=column.key,
                        message=f"Missing required value for {column.header}.",
                    )
                )
            values[column.key] = None
            continue
        parsed = _parse_value(column, raw_value, row_number)
        if isinstance(parsed, LoadIssue):
            issues.append(parsed)
            continue
        values[column.key] = parsed

    for column in job.columns:
        parsed = values.get(column.key)
        if _is_blank(parsed):
            continue
        if column.type == "ref":
            resolved = _resolve_value(
                column,
                parsed,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.type == "ref_or_id":
            resolved = _resolve_ref_or_id(
                column,
                parsed,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.type == "scoped_ref":
            resolved = _resolve_scoped_ref(
                column,
                parsed,
                values=values,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.type == "composition_list":
            resolved = _resolve_composition_list(
                column,
                parsed,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.value_set is not None:
            resolved = _resolve_value_set(
                column,
                parsed,
                row_number=row_number,
                value_set_indexes=value_set_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
    return values, issues


def _parse_value(column: LoadColumn, raw_value: Any, row_number: int) -> Any | LoadIssue:
    if column.type in {"text", "ref", "ref_or_id", "scoped_ref"}:
        return str(raw_value).strip()
    if column.type == "number":
        try:
            value = Decimal(str(raw_value).strip())
        except (InvalidOperation, ValueError, AttributeError):
            return LoadIssue(
                row=row_number,
                code="invalid_number",
                column=column.key,
                message=f"Value for {column.header} must be numeric.",
                sample=raw_value,
            )
        return int(value) if value == value.to_integral_value() else float(value)
    if column.type == "boolean":
        if isinstance(raw_value, bool):
            return raw_value
        text = str(raw_value).strip().casefold()
        if text in {"true", "yes", "y", "1"}:
            return True
        if text in {"false", "no", "n", "0"}:
            return False
        return LoadIssue(
            row=row_number,
            code="invalid_boolean",
            column=column.key,
            message=f"Value for {column.header} must be boolean.",
            sample=raw_value,
        )
    if column.type == "composition_list":
        return _parse_composition_entries(column, raw_value, row_number)
    return raw_value


def _parse_composition_entries(
    column: LoadColumn,
    raw_value: Any,
    row_number: int,
) -> list[tuple[Decimal, str]] | LoadIssue:
    text = str(raw_value).strip()
    entries = _composition_entries_from_text(text)
    if isinstance(entries, LoadIssue):
        return LoadIssue(
            row=row_number,
            code=entries.code,
            column=column.key,
            message=entries.message,
            sample=raw_value,
        )
    total = sum((percentage for percentage, _name in entries), Decimal("0"))
    if abs(total - Decimal("100")) > Decimal("0.0001"):
        return LoadIssue(
            row=row_number,
            code="composition_total_invalid",
            column=column.key,
            message=f"Composition total must be 100; got {_decimal_label(total)}.",
            sample=raw_value,
        )
    return entries


def _composition_entries_from_text(text: str) -> list[tuple[Decimal, str]] | LoadIssue:
    cleaned = text.strip().strip(".")
    if not cleaned:
        return LoadIssue(row=None, code="empty_composition", message="Composition is blank.")
    segments = [segment.strip() for segment in re.split(r"[,;+\n]+", cleaned) if segment.strip()]
    if len(segments) > 1:
        return _composition_entries_from_segments(segments)
    entries = _composition_entries_from_numeric_tokens(cleaned)
    if entries is not None:
        return entries
    return LoadIssue(
        row=None,
        code="composition_percentage_missing",
        message=f"Composition entry is missing a percentage: {cleaned!r}.",
    )


def _composition_entries_from_segments(
    segments: list[str],
) -> list[tuple[Decimal, str]] | LoadIssue:
    entries: list[tuple[Decimal, str]] = []
    for segment in segments:
        segment_entries = _composition_entries_from_numeric_tokens(segment)
        if isinstance(segment_entries, LoadIssue):
            return segment_entries
        if segment_entries is None:
            return LoadIssue(
                row=None,
                code="composition_percentage_missing",
                message=f"Composition entry is missing a percentage: {segment!r}.",
            )
        entries.extend(segment_entries)
    return entries


def _composition_entries_from_numeric_tokens(
    text: str,
) -> list[tuple[Decimal, str]] | LoadIssue | None:
    numbers = list(re.finditer(r"\d+(?:\.\d+)?\s*%?", text))
    if not numbers:
        return None

    percentages: list[Decimal] = []
    for number in numbers:
        percentage = _decimal_or_none(number.group().strip().rstrip("%"))
        if percentage is None:
            return LoadIssue(
                row=None,
                code="invalid_composition_percentage",
                message=f"Composition percentage must be numeric: {number.group()!r}.",
            )
        if percentage <= 0:
            return LoadIssue(
                row=None,
                code="invalid_composition_percentage",
                message=f"Composition percentage must be greater than 0: {number.group()!r}.",
            )
        percentages.append(percentage)

    parts = [text[: numbers[0].start()]]
    parts.extend(
        text[current.end() : next_number.start()]
        for current, next_number in zip(numbers, numbers[1:], strict=False)
    )
    parts.append(text[numbers[-1].end() :])
    names = [_clean_composition_name(part) for part in parts]
    if not any(names):
        return LoadIssue(
            row=None,
            code="composition_name_missing",
            message=f"Composition entry is missing a name: {text!r}.",
        )

    def assign(
        index: int,
        used_name_indexes: frozenset[int],
        entries: tuple[tuple[Decimal, str], ...],
    ) -> tuple[tuple[Decimal, str], ...] | None:
        if index == len(percentages):
            if all(
                not name or part_index in used_name_indexes for part_index, name in enumerate(names)
            ):
                return entries
            return None

        options: list[int] = []
        before_index = index
        after_index = index + 1
        if names[before_index] and before_index not in used_name_indexes:
            options.append(before_index)
        if names[after_index] and after_index not in used_name_indexes:
            options.append(after_index)

        for name_index in options:
            result = assign(
                index + 1,
                used_name_indexes | frozenset({name_index}),
                entries + ((percentages[index], names[name_index]),),
            )
            if result is not None:
                return result
        return None

    assigned = assign(0, frozenset(), ())
    if assigned is None:
        return LoadIssue(
            row=None,
            code="composition_name_missing",
            message=f"Composition entry is missing a name: {text!r}.",
        )
    return list(assigned)


def _clean_composition_name(value: str) -> str:
    cleaned = re.sub(r"^[\s,;/+._%-]+|[\s,;/+._%-]+$", "", value.strip())
    return re.sub(r"\s+", " ", cleaned).strip()


def _decimal_or_none(value: str) -> Decimal | None:
    try:
        return Decimal(value.strip())
    except (InvalidOperation, ValueError, AttributeError):
        return None


def _decimal_label(value: Decimal) -> str:
    normalized = value.normalize()
    return str(int(normalized)) if normalized == normalized.to_integral_value() else str(normalized)


def _resolve_value(
    column: LoadColumn,
    value: Any,
    *,
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> str | LoadIssue:
    resolve = column.resolve
    if resolve is None:
        raise ConfigError(f"Column {column.key} is missing resolve config.")
    matches = reference_indexes.get(_resolve_key(resolve), {}).get(_lookup_key(str(value)), [])
    if not matches:
        return LoadIssue(
            row=row_number,
            code="ref_not_found",
            column=column.key,
            message=(
                f"{column.header} {value!r} was not found in {resolve.endpoint}.{resolve.match}."
            ),
        )
    if len(matches) > 1:
        return LoadIssue(
            row=row_number,
            code="ref_ambiguous",
            column=column.key,
            message=(
                f"{column.header} {value!r} matched {len(matches)} records in "
                f"{resolve.endpoint}.{resolve.match}."
            ),
            sample=[match.get("id") for match in matches[:MAX_SAMPLES]],
        )
    resolved = matches[0].get(resolve.output)
    if _is_blank(resolved):
        return LoadIssue(
            row=row_number,
            code="ref_output_blank",
            column=column.key,
            message=f"Resolved {resolve.endpoint} record has blank {resolve.output!r}.",
            sample=matches[0].get("id"),
        )
    return str(resolved).strip()


def _resolve_ref_or_id(
    column: LoadColumn,
    value: Any,
    *,
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> str | LoadIssue:
    resolve = column.resolve
    if resolve is None:
        raise ConfigError(f"Column {column.key} is missing resolve config.")
    text = str(value).strip()
    direct_matches = reference_indexes.get(_resolve_direct_key(resolve), {}).get(
        _lookup_key(text),
        [],
    )
    if len(direct_matches) == 1:
        resolved = direct_matches[0].get(resolve.output)
        return str(resolved).strip()
    if len(direct_matches) > 1:
        return LoadIssue(
            row=row_number,
            code="ref_id_ambiguous",
            column=column.key,
            message=(
                f"{column.header} {text!r} matched {len(direct_matches)} records in "
                f"{resolve.endpoint}.{resolve.output}."
            ),
            sample=[match.get("id") for match in direct_matches[:MAX_SAMPLES]],
        )
    resolved = _resolve_value(
        column,
        value,
        row_number=row_number,
        reference_indexes=reference_indexes,
    )
    if isinstance(resolved, LoadIssue) and resolved.code == "ref_not_found":
        return LoadIssue(
            row=row_number,
            code="ref_or_id_not_found",
            column=column.key,
            message=(
                f"{column.header} {text!r} was not found in {resolve.endpoint} by "
                f"{resolve.output} or {resolve.match}."
            ),
        )
    return resolved


def _resolve_scoped_ref(
    column: LoadColumn,
    value: Any,
    *,
    values: dict[str, Any],
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> str | LoadIssue:
    resolve = column.resolve
    if resolve is None or resolve.scope is None:
        raise ConfigError(f"Column {column.key} is missing scoped resolve config.")
    scope = resolve.scope
    scope_value = values.get(scope.column)
    if _is_blank(scope_value):
        return LoadIssue(
            row=row_number,
            code="scope_value_missing",
            column=column.key,
            message=f"{column.header} requires a value for scope column {scope.column!r}.",
        )

    candidates = reference_indexes.get(_resolve_key(resolve), {}).get(_lookup_key(str(value)), [])
    if not candidates:
        return LoadIssue(
            row=row_number,
            code="scoped_ref_not_found",
            column=column.key,
            message=(
                f"{column.header} {value!r} was not found in "
                f"{resolve.endpoint}.{resolve.match}."
            ),
        )

    scope_index = reference_indexes.get(_scope_index_key(scope), {})
    scoped_matches: list[dict[str, Any]] = []
    missing_scope_refs: list[str] = []
    for candidate in candidates:
        scope_ref = _extract_path(candidate, scope.via)
        if _is_blank(scope_ref):
            continue
        scope_payloads = scope_index.get(_lookup_key(str(scope_ref)), [])
        if not scope_payloads:
            missing_scope_refs.append(str(scope_ref).strip())
            continue
        scope_lookup = _lookup_key(str(scope_value))
        if any(
            _lookup_key(str(_extract_path(scope_payload, scope.match))) == scope_lookup
            for scope_payload in scope_payloads
        ):
            scoped_matches.append(candidate)

    if not scoped_matches and missing_scope_refs:
        return LoadIssue(
            row=row_number,
            code="scope_ref_missing",
            column=column.key,
            message=(
                f"{column.header} {value!r} matched {resolve.endpoint} records, but "
                f"referenced {scope.endpoint} records were missing for "
                f"{resolve.endpoint}.{scope.via}."
            ),
            sample=sorted(set(missing_scope_refs))[:MAX_SAMPLES],
        )
    if not scoped_matches:
        return LoadIssue(
            row=row_number,
            code="scoped_ref_not_in_scope",
            column=column.key,
            message=(
                f"{column.header} {value!r} was not found under {scope.column} "
                f"{scope_value!r} via {resolve.endpoint}.{scope.via} -> "
                f"{scope.endpoint}.{scope.match}."
            ),
        )
    if len(scoped_matches) > 1:
        return LoadIssue(
            row=row_number,
            code="scoped_ref_ambiguous",
            column=column.key,
            message=(
                f"{column.header} {value!r} under {scope.column} {scope_value!r} "
                f"matched {len(scoped_matches)} records."
            ),
            sample=[match.get("id") for match in scoped_matches[:MAX_SAMPLES]],
        )

    resolved = scoped_matches[0].get(resolve.output)
    if _is_blank(resolved):
        return LoadIssue(
            row=row_number,
            code="scoped_ref_output_blank",
            column=column.key,
            message=f"Resolved {resolve.endpoint} record has blank {resolve.output!r}.",
            sample=scoped_matches[0].get("id"),
        )
    return str(resolved).strip()


def _resolve_composition_list(
    column: LoadColumn,
    entries: list[tuple[Decimal, str]],
    *,
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> list[dict[str, Any]] | LoadIssue:
    resolve = column.resolve
    if resolve is None:
        raise ConfigError(f"Column {column.key} is missing resolve config.")
    resolved_entries: dict[str, dict[str, Any]] = {}
    for percentage, name in entries:
        matches = _composition_reference_matches(
            reference_indexes.get(_resolve_key(resolve), {}),
            resolve,
            name,
        )
        if not matches:
            return LoadIssue(
                row=row_number,
                code="composition_not_found",
                column=column.key,
                message=(
                    f"Composition {name!r} was not found in {resolve.endpoint}.{resolve.match}."
                ),
            )
        if len(matches) > 1:
            return LoadIssue(
                row=row_number,
                code="composition_ambiguous",
                column=column.key,
                message=(
                    f"Composition {name!r} matched {len(matches)} records in "
                    f"{resolve.endpoint}.{resolve.match}."
                ),
                sample=[match.get("id") for match in matches[:MAX_SAMPLES]],
            )
        resolved = matches[0].get(resolve.output)
        if _is_blank(resolved):
            return LoadIssue(
                row=row_number,
                code="composition_output_blank",
                column=column.key,
                message=f"Resolved {resolve.endpoint} record has blank {resolve.output!r}.",
                sample=matches[0].get("id"),
            )
        resolved_id = str(resolved).strip()
        existing = resolved_entries.get(resolved_id)
        if existing is None:
            resolved_entries[resolved_id] = {
                "percentage": _number_value(percentage),
                "composition": resolved_id,
            }
        else:
            existing["percentage"] = _number_value(
                Decimal(str(existing["percentage"])) + percentage
            )
    return list(resolved_entries.values())


def _composition_reference_matches(
    reference_index: dict[str, list[dict[str, Any]]],
    resolve: LoadResolve,
    name: str,
) -> list[dict[str, Any]]:
    matches = reference_index.get(_lookup_key(name), [])
    if matches:
        return matches

    canonical_name = _composition_lookup_key(name)
    if not canonical_name:
        return []

    canonical_matches: list[dict[str, Any]] = []
    seen: set[str] = set()
    for payloads in reference_index.values():
        for payload in payloads:
            value = _extract_path(payload, resolve.match)
            if _is_blank(value) or _composition_lookup_key(str(value)) != canonical_name:
                continue
            dedupe_key = str(payload.get(resolve.output) or payload.get("id") or id(payload))
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            canonical_matches.append(payload)
    return canonical_matches


def _composition_lookup_key(value: str) -> str:
    tokens = re.findall(r"[a-z0-9]+", value.casefold())
    return " ".join(sorted(tokens))


def _number_value(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def _resolve_value_set(
    column: LoadColumn,
    value: Any,
    *,
    row_number: int,
    value_set_indexes: dict[str, LoadValueSetIndex],
) -> str | LoadIssue:
    value_set = column.value_set
    if value_set is None:
        raise ConfigError(f"Column {column.key} is missing value_set config.")
    index = value_set_indexes[value_set.name]
    text = str(value).strip()
    resolved = (
        index.exact.get(text)
        or index.normalized.get(_normalize_value_set_key(text))
        or index.loose.get(_loose_value_set_key(text))
    )
    if resolved is None:
        return LoadIssue(
            row=row_number,
            code="value_set_not_found",
            column=column.key,
            message=f"{column.header} {text!r} was not found in value set {value_set.name}.",
            sample=index.values[:MAX_SAMPLES],
        )
    return resolved


def _build_value_set_indexes(
    job: LoadJob,
    *,
    progress_callback: LoadProgressCallback | None = None,
) -> dict[str, LoadValueSetIndex]:
    value_set_names = {
        column.value_set.name for column in job.columns if column.value_set is not None
    }
    indexes: dict[str, LoadValueSetIndex] = {}
    for name in sorted(value_set_names):
        path = runtime_path(LOAD_VALUE_SETS_DIR / f"{name}.xlsx")
        indexes[name] = _load_value_set_index(name, path)
        _emit_progress(
            progress_callback,
            {
                "event": "load_values",
                "name": name,
                "path": str(path),
                "values": len(indexes[name].values),
            },
        )
    return indexes


def _load_value_set_index(name: str, path: Path) -> LoadValueSetIndex:
    if not path.is_file():
        raise ConfigError(f"Load value set {name!r} not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        values: list[str] = []
        seen_exact: set[str] = set()
        for (raw_value,) in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
            if _is_blank(raw_value):
                continue
            value = str(raw_value).strip()
            if value in seen_exact:
                continue
            seen_exact.add(value)
            values.append(value)
    finally:
        workbook.close()

    if not values:
        raise ConfigError(f"Load value set {name!r} has no values: {path}")
    exact = {value: value for value in values}
    normalized = _value_set_lookup(
        name,
        path,
        values,
        key_func=_normalize_value_set_key,
        label="normalized",
    )
    loose = _value_set_lookup(
        name,
        path,
        values,
        key_func=_loose_value_set_key,
        label="loose",
    )
    return LoadValueSetIndex(
        name=name,
        path=path,
        values=tuple(values),
        exact=exact,
        normalized=normalized,
        loose=loose,
    )


def _value_set_lookup(
    name: str,
    path: Path,
    values: list[str],
    *,
    key_func: Callable[[str], str],
    label: str,
) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for value in values:
        key = key_func(value)
        if not key:
            continue
        existing = lookup.get(key)
        if existing is not None and existing != value:
            raise ConfigError(
                f"Load value set {name!r} has ambiguous {label} values "
                f"{existing!r} and {value!r}: {path}"
            )
        lookup[key] = value
    return lookup


def _normalize_value_set_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).replace("\xa0", " ")
    return " ".join(normalized.strip().casefold().split())


def _loose_value_set_key(value: str) -> str:
    normalized = _normalize_value_set_key(value)
    tokens = re.findall(r"[a-z0-9]+", normalized)
    return "".join(_singular_value_set_token(token) for token in tokens)


def _singular_value_set_token(token: str) -> str:
    if len(token) > 3 and token.endswith("ies"):
        return f"{token[:-3]}y"
    if len(token) > 4 and token.endswith(("ches", "shes")):
        return token[:-2]
    if len(token) > 3 and token.endswith(("ses", "xes", "zes")):
        return token[:-2]
    if len(token) > 3 and token.endswith("s") and not token.endswith("ss"):
        return token[:-1]
    return token


def _build_reference_indexes(
    db_path: Path,
    job: LoadJob,
    *,
    progress_callback: LoadProgressCallback | None = None,
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    ref_columns = [column for column in job.columns if column.resolve is not None]
    refs = [column.resolve for column in ref_columns]
    if not refs:
        return {}
    with connect_readonly(db_path) as conn:
        if not table_exists(conn, "endpoint_records"):
            raise ConfigError(
                "Load reference resolution requires endpoint_records. Run fetch first."
            )
        missing_endpoints = sorted(
            {
                endpoint
                for resolve in refs
                for endpoint in _resolve_required_endpoints(resolve)
                if not endpoint_has_cache_evidence(conn, endpoint)
            }
        )
        if missing_endpoints:
            raise ConfigError(
                "Load reference resolution requires cached endpoint records for: "
                f"{', '.join(missing_endpoints)}. Run centric-api fetch for those endpoints first."
            )
        indexes: dict[str, dict[str, list[dict[str, Any]]]] = {}
        for resolve in refs:
            key = _resolve_key(resolve)
            if key not in indexes:
                indexes[key] = _reference_index(conn, resolve)
                _emit_progress(
                    progress_callback,
                    {
                        "event": "load_refs",
                        "endpoint": resolve.endpoint,
                        "match": resolve.match,
                        "output": resolve.output,
                        "filters": resolve.filters or {},
                        "matched": sum(len(matches) for matches in indexes[key].values()),
                        "values": len(indexes[key]),
                    },
                )
            if resolve.scope is not None:
                scope_key = _scope_index_key(resolve.scope)
                if scope_key not in indexes:
                    indexes[scope_key] = _reference_index(
                        conn,
                        LoadResolve(
                            endpoint=resolve.scope.endpoint,
                            match=resolve.scope.output,
                            output=resolve.scope.output,
                        ),
                    )
        for column in ref_columns:
            if column.type != "ref_or_id" or column.resolve is None:
                continue
            direct_key = _resolve_direct_key(column.resolve)
            if direct_key not in indexes:
                indexes[direct_key] = _reference_index(
                    conn,
                    column.resolve,
                    match_path=column.resolve.output,
                )
        return indexes


def _resolve_required_endpoints(resolve: LoadResolve) -> tuple[str, ...]:
    if resolve.scope is None:
        return (resolve.endpoint,)
    return (resolve.endpoint, resolve.scope.endpoint)


def _reference_index(
    conn: sqlite3.Connection,
    resolve: LoadResolve,
    *,
    match_path: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    rows = conn.execute(
        """
        SELECT payload_json
        FROM endpoint_records
        WHERE endpoint = ?
        ORDER BY record_id
        """,
        [resolve.endpoint],
    ).fetchall()
    index: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        payload = _json_dict(row["payload_json"])
        if not _matches_resolve_filters(payload, resolve):
            continue
        value = _extract_path(payload, match_path or resolve.match)
        if _is_blank(value):
            continue
        index.setdefault(_lookup_key(str(value)), []).append(payload)
    return index


def _matches_resolve_filters(payload: dict[str, Any], resolve: LoadResolve) -> bool:
    if not resolve.filters:
        return True
    return all(
        _extract_path(payload, path) == expected for path, expected in resolve.filters.items()
    )


def _request_path(job: LoadJob, values: dict[str, Any], *, row_number: int) -> str | LoadIssue:
    def replace(match: re.Match[str]) -> str:
        key = match.group(1)
        value = values.get(key)
        return "" if _is_blank(value) else str(value).strip()

    path = re.sub(r"{([A-Za-z_][A-Za-z0-9_]*)}", replace, job.path)
    if "{}" in path or re.search(r"{[A-Za-z_][A-Za-z0-9_]*}", path):
        return LoadIssue(
            row=row_number,
            code="path_template_unresolved",
            message=f"Request path could not be resolved from template {job.path!r}.",
        )
    if "//" in path:
        return LoadIssue(
            row=row_number,
            code="path_template_blank",
            message=f"Request path has a blank template value: {job.path!r}.",
        )
    return path


def _request_body(job: LoadJob, values: dict[str, Any]) -> Any:
    if isinstance(job.body, str):
        return values.get(job.body)
    body: dict[str, Any] = {}
    for target, source in job.body.items():
        value = values.get(source)
        if _is_blank(value):
            continue
        body[target] = value
    return body


def _execute_requests(
    auth_ctx: AuthContext,
    requests: tuple[LoadRequest, ...],
    *,
    progress_callback: LoadProgressCallback | None = None,
) -> list[LoadResponse]:
    responses: list[LoadResponse] = []
    started = time.time()
    total = len(requests)
    for index, request in enumerate(requests, start=1):
        try:
            response = auth_ctx.request(
                request.method,
                _request_url(auth_ctx, request.path),
                json_body=request.body,
            )
            status_code = response.status_code
            body = _response_body(response)
        except Exception as exc:
            status_code = 0
            body = _exception_body(exc)
        _emit_progress(
            progress_callback,
            {
                "event": "load_send",
                "index": index,
                "total": total,
                "row": request.row,
                "method": request.method,
                "path": request.path,
                "status_code": status_code,
                "elapsed_seconds": time.time() - started,
            },
        )
        responses.append(
            LoadResponse(
                row=request.row,
                status_code=status_code,
                ok=0 < status_code < 400,
                body=body,
            )
        )
    return responses


def _write_review_workbook(
    result: LoadMaterialized,
    *,
    responses: tuple[LoadResponse, ...],
    run_id: str,
    processed_at: str,
    output_path: Path,
) -> Path:
    workbook = load_workbook(result.workbook_path)
    try:
        worksheet = _select_sheet(workbook, result.sheet)
        columns = _review_column_indexes(worksheet, result.header_row)
        _clear_review_rows(worksheet, columns, result.header_row)
        requests_by_row = {request.row: request for request in result.requests}
        issues_by_row: dict[int, list[LoadIssue]] = {}
        for issue in result.issues:
            if issue.row is not None:
                issues_by_row.setdefault(issue.row, []).append(issue)
        for row_number, issues in issues_by_row.items():
            _write_review_row(
                worksheet,
                row_number,
                columns,
                run_id=run_id,
                status="validation_error",
                status_code=None,
                message="; ".join(issue.message for issue in issues),
                request_path=requests_by_row.get(row_number).path
                if row_number in requests_by_row
                else "",
                response_id="",
                processed_at=processed_at,
            )
        for response in responses:
            request = requests_by_row.get(response.row)
            status = "success" if response.ok else "failed"
            _write_review_row(
                worksheet,
                response.row,
                columns,
                run_id=run_id,
                status=status,
                status_code=response.status_code,
                message=_response_message(response),
                request_path=request.path if request else "",
                response_id=_response_id(response),
                processed_at=processed_at,
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = _temp_output_path(output_path)
        try:
            workbook.save(temp_path)
            temp_path.replace(output_path)
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise
        return output_path
    finally:
        workbook.close()


def _review_column_indexes(worksheet: Any, header_row: int) -> dict[str, int]:
    existing: dict[str, int] = {}
    for cell in worksheet[header_row]:
        value = str(cell.value or "").strip()
        if value:
            existing[_lookup_key(value)] = int(cell.column)
    next_column = int(worksheet.max_column) + 1
    indexes: dict[str, int] = {}
    for header in REVIEW_COLUMN_HEADERS:
        key = _lookup_key(header)
        column = existing.get(key)
        if column is None:
            column = next_column
            worksheet.cell(row=header_row, column=column, value=header)
            next_column += 1
        indexes[header] = column
    return indexes


def _clear_review_rows(worksheet: Any, columns: dict[str, int], header_row: int) -> None:
    for row_number in range(header_row + 1, int(worksheet.max_row) + 1):
        for column in columns.values():
            worksheet.cell(row=row_number, column=column).value = None


def _write_review_row(
    worksheet: Any,
    row_number: int,
    columns: dict[str, int],
    *,
    run_id: str,
    status: str,
    status_code: int | None,
    message: str,
    request_path: str,
    response_id: str,
    processed_at: str,
) -> None:
    if status not in REVIEW_STATUSES:
        raise ConfigError(f"Unknown load review status: {status}")
    values = {
        "_cent_load_run_id": run_id,
        "_cent_load_status": status,
        "_cent_load_status_code": status_code,
        "_cent_load_message": _safe_review_text(message),
        "_cent_load_request_path": _safe_review_text(request_path),
        "_cent_load_response_id": _safe_review_text(response_id),
        "_cent_load_processed_at": processed_at,
    }
    for header, value in values.items():
        worksheet.cell(row=row_number, column=columns[header], value=value)


def _response_message(response: LoadResponse) -> str:
    if response.ok:
        return "ok"
    body = response.body
    if isinstance(body, dict):
        for key in ("error", "message", "detail"):
            value = body.get(key)
            if not _is_blank(value):
                return str(value)
    if _is_blank(body):
        return f"HTTP {response.status_code}"
    return str(body)


def _response_id(response: LoadResponse) -> str:
    if not isinstance(response.body, dict):
        return ""
    value = response.body.get("id")
    return "" if _is_blank(value) else str(value)


def _emit_progress(
    progress_callback: LoadProgressCallback | None,
    event: dict[str, Any],
) -> None:
    if progress_callback is not None:
        progress_callback(event)


def _select_sheet(workbook: Any, sheet: str | None) -> Any:
    if sheet is None:
        return workbook.worksheets[0]
    if sheet not in workbook.sheetnames:
        names = ", ".join(workbook.sheetnames)
        raise ConfigError(f"Workbook sheet {sheet!r} not found. Available: {names}")
    return workbook[sheet]


def _iter_data_rows(worksheet: Any, header_row: int) -> Iterator[tuple[int, tuple[Any, ...]]]:
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        row_values = tuple(values)
        if all(_is_blank(value) for value in row_values):
            continue
        yield row_number, row_values


def _cell_value(row_values: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row_values):
        return None
    return row_values[index]


def _request_url(auth_ctx: AuthContext, path: str) -> str:
    normalized = path.strip().strip("/")
    return f"{auth_ctx.base_url}/api/{normalized}"


def _response_body(response: Any) -> Any:
    if not response.text:
        return None
    try:
        return response.json()
    except ValueError:
        return response.text


def _exception_body(exc: Exception) -> dict[str, str]:
    return {
        "error": str(exc),
        "type": exc.__class__.__name__,
    }


def _write_requests(path: Path, requests: tuple[LoadRequest, ...]) -> None:
    _write_jsonl(path, [_request_record(request) for request in requests])


def _write_responses(path: Path, responses: tuple[LoadResponse, ...]) -> None:
    _write_jsonl(path, [_response_record(response) for response in responses])


def _write_summary(path: Path, result: LoadRunResult, config: LoadConfig) -> None:
    payload = run_record(result)
    payload["config"] = [str(path) for path in config.paths]
    _write_text_atomic(path, json.dumps(payload, default=str, indent=2, sort_keys=True) + "\n")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    temp_path = _temp_output_path(path)
    try:
        with temp_path.open("w", encoding="utf-8") as fh:
            for row in rows:
                fh.write(json.dumps(row, default=str, sort_keys=True))
                fh.write("\n")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_text_atomic(path: Path, text: str) -> None:
    temp_path = _temp_output_path(path)
    try:
        temp_path.write_text(text, encoding="utf-8")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _temp_output_path(path: Path) -> Path:
    return path.parent / f".{path.name}.tmp"


def _safe_review_text(value: str) -> str:
    if not value:
        return value
    return f"'{value}" if value[0] in {"=", "+", "-", "@"} else value


def materialized_record(result: LoadMaterialized) -> dict[str, Any]:
    return {
        "job": result.job_name,
        "title": result.title,
        "workbook": str(result.workbook_path),
        "sheet": result.sheet,
        "header_row": result.header_row,
        "rows_scanned": result.rows_scanned,
        "valid_rows": result.valid_rows,
        "error_rows": result.error_rows,
        "issues": [_issue_record(issue) for issue in result.issues],
        "request_samples": [_request_record(request) for request in result.requests[:MAX_SAMPLES]],
    }


def run_record(result: LoadRunResult) -> dict[str, Any]:
    return {
        "run_id": result.run_id,
        "job": result.job_name,
        "title": result.title,
        "mode": result.mode,
        "dry_run": result.dry_run,
        "workbook": str(result.workbook_path),
        "sheet": result.sheet,
        "started_at": result.started_at,
        "finished_at": result.finished_at,
        "rows_scanned": result.rows_scanned,
        "valid_rows": result.valid_rows,
        "error_rows": result.error_rows,
        "requests": result.request_count,
        "successes": result.success_count,
        "failures": result.failure_count,
        "run_dir": str(result.run_dir),
        "review_workbook": str(result.review_path) if result.review_path else None,
        "issues": [_issue_record(issue) for issue in result.issues],
        "request_samples": [_request_record(request) for request in result.requests[:MAX_SAMPLES]],
        "response_samples": [
            _response_record(response) for response in result.responses[:MAX_SAMPLES]
        ],
    }


def _issue_record(issue: LoadIssue) -> dict[str, Any]:
    return {
        "row": issue.row,
        "code": issue.code,
        "message": issue.message,
        "column": issue.column,
        "sample": issue.sample,
    }


def _request_record(request: LoadRequest) -> dict[str, Any]:
    return {
        "row": request.row,
        "method": request.method,
        "path": request.path,
        "body": request.body,
    }


def _response_record(response: LoadResponse) -> dict[str, Any]:
    return {
        "row": response.row,
        "status_code": response.status_code,
        "ok": response.ok,
        "body": response.body,
    }


def _resolve_key(resolve: LoadResolve) -> str:
    filters = json.dumps(resolve.filters or {}, default=str, sort_keys=True)
    return f"{resolve.endpoint}:{resolve.match}:{resolve.output}:{filters}"


def _resolve_direct_key(resolve: LoadResolve) -> str:
    filters = json.dumps(resolve.filters or {}, default=str, sort_keys=True)
    return f"{resolve.endpoint}:{resolve.output}:{resolve.output}:{filters}"


def _scope_index_key(scope: LoadScope) -> str:
    return f"{scope.endpoint}:{scope.output}:{scope.output}:scope"


def _lookup_key(value: str) -> str:
    return " ".join(value.strip().casefold().split())


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def _extract_path(payload: Any, path: str) -> Any:
    current = payload
    for part in path.split("."):
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    return current


def _json_dict(value: str) -> dict[str, Any]:
    payload = json.loads(value)
    return payload if isinstance(payload, dict) else {}


def _run_id(job_name: str) -> str:
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    suffix = uuid.uuid4().hex[:8]
    safe_job = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "-" for ch in job_name)
    return f"{timestamp}-{safe_job}-{suffix}"


def _utc_iso() -> str:
    return datetime.now(UTC).isoformat(timespec="milliseconds").replace("+00:00", "Z")
