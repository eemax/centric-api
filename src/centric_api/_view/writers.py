from __future__ import annotations

import csv
import re
from collections.abc import Iterable
from pathlib import Path
from typing import TYPE_CHECKING, Any

from ..artifact_names import allocate_artifact_path
from ..config import ConfigError
from ..view_config import ViewColumn, ViewDefinition
from .values import _cell_text, _parse_datetime

if TYPE_CHECKING:
    from .materialize import ViewMaterialized

HEADER_ROW_HEIGHT = 18


def _write_csv(path: Path, materialized: ViewMaterialized) -> None:
    _write_csv_rows(path, materialized.headers, materialized.columns, materialized.rows)


def _write_csv_streaming(
    path: Path,
    headers: tuple[str, ...],
    columns: tuple[ViewColumn, ...],
    rows: Iterable[tuple[Any, ...]],
) -> int:
    return _write_csv_rows(path, headers, columns, rows)


def _write_csv_rows(
    path: Path,
    headers: tuple[str, ...],
    columns: tuple[ViewColumn, ...],
    rows: Iterable[tuple[Any, ...]],
) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = _temp_output_path(path)
    row_count = 0
    try:
        with temp_path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(
                    [_csv_value(value, column) for value, column in zip(row, columns, strict=True)]
                )
                row_count += 1
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return row_count


def _write_xlsx(path: Path, materialized: ViewMaterialized, view: ViewDefinition) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ConfigError("XLSX export requires openpyxl.") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = _temp_output_path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name(view)
    sheet.append(list(materialized.headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_border = Border(bottom=Side(style="thin", color="B7B7B7"))
    header_alignment = Alignment(horizontal="left", vertical="top")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    for raw_row in materialized.rows:
        sheet.append(
            [
                _xlsx_value(value, column)
                for value, column in zip(raw_row, materialized.columns, strict=True)
            ]
        )
    for column_index, column in enumerate(materialized.columns, start=1):
        letter = get_column_letter(column_index)
        for cell in sheet[letter][1:]:
            if cell.value is None:
                continue
            cell.alignment = Alignment(vertical="top")
            if column.number_format:
                cell.number_format = column.number_format
            elif column.type == "date":
                cell.number_format = "yyyy-mm-dd"
            elif column.type == "datetime":
                cell.number_format = "yyyy-mm-dd hh:mm"
        width = column.width
        if width is None and view.options.autosize:
            values = [
                column.header,
                *(_cell_text(row[column_index - 1]) for row in materialized.rows),
            ]
            width = min(max((len(value) for value in values), default=10) + 2, 80)
        if width is not None:
            sheet.column_dimensions[letter].width = width
    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT
    if view.options.freeze_header:
        sheet.freeze_panes = "A2"
    has_table = bool(materialized.headers and materialized.rows)
    if view.options.autofilter and materialized.headers and not has_table:
        sheet.auto_filter.ref = sheet.dimensions
    if has_table:
        table = Table(displayName=_excel_table_name(view.name), ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_xlsx_streaming(
    path: Path,
    headers: tuple[str, ...],
    columns: tuple[ViewColumn, ...],
    rows: Iterable[tuple[Any, ...]],
    view: ViewDefinition,
    *,
    widths: tuple[int | None, ...],
    row_count: int,
) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ConfigError("XLSX export requires openpyxl.") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = _temp_output_path(path)
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(title=_sheet_name(view))
    if view.options.freeze_header:
        sheet.freeze_panes = "A2"
    for column_index, width in enumerate(widths, start=1):
        if width is not None:
            sheet.column_dimensions[get_column_letter(column_index)].width = width
    if view.options.autofilter and headers:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_count + 1}"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_border = Border(bottom=Side(style="thin", color="B7B7B7"))
    header_alignment = Alignment(horizontal="left", vertical="top")
    body_alignment = Alignment(vertical="top")
    header_row = []
    for header in headers:
        cell = WriteOnlyCell(sheet, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
        header_row.append(cell)
    sheet.append(header_row)
    written_count = 0
    for raw_row in rows:
        sheet.append(
            [
                _streaming_xlsx_cell(sheet, value, column, body_alignment, WriteOnlyCell)
                for value, column in zip(raw_row, columns, strict=True)
            ]
        )
        written_count += 1
    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    return written_count


def _measure_xlsx_streaming(
    headers: tuple[str, ...],
    columns: tuple[ViewColumn, ...],
    rows: Iterable[tuple[Any, ...]],
    view: ViewDefinition,
) -> tuple[int, tuple[int | None, ...]]:
    row_count = 0
    auto_widths = [len(header) for header in headers]
    for raw_row in rows:
        if view.options.autosize:
            for index, value in enumerate(raw_row):
                if columns[index].width is None:
                    auto_widths[index] = max(auto_widths[index], len(_cell_text(value)))
        row_count += 1
    widths: list[int | None] = []
    for index, column in enumerate(columns):
        if column.width is not None:
            widths.append(column.width)
        elif view.options.autosize:
            widths.append(min(auto_widths[index] + 2, 80))
        else:
            widths.append(None)
    return row_count, tuple(widths)


def _streaming_xlsx_cell(
    sheet: Any,
    value: Any,
    column: ViewColumn,
    alignment: Any,
    cell_type: Any,
) -> Any:
    typed_value = _xlsx_value(value, column)
    if typed_value is None:
        return None
    number_format = _xlsx_number_format(column)
    if number_format is None:
        return typed_value
    cell = cell_type(sheet, value=typed_value)
    cell.alignment = alignment
    cell.number_format = number_format
    return cell


def _xlsx_number_format(column: ViewColumn) -> str | None:
    if column.number_format:
        return column.number_format
    if column.type == "date":
        return "yyyy-mm-dd"
    if column.type == "datetime":
        return "yyyy-mm-dd hh:mm"
    return None


def _xlsx_value(value: Any, column: ViewColumn) -> Any:
    if value is None:
        return None
    if column.type in {"date", "datetime"}:
        parsed = _parse_datetime(value)
        if parsed is None:
            return _xlsx_text(value)
        return parsed.date() if column.type == "date" else parsed.replace(tzinfo=None)
    if column.type == "number":
        return _number_value(value)
    if column.type == "integer":
        return _integer_value(value)
    if column.type == "boolean":
        return _bool_value(value)
    return _xlsx_text(value)


def _csv_value(value: Any, column: ViewColumn) -> str:
    if value is None:
        return ""
    if column.type == "number":
        return _cell_text(_number_value(value))
    if column.type == "integer":
        return _cell_text(_integer_value(value))
    if column.type == "boolean":
        return _cell_text(_bool_value(value))
    text = _cell_text(value)
    if text and text[0] in {"=", "+", "-", "@"}:
        return f"'{text}"
    return text


def _number_value(value: Any) -> int | float | str | None:
    if isinstance(value, bool):
        return _xlsx_text(value)
    if isinstance(value, int | float):
        return value
    text = str(value).strip()
    if text == "":
        return None
    try:
        number = float(text)
    except ValueError:
        return _xlsx_text(value)
    return int(number) if number.is_integer() else number


def _integer_value(value: Any) -> int | str | None:
    number = _number_value(value)
    if isinstance(number, int):
        return number
    if isinstance(number, float) and number.is_integer():
        return int(number)
    if isinstance(number, float):
        return _xlsx_text(value)
    return number


def _bool_value(value: Any) -> bool | str | None:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "1"}:
        return True
    if text in {"false", "no", "0"}:
        return False
    return _xlsx_text(value)


def _sheet_name(view: ViewDefinition) -> str:
    raw = view.options.sheet_name or view.title or view.name
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", raw).strip() or "Export"
    return cleaned[:31]


def _default_output_path(output_dir: Path, view: ViewDefinition, export_format: str) -> Path:
    return allocate_artifact_path(output_dir, view.name, extension=export_format)


def _temp_output_path(path: Path) -> Path:
    return path.parent / f".{path.name}.tmp"


def _safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip()) or "view"


def _excel_table_name(value: str) -> str:
    cleaned = re.sub(r"\W+", "_", value.strip())
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"View_{cleaned}"
    return cleaned[:255]


def _xlsx_text(value: Any) -> str | None:
    text = _cell_text(value)
    if text == "":
        return None
    if text[0] in {"=", "+", "-", "@"}:
        return f"'{text}"
    return text
