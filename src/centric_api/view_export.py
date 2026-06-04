from __future__ import annotations

import csv
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Literal

from .config import ConfigError
from .store import connect_readonly, table_exists
from .view_config import ViewColumn, ViewConfig, ViewDefinition, ViewFilter, ViewJoin

ExportFormat = Literal["xlsx", "csv"]
SUPPORTED_EXPORT_FORMATS = {"xlsx", "csv"}
HEADER_ROW_HEIGHT = 18
MISSING_JOIN_SAMPLE_LIMIT = 10
SQL_IDENTIFIER_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
SourceKey = tuple[str, str]


@dataclass(frozen=True)
class MissingJoinDetail:
    alias: str
    source_type: str
    source_name: str
    from_path: str
    to_path: str
    missing_count: int
    missing_source_count: int
    missing_ref_count: int
    filtered_out_count: int
    missing_endpoint: bool
    filters_applied: bool
    sample_keys: tuple[str, ...]

    @property
    def endpoint(self) -> str:
        return self.source_name


@dataclass(frozen=True)
class ViewMaterialized:
    root_row_count: int
    headers: tuple[str, ...]
    columns: tuple[ViewColumn, ...]
    rows: tuple[tuple[Any, ...], ...]
    missing_join_count: int
    missing_join_details: tuple[MissingJoinDetail, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class ViewExportResult:
    view_name: str
    title: str
    format: str
    output_path: Path
    row_count: int
    column_count: int
    missing_join_count: int
    missing_join_details: tuple[MissingJoinDetail, ...]
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class ViewCheckResult:
    view_name: str
    title: str
    root_row_count: int
    row_count: int
    column_count: int
    missing_join_count: int
    missing_join_details: tuple[MissingJoinDetail, ...]
    warnings: tuple[str, ...]


def materialize_view(db_path: Path, view: ViewDefinition) -> ViewMaterialized:
    with connect_readonly(db_path) as conn:
        if not table_exists(conn, "endpoint_records"):
            raise ConfigError("View export requires endpoint_records. Run centric-api fetch first.")
        records_by_source, source_exists = _load_records_by_source(conn, _view_sources(view))
    root_key = _source_key(view.root.source_type, view.root.source_name)
    if view.root.source_type == "table" and not source_exists.get(root_key, False):
        message = (
            f"View root table not found: {view.root.source_name}. "
            "Run the model that creates it first."
        )
        raise ConfigError(message)
    indexes = _build_join_indexes(view, records_by_source)
    root_records = records_by_source.get(root_key, [])
    contexts: list[dict[str, Any]] = [{view.root.alias: record} for record in root_records]
    available_aliases = {view.root.alias}
    pending_filters = list(view.filters)
    contexts, pending_filters = _apply_available_filters(
        contexts,
        pending_filters,
        available_aliases,
    )
    missing_join_count = 0
    missing_join_details: list[MissingJoinDetail] = []
    warnings: list[str] = []
    for join in view.joins:
        contexts, detail = _apply_join(
            contexts,
            join,
            indexes[(join.alias, join.to_path)],
            source_exists=source_exists.get(_source_key(join.source_type, join.source_name), False),
            view=view,
            warnings=warnings,
        )
        missing_join_count += detail.missing_count
        if detail.missing_count:
            missing_join_details.append(detail)
        available_aliases.add(join.alias)
        contexts, pending_filters = _apply_available_filters(
            contexts,
            pending_filters,
            available_aliases,
        )
    if pending_filters:
        contexts = [
            context for context in contexts if _matches_filters(context, tuple(pending_filters))
        ]
    rows = tuple(_materialized_row(context, view) for context in contexts)
    return ViewMaterialized(
        root_row_count=len(root_records),
        headers=tuple(column.header for column in view.columns),
        columns=view.columns,
        rows=rows,
        missing_join_count=missing_join_count,
        missing_join_details=tuple(missing_join_details),
        warnings=tuple(warnings),
    )


def check_view(db_path: Path, view: ViewDefinition) -> ViewCheckResult:
    materialized = materialize_view(db_path, view)
    return ViewCheckResult(
        view_name=view.name,
        title=view.title,
        root_row_count=materialized.root_row_count,
        row_count=len(materialized.rows),
        column_count=len(materialized.headers),
        missing_join_count=materialized.missing_join_count,
        missing_join_details=materialized.missing_join_details,
        warnings=materialized.warnings,
    )


def export_view(
    db_path: Path,
    config: ViewConfig,
    view: ViewDefinition,
    *,
    export_format: str = "xlsx",
    output_path: Path | None = None,
) -> ViewExportResult:
    if export_format not in SUPPORTED_EXPORT_FORMATS:
        raise ConfigError(
            f"view export format must be one of: {', '.join(sorted(SUPPORTED_EXPORT_FORMATS))}."
        )
    resolved_output_path = output_path or _default_output_path(
        config.output_dir, view, export_format
    )
    materialized = materialize_view(db_path, view)
    if export_format == "csv":
        _write_csv(resolved_output_path, materialized)
    else:
        _write_xlsx(resolved_output_path, materialized, view)
    return ViewExportResult(
        view_name=view.name,
        title=view.title,
        format=export_format,
        output_path=resolved_output_path,
        row_count=len(materialized.rows),
        column_count=len(materialized.headers),
        missing_join_count=materialized.missing_join_count,
        missing_join_details=materialized.missing_join_details,
        warnings=materialized.warnings,
    )


def infer_export_format(output_path: Path | None, requested_format: str | None) -> str:
    if requested_format is not None:
        if requested_format not in SUPPORTED_EXPORT_FORMATS:
            raise ConfigError(
                f"view export format must be one of: {', '.join(sorted(SUPPORTED_EXPORT_FORMATS))}."
            )
        if output_path is not None:
            suffix = output_path.suffix.lower().lstrip(".")
            if suffix in SUPPORTED_EXPORT_FORMATS and suffix != requested_format:
                raise ConfigError("view export --format must match --output extension.")
        return requested_format
    if output_path is not None:
        suffix = output_path.suffix.lower().lstrip(".")
        if suffix in SUPPORTED_EXPORT_FORMATS:
            return suffix
        if suffix:
            raise ConfigError("view export output extension must be .xlsx or .csv.")
    return "xlsx"


def _load_records_by_source(
    conn: sqlite3.Connection,
    sources: set[SourceKey],
) -> tuple[dict[SourceKey, list[dict[str, Any]]], dict[SourceKey, bool]]:
    records: dict[SourceKey, list[dict[str, Any]]] = {}
    exists: dict[SourceKey, bool] = {}
    for source_type, source_name in sorted(sources):
        key = _source_key(source_type, source_name)
        if source_type == "endpoint":
            rows = conn.execute(
                """
                SELECT record_id, payload_json
                FROM endpoint_records
                WHERE endpoint = ?
                ORDER BY record_id
                """,
                [source_name],
            ).fetchall()
            records[key] = [_json_dict(row["payload_json"]) for row in rows]
            exists[key] = bool(rows)
            continue
        _validate_sql_identifier(source_name, "view source table")
        if not table_exists(conn, source_name):
            records[key] = []
            exists[key] = False
            continue
        rows = conn.execute(
            f"SELECT * FROM {_quote_identifier(source_name)} ORDER BY rowid"
        ).fetchall()
        records[key] = [dict(row) for row in rows]
        exists[key] = True
    return records, exists


def _view_sources(view: ViewDefinition) -> set[SourceKey]:
    return {
        _source_key(view.root.source_type, view.root.source_name),
        *(_source_key(join.source_type, join.source_name) for join in view.joins),
    }


def _build_join_indexes(
    view: ViewDefinition,
    records_by_source: dict[SourceKey, list[dict[str, Any]]],
) -> dict[tuple[str, str], dict[str, list[dict[str, Any]]]]:
    indexes: dict[tuple[str, str], dict[str, list[dict[str, Any]]]] = {}
    for join in view.joins:
        key = (join.alias, join.to_path)
        if key in indexes:
            continue
        index: dict[str, list[dict[str, Any]]] = {}
        source_key = _source_key(join.source_type, join.source_name)
        for payload in records_by_source.get(source_key, []):
            for value in _join_keys(_extract_path(payload, join.to_path)):
                index.setdefault(value, []).append(payload)
        indexes[key] = index
    return indexes


def _apply_join(
    contexts: list[dict[str, Any]],
    join: ViewJoin,
    index: dict[str, list[dict[str, Any]]],
    *,
    source_exists: bool,
    view: ViewDefinition,
    warnings: list[str],
) -> tuple[list[dict[str, Any]], MissingJoinDetail]:
    output: list[dict[str, Any]] = []
    missing_count = 0
    missing_source_count = 0
    missing_ref_count = 0
    filtered_out_count = 0
    sample_keys: list[str] = []
    policy = join.missing or view.options.missing
    for context in contexts:
        keys = _join_keys(_context_value(context, join.from_path))
        raw_matches = _unique_records(record for key in keys for record in index.get(key, []))
        matches = raw_matches
        if join.filters:
            matches = [
                match
                for match in matches
                if _matches_filters({**context, join.alias: match}, join.filters)
            ]
        if not matches:
            missing_count += 1
            if not keys:
                missing_source_count += 1
            elif raw_matches:
                filtered_out_count += 1
            else:
                missing_ref_count += 1
            _append_sample_keys(sample_keys, keys)
            if policy == "error":
                raise ConfigError(
                    f"view {view.name!r} missing join {join.alias!r} from {join.from_path!r}."
                )
            if policy == "drop":
                continue
            output.append(
                {**context, join.alias: [] if join.relationship == "many_concat" else None}
            )
            continue
        if join.relationship == "one":
            if len(matches) > 1:
                warnings.append(
                    f"join {join.alias} matched {len(matches)} rows; first match was used"
                )
            output.append({**context, join.alias: matches[0]})
        elif join.relationship == "many_concat":
            output.append({**context, join.alias: matches})
        else:
            output.extend({**context, join.alias: match} for match in matches)
    return output, MissingJoinDetail(
        alias=join.alias,
        source_type=join.source_type,
        source_name=join.source_name,
        from_path=join.from_path,
        to_path=join.to_path,
        missing_count=missing_count,
        missing_source_count=missing_source_count,
        missing_ref_count=missing_ref_count,
        filtered_out_count=filtered_out_count,
        missing_endpoint=not source_exists,
        filters_applied=bool(join.filters),
        sample_keys=tuple(sample_keys),
    )


def _source_key(source_type: str, source_name: str) -> SourceKey:
    return source_type, source_name


def _validate_sql_identifier(value: str, label: str) -> None:
    if not SQL_IDENTIFIER_PATTERN.match(value):
        raise ConfigError(f"{label} must be a SQLite-safe identifier.")


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _apply_available_filters(
    contexts: list[dict[str, Any]],
    filters: list[ViewFilter],
    available_aliases: set[str],
) -> tuple[list[dict[str, Any]], list[ViewFilter]]:
    available_filters = [item for item in filters if _filter_alias(item) in available_aliases]
    if not available_filters:
        return contexts, filters
    remaining_filters = [item for item in filters if _filter_alias(item) not in available_aliases]
    filtered_contexts = [
        context for context in contexts if _matches_filters(context, tuple(available_filters))
    ]
    return filtered_contexts, remaining_filters


def _filter_alias(item: ViewFilter) -> str:
    return item.path.split(".", 1)[0]


def _matches_filters(context: dict[str, Any], filters: tuple[ViewFilter, ...]) -> bool:
    return all(_matches_filter(context, item) for item in filters)


def _matches_filter(context: dict[str, Any], item: ViewFilter) -> bool:
    found_values = _filter_values(context, item.path)
    if item.operator == "exists":
        return bool(found_values) == item.exists
    if not found_values:
        return False
    return any(_matches_filter_value(value, item) for value in found_values)


def _matches_filter_value(value: Any, item: ViewFilter) -> bool:
    if item.operator == "equals":
        return value == item.equals
    if item.operator == "in":
        return item.in_values is not None and value in item.in_values
    if item.operator == "contains":
        return _contains(value, item.contains)
    if item.operator == "matches":
        return item.matches is not None and re.search(item.matches, str(value or "")) is not None
    if item.operator == "gt":
        return _compare_filter_value(value, item.gt) > 0
    if item.operator == "gte":
        return _compare_filter_value(value, item.gte) >= 0
    if item.operator == "lt":
        return _compare_filter_value(value, item.lt) < 0
    if item.operator == "lte":
        return _compare_filter_value(value, item.lte) <= 0
    return False


def _contains(value: Any, expected: Any) -> bool:
    if isinstance(value, str) and isinstance(expected, str):
        return expected in value
    if isinstance(value, list):
        return expected in value
    return value == expected


def _compare_filter_value(value: Any, expected: Any) -> int:
    left_number = _coerce_number(value)
    right_number = _coerce_number(expected)
    if left_number is not None and right_number is not None:
        return (left_number > right_number) - (left_number < right_number)
    left_datetime = _parse_datetime(value)
    right_datetime = _parse_datetime(expected)
    if left_datetime is not None and right_datetime is not None:
        return (left_datetime > right_datetime) - (left_datetime < right_datetime)
    left = str(value)
    right = str(expected)
    return (left > right) - (left < right)


def _coerce_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        return float(value.strip())
    except ValueError:
        return None


def _materialized_row(context: dict[str, Any], view: ViewDefinition) -> tuple[Any, ...]:
    return tuple(_column_value(context, column.path, view) for column in view.columns)


def _column_value(context: dict[str, Any], path: str, view: ViewDefinition) -> Any:
    alias, _, field_path = path.partition(".")
    value = context.get(alias)
    if isinstance(value, list):
        separator = _separator_for_alias(alias, view)
        parts = [
            _cell_text(_extract_path(item, field_path) if field_path else item)
            for item in value
            if item is not None
        ]
        return separator.join(part for part in parts if part)
    if value is None:
        return None
    return _extract_path(value, field_path) if field_path else value


def _separator_for_alias(alias: str, view: ViewDefinition) -> str:
    for join in view.joins:
        if join.alias == alias:
            return join.separator or view.options.many_separator
    return view.options.many_separator


def _context_value(context: dict[str, Any], path: str) -> Any:
    alias, _, field_path = path.partition(".")
    value = context.get(alias)
    if isinstance(value, list):
        return [_extract_path(item, field_path) if field_path else item for item in value]
    if value is None:
        return None
    return _extract_path(value, field_path) if field_path else value


def _filter_values(context: dict[str, Any], path: str) -> list[Any]:
    value = _context_value(context, path)
    return [item for item in _flatten_values(value) if item is not None]


def _extract_path(payload: Any, path: str) -> Any:
    if not path:
        return payload
    current = payload
    for part in path.split("."):
        if isinstance(current, dict):
            if part not in current:
                return None
            current = current[part]
        elif isinstance(current, list) and part.isdigit():
            index = int(part)
            if index >= len(current):
                return None
            current = current[index]
        else:
            return None
    return current


def _join_keys(value: Any) -> list[str]:
    keys: list[str] = []
    for item in _flatten_values(value):
        if item is None:
            continue
        text = str(item).strip()
        if text and text != "centric:":
            keys.append(text)
    return sorted(set(keys))


def _append_sample_keys(samples: list[str], keys: list[str]) -> None:
    if len(samples) >= MISSING_JOIN_SAMPLE_LIMIT:
        return
    for key in keys:
        if key in samples:
            continue
        samples.append(key)
        if len(samples) >= MISSING_JOIN_SAMPLE_LIMIT:
            return


def _flatten_values(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        flattened: list[Any] = []
        for item in value:
            flattened.extend(_flatten_values(item))
        return flattened
    if isinstance(value, dict):
        flattened = []
        for item in value.values():
            flattened.extend(_flatten_values(item))
        return flattened
    return [value]


def _unique_records(records: Any) -> list[dict[str, Any]]:
    unique: dict[str, dict[str, Any]] = {}
    for record in records:
        key = str(record.get("id") or json.dumps(record, sort_keys=True))
        unique[key] = record
    return [unique[key] for key in sorted(unique)]


def _write_csv(path: Path, materialized: ViewMaterialized) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = _temp_output_path(path)
    try:
        with temp_path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(materialized.headers)
            for row in materialized.rows:
                writer.writerow(
                    [
                        _csv_value(value, column)
                        for value, column in zip(row, materialized.columns, strict=True)
                    ]
                )
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_xlsx(path: Path, materialized: ViewMaterialized, view: ViewDefinition) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ConfigError("XLSX export requires openpyxl.") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = _temp_output_path(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name(view)
    sheet.append(list(materialized.headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_border = Border(bottom=Side(style="thin", color="B7B7B7"))
    header_alignment = Alignment(horizontal="left", vertical="top")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    for raw_row in materialized.rows:
        sheet.append(
            [
                _xlsx_value(value, column)
                for value, column in zip(raw_row, materialized.columns, strict=True)
            ]
        )
    for column_index, column in enumerate(materialized.columns, start=1):
        letter = get_column_letter(column_index)
        for cell in sheet[letter][1:]:
            if cell.value is None:
                continue
            cell.alignment = Alignment(vertical="top")
            if column.number_format:
                cell.number_format = column.number_format
            elif column.type == "date":
                cell.number_format = "yyyy-mm-dd"
            elif column.type == "datetime":
                cell.number_format = "yyyy-mm-dd hh:mm"
        width = column.width
        if width is None and view.options.autosize:
            values = [
                column.header,
                *(_cell_text(row[column_index - 1]) for row in materialized.rows),
            ]
            width = min(max((len(value) for value in values), default=10) + 2, 80)
        if width is not None:
            sheet.column_dimensions[letter].width = width
    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT
    if view.options.freeze_header:
        sheet.freeze_panes = "A2"
    has_table = bool(materialized.headers and materialized.rows)
    if view.options.autofilter and materialized.headers and not has_table:
        sheet.auto_filter.ref = sheet.dimensions
    if has_table:
        table = Table(displayName=_excel_table_name(view.name), ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _xlsx_value(value: Any, column: ViewColumn) -> Any:
    if value is None:
        return None
    if column.type in {"date", "datetime"}:
        parsed = _parse_datetime(value)
        if parsed is None:
            return _xlsx_text(value)
        return parsed.date() if column.type == "date" else parsed.replace(tzinfo=None)
    if column.type == "number":
        return _number_value(value)
    if column.type == "integer":
        return _integer_value(value)
    if column.type == "boolean":
        return _bool_value(value)
    return _xlsx_text(value)


def _csv_value(value: Any, column: ViewColumn) -> str:
    if value is None:
        return ""
    if column.type == "number":
        return _cell_text(_number_value(value))
    if column.type == "integer":
        return _cell_text(_integer_value(value))
    if column.type == "boolean":
        return _cell_text(_bool_value(value))
    text = _cell_text(value)
    if text and text[0] in {"=", "+", "-", "@"}:
        return f"'{text}"
    return text


def _number_value(value: Any) -> int | float | str | None:
    if isinstance(value, bool):
        return _xlsx_text(value)
    if isinstance(value, int | float):
        return value
    text = str(value).strip()
    if text == "":
        return None
    try:
        number = float(text)
    except ValueError:
        return _xlsx_text(value)
    return int(number) if number.is_integer() else number


def _integer_value(value: Any) -> int | str | None:
    number = _number_value(value)
    if isinstance(number, int):
        return number
    if isinstance(number, float) and number.is_integer():
        return int(number)
    if isinstance(number, float):
        return _xlsx_text(value)
    return number


def _bool_value(value: Any) -> bool | str | None:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "1"}:
        return True
    if text in {"false", "no", "0"}:
        return False
    return _xlsx_text(value)


def _parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.astimezone(UTC) if value.tzinfo is not None else value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if not isinstance(value, str) or not value.strip():
        return None
    text = value.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    return parsed.astimezone(UTC) if parsed.tzinfo is not None else parsed


def _sheet_name(view: ViewDefinition) -> str:
    raw = view.options.sheet_name or view.title or view.name
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", raw).strip() or "Export"
    return cleaned[:31]


def _default_output_path(output_dir: Path, view: ViewDefinition, export_format: str) -> Path:
    timestamp = datetime.now(UTC).strftime("%Y-%m-%dT%H%M%SZ")
    return output_dir / f"{_safe_name(view.name)}-{timestamp}.{export_format}"


def _temp_output_path(path: Path) -> Path:
    return path.parent / f".{path.name}.tmp"


def _safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip()) or "view"


def _excel_table_name(value: str) -> str:
    cleaned = re.sub(r"\W+", "_", value.strip())
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"View_{cleaned}"
    return cleaned[:255]


def _xlsx_text(value: Any) -> str | None:
    text = _cell_text(value)
    if text == "":
        return None
    if text[0] in {"=", "+", "-", "@"}:
        return f"'{text}"
    return text


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int | float):
        return str(value)
    return json.dumps(value, sort_keys=True, ensure_ascii=False)


def _json_dict(value: str | None) -> dict[str, Any]:
    if not value:
        return {}
    payload = json.loads(value)
    return payload if isinstance(payload, dict) else {}
