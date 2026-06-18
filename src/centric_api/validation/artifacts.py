from __future__ import annotations

import json
import re
from dataclasses import asdict
from datetime import UTC, datetime
from itertools import islice
from pathlib import Path
from typing import Any

from .contracts import ValidationFinding, ValidationHistoryMetric, ValidationResult

SUMMARY_COLUMNS = ("metric", "value")
VALIDATION_ARTIFACT_TIMESTAMP_FORMAT = "%y-%m-%d-%H%M"
FINDING_COLUMNS = (
    "severity",
    "code",
    "message",
    "style_id",
    "style_name",
    "brand",
    "season",
    "endpoint",
    "record_id",
    "record_name",
    "source_endpoint",
    "source_record_id",
    "source_field",
    "details",
)


def write_validation_artifacts(
    output_dir: Path,
    result: ValidationResult,
    *,
    run_record: dict[str, Any],
) -> tuple[Path, Path, Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = output_dir / "summary.json"
    findings_path = output_dir / "findings.json"
    history_path = output_dir / "history.json"
    artifact_timestamp = str(
        run_record.get("artifact_timestamp")
        or validation_artifact_timestamp(run_record.get("started_at"))
    )
    report_path = output_dir / f"report_{artifact_timestamp}.xlsx"
    artifact_run_record = {
        **run_record,
        "artifact_timestamp": artifact_timestamp,
        "report_path": str(report_path),
    }

    summary_payload = {**artifact_run_record, "summary": result.summary}
    finding_rows = _exported_finding_records(result)
    findings_payload = {
        "total_findings": artifact_run_record["findings"],
        "errors": artifact_run_record["errors"],
        "warnings": artifact_run_record["warnings"],
        "info": artifact_run_record["info"],
        "exported_findings": len(finding_rows),
        "truncated": len(finding_rows) < int(artifact_run_record["findings"]),
        "findings": finding_rows,
    }
    _write_json(summary_path, summary_payload)
    _write_json(findings_path, findings_payload)
    _write_json(history_path, _history_payload(result, artifact_run_record))
    write_validation_workbook(report_path, result, run_record=artifact_run_record)
    return report_path, summary_path, findings_path, history_path


def validation_artifact_timestamp(value: str | datetime | None = None) -> str:
    parsed = _artifact_datetime(value)
    return parsed.strftime(VALIDATION_ARTIFACT_TIMESTAMP_FORMAT)


def write_validation_workbook(
    path: Path,
    result: ValidationResult,
    *,
    run_record: dict[str, Any],
) -> None:
    if result.report_workbook is not None:
        _write_bytes(path, result.report_workbook)
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise RuntimeError("XLSX validation reports require openpyxl.") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _append_sheet_rows(
        summary,
        SUMMARY_COLUMNS,
        _summary_rows(result, run_record),
    )
    for sheet in result.sheets:
        _append_sheet_rows(
            workbook.create_sheet(_sheet_name(sheet.name)),
            sheet.columns or _headers_from_rows(sheet.rows),
            sheet.rows,
        )
    _append_sheet_rows(
        workbook.create_sheet("Findings"),
        FINDING_COLUMNS,
        _exported_finding_records(result),
    )
    for sheet in workbook.worksheets:
        _format_sheet(
            sheet,
            get_column_letter,
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
    temp_path = path.parent / f".{path.name}.tmp"
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def finding_record(finding: ValidationFinding) -> dict[str, Any]:
    payload = asdict(finding)
    payload["details"] = (
        json.dumps(finding.details, default=str, sort_keys=True) if finding.details else None
    )
    return payload


def history_metric_record(metric: ValidationHistoryMetric) -> dict[str, Any]:
    return {
        "scope": metric.scope,
        "brand": metric.brand,
        "metric": metric.metric,
        "value": metric.value,
        "unit": metric.unit,
        "trend": metric.trend,
        "numerator": metric.numerator,
        "denominator": metric.denominator,
        "dimensions": metric.dimensions,
    }


def _history_payload(
    result: ValidationResult,
    run_record: dict[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": 2,
        "validator": run_record["validator"],
        "title": run_record["title"],
        "run_id": run_record["run_id"],
        "status": run_record["status"],
        "started_at": run_record["started_at"],
        "finished_at": run_record["finished_at"],
        "artifact_timestamp": run_record.get("artifact_timestamp"),
        "report_path": run_record.get("report_path"),
        "metrics": [history_metric_record(metric) for metric in result.history_metrics],
    }


def _artifact_datetime(value: str | datetime | None) -> datetime:
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(UTC)
        return value.replace(tzinfo=UTC)
    if isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            pass
        else:
            if parsed.tzinfo is not None:
                return parsed.astimezone(UTC)
            return parsed.replace(tzinfo=UTC)
    return datetime.now(UTC)


def _exported_finding_records(result: ValidationResult) -> list[dict[str, Any]]:
    findings = _exported_findings(result)
    return [finding_record(finding) for finding in findings]


def _exported_findings(result: ValidationResult) -> tuple[ValidationFinding, ...]:
    findings = result.finding_samples or result.findings
    limit = _findings_export_limit(result)
    if limit is not None:
        return tuple(islice(findings, limit))
    return findings


def _findings_export_limit(result: ValidationResult) -> int | None:
    if result.findings_export_limit is None:
        return None
    try:
        limit = int(result.findings_export_limit)
    except (TypeError, ValueError):
        return None
    return max(limit, 0)


def _summary_rows(result: ValidationResult, run_record: dict[str, Any]) -> list[dict[str, Any]]:
    rows = [
        {"metric": "validator", "value": run_record["validator"]},
        {"metric": "title", "value": run_record["title"]},
        {"metric": "status", "value": run_record["status"]},
        {"metric": "run_id", "value": run_record["run_id"]},
        {"metric": "started_at", "value": run_record["started_at"]},
        {"metric": "finished_at", "value": run_record["finished_at"]},
        {"metric": "findings", "value": run_record["findings"]},
        {"metric": "errors", "value": run_record["errors"]},
        {"metric": "warnings", "value": run_record["warnings"]},
        {"metric": "info", "value": run_record["info"]},
    ]
    rows.extend({"metric": key, "value": value} for key, value in result.summary.items())
    return rows


def _write_json(path: Path, payload: Any) -> None:
    temp_path = path.parent / f".{path.name}.tmp"
    try:
        temp_path.write_text(json.dumps(payload, indent=2, default=str, sort_keys=True) + "\n")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_bytes(path: Path, payload: bytes) -> None:
    temp_path = path.parent / f".{path.name}.tmp"
    try:
        temp_path.write_bytes(payload)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _append_sheet_rows(sheet: Any, headers: tuple[str, ...], rows: Any) -> None:
    header_list = tuple(headers)
    sheet.append([_display_header(header) for header in header_list])
    for row in rows:
        sheet.append([_cell_value(row.get(header)) for header in header_list])


def _headers_from_rows(rows: tuple[dict[str, Any], ...]) -> tuple[str, ...]:
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return tuple(headers) or ("value",)


def _format_sheet(
    sheet: Any,
    get_column_letter: Any,
    alignment_cls: Any,
    border_cls: Any,
    font_cls: Any,
    pattern_fill_cls: Any,
    side_cls: Any,
) -> None:
    header_fill = pattern_fill_cls("solid", fgColor="1F4E78")
    header_font = font_cls(bold=True, color="FFFFFF", size=11)
    header_border = border_cls(bottom=side_cls(style="thin", color="B7B7B7"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = alignment_cls(horizontal="left", vertical="top")
    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        values = [
            str(sheet.cell(row=row_index, column=column_index).value or "")
            for row_index in range(1, min(sheet.max_row, 250) + 1)
        ]
        width = min(max((len(value) for value in values), default=10) + 2, 70)
        sheet.column_dimensions[letter].width = width
    sheet.row_dimensions[1].height = 18
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool | int | float):
        return value
    if isinstance(value, dict | list | tuple):
        text = json.dumps(value, default=str, sort_keys=True)
    else:
        text = str(value)
    if text == "":
        return None
    if text[0] in {"=", "+", "-", "@"}:
        return f"'{text}"
    return text


def _display_header(value: str) -> str:
    return value.replace("_", " ").title()


def _sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", value).strip() or "Sheet"
    return cleaned[:31]
