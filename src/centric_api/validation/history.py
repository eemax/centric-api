from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from importlib import resources
from math import isfinite
from pathlib import Path
from typing import Any, Literal

from ..config import ConfigError, runtime_home

HistoryGroup = Literal["day", "week", "month"]

DEFAULT_VALIDATION_HISTORY_DIR = Path("validation/history")
DEFAULT_VALIDATION_RUNS_DIR = Path("validation/runs")
HISTORY_SCHEMA_VERSION = 1
HISTORY_ARTIFACT_NAME = "history.json"
HISTORY_OUTPUT_JSON = "history.json"
HISTORY_OUTPUT_XLSX = "history.xlsx"
HISTORY_OUTPUT_HTML = "history.html"
HISTORY_TEMPLATE_NAME = "validation-history.html"


@dataclass(frozen=True)
class ValidationHistoryOutput:
    group: HistoryGroup
    runs_dir: Path
    output_dir: Path
    json_path: Path
    workbook_path: Path
    html_path: Path
    raw_metric_count: int
    point_count: int
    run_count: int
    validators: tuple[str, ...]
    metrics: tuple[str, ...]


def build_validation_history(
    *,
    runs_dir: str | Path | None = None,
    output_dir: str | Path | None = None,
    group: HistoryGroup = "week",
    validators: tuple[str, ...] = (),
) -> ValidationHistoryOutput:
    resolved_runs_dir = (
        Path(runs_dir).expanduser() if runs_dir else runtime_home() / DEFAULT_VALIDATION_RUNS_DIR
    )
    resolved_output_dir = (
        Path(output_dir).expanduser()
        if output_dir
        else runtime_home() / DEFAULT_VALIDATION_HISTORY_DIR
    )
    if group not in {"day", "week", "month"}:
        raise ConfigError("History group must be one of: day, week, month.")

    raw_points = _load_history_metrics(resolved_runs_dir, validators=set(validators))
    points = _group_latest_points(raw_points, group=group)
    payload = _history_payload(
        group=group,
        runs_dir=resolved_runs_dir,
        output_dir=resolved_output_dir,
        raw_points=raw_points,
        points=points,
    )
    resolved_output_dir.mkdir(parents=True, exist_ok=True)
    json_path = resolved_output_dir / HISTORY_OUTPUT_JSON
    workbook_path = resolved_output_dir / HISTORY_OUTPUT_XLSX
    html_path = resolved_output_dir / HISTORY_OUTPUT_HTML
    _write_json(json_path, payload)
    _write_workbook(workbook_path, payload)
    _write_html(html_path, payload)
    return ValidationHistoryOutput(
        group=group,
        runs_dir=resolved_runs_dir,
        output_dir=resolved_output_dir,
        json_path=json_path,
        workbook_path=workbook_path,
        html_path=html_path,
        raw_metric_count=len(raw_points),
        point_count=len(points),
        run_count=_run_count(raw_points),
        validators=tuple(payload["validators"]),
        metrics=tuple(payload["metrics"]),
    )


def _load_history_metrics(runs_dir: Path, *, validators: set[str]) -> list[dict[str, Any]]:
    if not runs_dir.exists():
        return []
    points: list[dict[str, Any]] = []
    for history_path in sorted(runs_dir.glob(f"*/*/{HISTORY_ARTIFACT_NAME}")):
        payload = _load_json_object(history_path)
        if not payload:
            continue
        if _schema_version(payload) != HISTORY_SCHEMA_VERSION:
            continue
        validator = str(payload.get("validator") or history_path.parent.parent.name)
        if validators and validator not in validators:
            continue
        run_id = str(payload.get("run_id") or history_path.parent.name)
        started_at = str(payload.get("started_at") or "")
        started_dt = _parse_datetime(started_at)
        if started_dt is None:
            continue
        metrics = payload.get("metrics")
        if not isinstance(metrics, list):
            continue
        for metric in metrics:
            point = _history_point(
                metric,
                payload=payload,
                history_path=history_path,
                validator=validator,
                run_id=run_id,
                started_at=started_dt,
            )
            if point is not None:
                points.append(point)
    return sorted(points, key=lambda item: (item["started_at"], item["validator"], item["metric"]))


def _history_point(
    metric: Any,
    *,
    payload: dict[str, Any],
    history_path: Path,
    validator: str,
    run_id: str,
    started_at: datetime,
) -> dict[str, Any] | None:
    if not isinstance(metric, dict):
        return None
    metric_name = str(metric.get("metric") or "").strip()
    value = _numeric_value(metric.get("value"))
    if not metric_name or value is None:
        return None
    unit = str(metric.get("unit") or "number")
    scope = str(metric.get("scope") or "overall")
    brand = metric.get("brand")
    dimensions = metric.get("dimensions")
    return {
        "validator": validator,
        "title": payload.get("title"),
        "run_id": run_id,
        "status": payload.get("status"),
        "started_at": _isoformat(started_at),
        "finished_at": payload.get("finished_at"),
        "scope": scope,
        "brand": str(brand) if brand is not None else None,
        "metric": metric_name,
        "value": value,
        "unit": unit,
        "numerator": _numeric_value(metric.get("numerator")),
        "denominator": _numeric_value(metric.get("denominator")),
        "dimensions": dimensions if isinstance(dimensions, dict) else {},
        "history_path": str(history_path),
        "report_path": str(history_path.parent / "report.xlsx"),
    }


def _group_latest_points(
    points: list[dict[str, Any]], *, group: HistoryGroup
) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str, str | None, str], dict[str, Any]] = {}
    for point in points:
        started_at = _parse_datetime(str(point["started_at"]))
        if started_at is None:
            continue
        bucket_start = _bucket_start(started_at, group)
        key = (
            str(point["validator"]),
            str(point["metric"]),
            str(point["unit"]),
            str(point["scope"]),
            point.get("brand"),
            _isoformat(bucket_start),
        )
        current = grouped.get(key)
        if current is None or str(point["started_at"]) > str(current["started_at"]):
            grouped[key] = {
                **point,
                "bucket": _bucket_label(bucket_start, group),
                "bucket_start": _isoformat(bucket_start),
            }
    return sorted(
        grouped.values(),
        key=lambda item: (
            item["validator"],
            item["metric"],
            item["scope"],
            item.get("brand") or "",
            item["bucket_start"],
        ),
    )


def _history_payload(
    *,
    group: HistoryGroup,
    runs_dir: Path,
    output_dir: Path,
    raw_points: list[dict[str, Any]],
    points: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "generated_at": _isoformat(datetime.now(UTC)),
        "group": group,
        "runs_dir": str(runs_dir),
        "output_dir": str(output_dir),
        "run_count": _run_count(raw_points),
        "raw_metric_count": len(raw_points),
        "point_count": len(points),
        "validators": sorted({point["validator"] for point in points}),
        "metrics": sorted({point["metric"] for point in points}),
        "points": points,
        "raw_points": raw_points,
    }


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    temp_path = path.parent / f".{path.name}.tmp"
    try:
        temp_path.write_text(
            json.dumps(payload, indent=2, sort_keys=True, default=str) + "\n",
            encoding="utf-8",
        )
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_workbook(path: Path, payload: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ConfigError("XLSX validation history requires openpyxl.") from exc

    workbook = Workbook()
    points_sheet = workbook.active
    points_sheet.title = "History"
    _append_rows(points_sheet, _point_rows(payload["points"]))
    latest_sheet = workbook.create_sheet("Latest")
    _append_rows(latest_sheet, _latest_rows(payload["points"]))
    runs_sheet = workbook.create_sheet("Runs")
    _append_rows(runs_sheet, _run_rows(payload["raw_points"]))
    for sheet in workbook.worksheets:
        _format_sheet(sheet, get_column_letter, Alignment, Border, Font, PatternFill, Side)
        if sheet.max_row > 1 and sheet.max_column > 1:
            table = Table(displayName=_table_name(sheet.title), ref=sheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
    temp_path = path.parent / f".{path.name}.tmp"
    try:
        workbook.save(temp_path)
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _append_rows(sheet: Any, rows: list[dict[str, Any]]) -> None:
    headers = tuple(rows[0]) if rows else ("message",)
    sheet.append([_display_header(header) for header in headers])
    for row in rows:
        sheet.append([_cell_value(row.get(header)) for header in headers])


def _point_rows(points: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "bucket": point["bucket"],
            "bucket_start": point["bucket_start"],
            "validator": point["validator"],
            "metric": point["metric"],
            "scope": point["scope"],
            "brand": point.get("brand"),
            "value": point["value"],
            "unit": point["unit"],
            "numerator": point.get("numerator"),
            "denominator": point.get("denominator"),
            "run_id": point["run_id"],
            "started_at": point["started_at"],
            "status": point.get("status"),
            "report_path": point.get("report_path"),
        }
        for point in points
    ] or [{"message": "No validation history metrics found."}]


def _latest_rows(points: list[dict[str, Any]]) -> list[dict[str, Any]]:
    latest: dict[tuple[str, str, str, str | None], dict[str, Any]] = {}
    for point in points:
        key = (point["validator"], point["metric"], point["scope"], point.get("brand"))
        current = latest.get(key)
        if current is None or point["bucket_start"] > current["bucket_start"]:
            latest[key] = point
    rows = [
        {
            "validator": point["validator"],
            "metric": point["metric"],
            "scope": point["scope"],
            "brand": point.get("brand"),
            "latest_bucket": point["bucket"],
            "value": point["value"],
            "unit": point["unit"],
            "numerator": point.get("numerator"),
            "denominator": point.get("denominator"),
            "run_id": point["run_id"],
            "started_at": point["started_at"],
        }
        for point in sorted(
            latest.values(),
            key=lambda item: (
                item["validator"],
                item["metric"],
                item["scope"],
                item.get("brand") or "",
            ),
        )
    ]
    return rows or [{"message": "No validation history metrics found."}]


def _run_rows(points: list[dict[str, Any]]) -> list[dict[str, Any]]:
    runs: dict[tuple[str, str], dict[str, Any]] = {}
    for point in points:
        runs[(point["validator"], point["run_id"])] = {
            "validator": point["validator"],
            "run_id": point["run_id"],
            "started_at": point["started_at"],
            "finished_at": point.get("finished_at"),
            "status": point.get("status"),
            "history_path": point.get("history_path"),
            "report_path": point.get("report_path"),
        }
    return sorted(runs.values(), key=lambda item: (item["validator"], item["started_at"])) or [
        {"message": "No validation history runs found."}
    ]


def _format_sheet(
    sheet: Any,
    get_column_letter: Any,
    alignment_cls: Any,
    border_cls: Any,
    font_cls: Any,
    pattern_fill_cls: Any,
    side_cls: Any,
) -> None:
    header_fill = pattern_fill_cls("solid", fgColor="1F4E78")
    header_font = font_cls(bold=True, color="FFFFFF", size=11)
    header_border = border_cls(bottom=side_cls(style="thin", color="B7B7B7"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = alignment_cls(horizontal="left", vertical="top")
    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        values = [
            str(sheet.cell(row=row_index, column=column_index).value or "")
            for row_index in range(1, min(sheet.max_row, 250) + 1)
        ]
        sheet.column_dimensions[letter].width = min(
            max((len(value) for value in values), default=10) + 2,
            72,
        )
    sheet.row_dimensions[1].height = 18
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions


def _write_html(path: Path, payload: dict[str, Any]) -> None:
    html = _validation_history_template()
    replacements = {
        "__GENERATED_AT__": str(payload["generated_at"]),
        "__GROUP__": str(payload["group"]),
        "__POINT_COUNT__": str(payload["point_count"]),
        "__RUN_COUNT__": str(payload["run_count"]),
        "__HISTORY_JSON__": _script_json(payload, sort_keys=True),
    }
    for placeholder, value in replacements.items():
        html = html.replace(placeholder, value)
    temp_path = path.parent / f".{path.name}.tmp"
    try:
        temp_path.write_text(html, encoding="utf-8")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _validation_history_template() -> str:
    return (
        resources.files("centric_api.templates")
        .joinpath(HISTORY_TEMPLATE_NAME)
        .read_text(encoding="utf-8")
    )


def _script_json(value: Any, *, sort_keys: bool = False) -> str:
    return json.dumps(value, sort_keys=sort_keys, default=str).replace("</", "<\\/")


def _load_json_object(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _schema_version(payload: dict[str, Any]) -> int:
    try:
        return int(payload.get("schema_version") or 0)
    except (TypeError, ValueError):
        return 0


def _parse_datetime(value: str) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def _bucket_start(value: datetime, group: HistoryGroup) -> datetime:
    value = value.astimezone(UTC)
    if group == "day":
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    if group == "month":
        return value.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start = value - timedelta(days=value.weekday())
    return start.replace(hour=0, minute=0, second=0, microsecond=0)


def _bucket_label(value: datetime, group: HistoryGroup) -> str:
    if group == "day":
        return value.strftime("%Y-%m-%d")
    if group == "month":
        return value.strftime("%Y-%m")
    year, week, _day = value.isocalendar()
    return f"{year}-W{week:02d}"


def _isoformat(value: datetime) -> str:
    return value.astimezone(UTC).isoformat().replace("+00:00", "Z")


def _numeric_value(value: Any) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return value if isfinite(float(value)) else None
    try:
        parsed = float(str(value).strip().removesuffix("%"))
    except ValueError:
        return None
    if not isfinite(parsed):
        return None
    return int(parsed) if parsed.is_integer() else parsed


def _run_count(points: list[dict[str, Any]]) -> int:
    return len({(point["validator"], point["run_id"]) for point in points})


def _display_header(value: str) -> str:
    return value.replace("_", " ").title()


def _cell_value(value: Any) -> Any:
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return value


def _table_name(value: str) -> str:
    return f"Validation{''.join(character for character in value.title() if character.isalnum())}"
