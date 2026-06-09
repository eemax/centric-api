from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..auth import AuthContext
from ..config import ConfigError, runtime_path
from ..load_config import LoadConfig, LoadJob
from .artifacts import (
    _emit_progress,
    _execute_chained_load_request,
    _has_row_issues,
    _response_field,
    _write_requests,
    _write_responses,
    _write_review_workbook,
    _write_summary,
)
from .excel import (
    _include_retry_row,
    _iter_data_rows,
    _map_headers,
    _retry_status_index,
    _select_sheet,
)
from .models import (
    LOAD_RUNS_DIR,
    REVIEW_WORKBOOK_NAME,
    LoadIssue,
    LoadMaterialized,
    LoadProgressCallback,
    LoadRequest,
    LoadResponse,
    LoadRunResult,
    MaterialCreateCompositionQuoteRow,
    MaterialCreateCompositionRow,
    MaterialSupplierQuoteRow,
)
from .references import _build_reference_indexes, _build_value_set_indexes, _row_values
from .style_supplier_quote import (
    _resolve_style_supplier_quote_memberships,
    _style_supplier_quote_references,
    _supplier_quote_item_request,
    _supplier_quote_product_source_request,
    _supplier_quote_production_request,
    _supplier_quote_revision_request,
)
from .utils import _is_blank, _run_id, _utc_iso


def materialize_material_create_with_composition_workflow(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None = None,
    limit: int | None = None,
    mode: str = "check",
    retry_statuses: set[str] | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadMaterialized:
    parsed = _material_create_composition_rows(
        db_path,
        job,
        workbook_path,
        sheet=sheet,
        limit=limit,
        mode=mode,
        retry_statuses=retry_statuses,
        progress_callback=progress_callback,
    )
    requests = _material_create_composition_planned_requests(parsed["rows"])
    _emit_progress(
        progress_callback,
        {
            "event": "load_validate",
            "scanned": parsed["rows_scanned"],
            "valid": len(parsed["rows"]),
            "errors": parsed["error_rows"],
        },
    )
    return LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        issues=tuple(parsed["issues"]),
        requests=tuple(requests),
    )


def materialize_material_create_with_composition_and_quote_workflow(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None = None,
    limit: int | None = None,
    mode: str = "check",
    retry_statuses: set[str] | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadMaterialized:
    parsed = _material_create_composition_quote_rows(
        db_path,
        job,
        workbook_path,
        sheet=sheet,
        limit=limit,
        mode=mode,
        retry_statuses=retry_statuses,
        progress_callback=progress_callback,
    )
    requests = _material_create_composition_quote_planned_requests(parsed["rows"])
    _emit_progress(
        progress_callback,
        {
            "event": "load_validate",
            "scanned": parsed["rows_scanned"],
            "valid": len(parsed["rows"]),
            "errors": parsed["error_rows"],
        },
    )
    return LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        issues=tuple(parsed["issues"]),
        requests=tuple(requests),
    )


def run_material_create_with_composition_workflow(
    db_path: Path,
    config: LoadConfig,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    dry_run: bool,
    yes: bool,
    retry_statuses: set[str] | None = None,
    materialized: LoadMaterialized | None = None,
    auth_ctx: AuthContext | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadRunResult:
    if not dry_run and not yes:
        raise ConfigError("Non-dry-run load requires --yes.")
    mode = (
        "retry-dry-run"
        if retry_statuses and dry_run
        else ("retry" if retry_statuses else ("dry-run" if dry_run else "run"))
    )
    parsed = _material_create_composition_rows(
        db_path,
        job,
        workbook_path,
        sheet=sheet,
        limit=limit,
        mode=mode,
        retry_statuses=retry_statuses,
        progress_callback=progress_callback,
    )
    started_at = _utc_iso()
    run_id = _run_id(job.name)
    run_dir = runtime_path(LOAD_RUNS_DIR / run_id)
    run_dir.mkdir(parents=True, exist_ok=True)

    materialized = materialized or LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        issues=tuple(parsed["issues"]),
        requests=tuple(_material_create_composition_planned_requests(parsed["rows"])),
    )
    requests = list(materialized.requests)
    responses: list[LoadResponse] = []
    issues = list(materialized.issues)
    workflow_issues: list[LoadIssue] = []
    if dry_run:
        _write_requests(run_dir / "requests.jsonl", tuple(requests))
    else:
        if parsed["rows"] and auth_ctx is None:
            raise ConfigError("Load run requires an auth context.")
        requests = []
        responses = []
        if parsed["rows"]:
            workflow_issues = _execute_material_create_composition_workflow(
                auth_ctx,
                parsed["rows"],
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            issues.extend(workflow_issues)
        _write_requests(run_dir / "requests.jsonl", tuple(requests))
        _write_responses(run_dir / "responses.jsonl", tuple(responses))
    _emit_progress(
        progress_callback,
        {
            "event": "load_artifacts",
            "run_dir": str(run_dir),
            "requests": len(requests),
        },
    )

    finished_at = _utc_iso()
    materialized = LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]) + len(workflow_issues if not dry_run else ()),
        issues=tuple(issues),
        requests=tuple(requests),
    )
    review_path = None
    if responses or _has_row_issues(materialized.issues):
        review_path = _write_review_workbook(
            materialized,
            responses=tuple(responses),
            run_id=run_id,
            processed_at=finished_at,
            output_path=run_dir / REVIEW_WORKBOOK_NAME,
        )
    result = LoadRunResult(
        run_id=run_id,
        job_name=job.name,
        title=job.title,
        mode=mode,
        dry_run=dry_run,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=materialized.error_rows,
        request_count=len(requests),
        success_count=sum(1 for response in responses if response.ok),
        failure_count=sum(1 for response in responses if not response.ok),
        issues=tuple(issues),
        requests=tuple(requests),
        responses=tuple(responses),
        run_dir=run_dir,
        review_path=review_path,
        started_at=started_at,
        finished_at=finished_at,
    )
    _write_summary(run_dir / "summary.json", result, config)
    return result


def run_material_create_with_composition_and_quote_workflow(
    db_path: Path,
    config: LoadConfig,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    dry_run: bool,
    yes: bool,
    retry_statuses: set[str] | None = None,
    materialized: LoadMaterialized | None = None,
    auth_ctx: AuthContext | None = None,
    progress_callback: LoadProgressCallback | None = None,
) -> LoadRunResult:
    if not dry_run and not yes:
        raise ConfigError("Non-dry-run load requires --yes.")
    mode = (
        "retry-dry-run"
        if retry_statuses and dry_run
        else ("retry" if retry_statuses else ("dry-run" if dry_run else "run"))
    )
    parsed = _material_create_composition_quote_rows(
        db_path,
        job,
        workbook_path,
        sheet=sheet,
        limit=limit,
        mode=mode,
        retry_statuses=retry_statuses,
        progress_callback=progress_callback,
    )
    started_at = _utc_iso()
    run_id = _run_id(job.name)
    run_dir = runtime_path(LOAD_RUNS_DIR / run_id)
    run_dir.mkdir(parents=True, exist_ok=True)

    materialized = materialized or LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]),
        issues=tuple(parsed["issues"]),
        requests=tuple(_material_create_composition_quote_planned_requests(parsed["rows"])),
    )
    requests = list(materialized.requests)
    responses: list[LoadResponse] = []
    issues = list(materialized.issues)
    workflow_issues: list[LoadIssue] = []
    if dry_run:
        _write_requests(run_dir / "requests.jsonl", tuple(requests))
    else:
        if parsed["rows"] and auth_ctx is None:
            raise ConfigError("Load run requires an auth context.")
        requests = []
        responses = []
        if parsed["rows"]:
            workflow_issues = _execute_material_create_composition_quote_workflow(
                auth_ctx,
                parsed["rows"],
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            issues.extend(workflow_issues)
        _write_requests(run_dir / "requests.jsonl", tuple(requests))
        _write_responses(run_dir / "responses.jsonl", tuple(responses))
    _emit_progress(
        progress_callback,
        {
            "event": "load_artifacts",
            "run_dir": str(run_dir),
            "requests": len(requests),
        },
    )

    finished_at = _utc_iso()
    materialized = LoadMaterialized(
        job_name=job.name,
        title=job.title,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        header_row=job.input.header_row,
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=int(parsed["error_rows"]) + len(workflow_issues if not dry_run else ()),
        issues=tuple(issues),
        requests=tuple(requests),
    )
    review_path = None
    if responses or _has_row_issues(materialized.issues):
        review_path = _write_review_workbook(
            materialized,
            responses=tuple(responses),
            run_id=run_id,
            processed_at=finished_at,
            output_path=run_dir / REVIEW_WORKBOOK_NAME,
        )
    result = LoadRunResult(
        run_id=run_id,
        job_name=job.name,
        title=job.title,
        mode=mode,
        dry_run=dry_run,
        workbook_path=Path(workbook_path).expanduser(),
        sheet=str(parsed["sheet"]),
        rows_scanned=int(parsed["rows_scanned"]),
        valid_rows=len(parsed["rows"]),
        error_rows=materialized.error_rows,
        request_count=len(requests),
        success_count=sum(1 for response in responses if response.ok),
        failure_count=sum(1 for response in responses if not response.ok),
        issues=tuple(issues),
        requests=tuple(requests),
        responses=tuple(responses),
        run_dir=run_dir,
        review_path=review_path,
        started_at=started_at,
        finished_at=finished_at,
    )
    _write_summary(run_dir / "summary.json", result, config)
    return result


def _material_create_composition_rows(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    mode: str,
    retry_statuses: set[str] | None,
    progress_callback: LoadProgressCallback | None,
) -> dict[str, Any]:
    _require_material_create_composition_columns(job)
    workbook_path = workbook_path.expanduser()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = _select_sheet(workbook, sheet)
        _emit_progress(
            progress_callback,
            {
                "event": "load_planning",
                "job": job.name,
                "mode": mode,
                "workbook": str(workbook_path),
                "sheet": worksheet.title,
            },
        )
        header_map, header_issues, header_stats = _map_headers(job, worksheet)
        retry_status_index = _retry_status_index(
            worksheet,
            job.input.header_row,
            retry_statuses=retry_statuses,
        )
        _emit_progress(
            progress_callback,
            {
                "event": "load_headers",
                "matched": header_stats["matched"],
                "columns": len(job.columns),
                "required_matched": header_stats["required_matched"],
                "required": header_stats["required"],
                "aliases": header_stats["aliases"],
                "issues": len(header_issues),
            },
        )
        reference_indexes = _build_reference_indexes(
            db_path,
            job,
            progress_callback=progress_callback,
        )
        value_set_indexes = _build_value_set_indexes(
            job,
            progress_callback=progress_callback,
        )
        rows: list[MaterialCreateCompositionRow] = []
        issues: list[LoadIssue] = list(header_issues)
        rows_scanned = 0
        error_rows = 0
        if not header_issues:
            for row_number, row_values in _iter_data_rows(worksheet, job.input.header_row):
                if not _include_retry_row(row_values, retry_status_index, retry_statuses):
                    continue
                if limit is not None and rows_scanned >= limit:
                    break
                rows_scanned += 1
                values, row_issues = _row_values(
                    job,
                    row_number=row_number,
                    row_values=row_values,
                    header_map=header_map,
                    reference_indexes=reference_indexes,
                    value_set_indexes=value_set_indexes,
                )
                if row_issues:
                    error_rows += 1
                    issues.extend(row_issues)
                    continue
                rows.append(MaterialCreateCompositionRow(row=row_number, values=values))
        return {
            "sheet": worksheet.title,
            "rows_scanned": rows_scanned,
            "error_rows": error_rows,
            "issues": issues,
            "rows": tuple(rows),
        }
    finally:
        workbook.close()


def _material_create_composition_quote_rows(
    db_path: Path,
    job: LoadJob,
    workbook_path: Path,
    *,
    sheet: str | None,
    limit: int | None,
    mode: str,
    retry_statuses: set[str] | None,
    progress_callback: LoadProgressCallback | None,
) -> dict[str, Any]:
    _require_material_create_composition_quote_columns(job)
    workbook_path = workbook_path.expanduser()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    supplier_refs = _style_supplier_quote_references(db_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = _select_sheet(workbook, sheet)
        _emit_progress(
            progress_callback,
            {
                "event": "load_planning",
                "job": job.name,
                "mode": mode,
                "workbook": str(workbook_path),
                "sheet": worksheet.title,
            },
        )
        header_map, header_issues, header_stats = _map_headers(job, worksheet)
        retry_status_index = _retry_status_index(
            worksheet,
            job.input.header_row,
            retry_statuses=retry_statuses,
        )
        _emit_progress(
            progress_callback,
            {
                "event": "load_headers",
                "matched": header_stats["matched"],
                "columns": len(job.columns),
                "required_matched": header_stats["required_matched"],
                "required": header_stats["required"],
                "aliases": header_stats["aliases"],
                "issues": len(header_issues),
            },
        )
        reference_indexes = _build_reference_indexes(
            db_path,
            job,
            progress_callback=progress_callback,
        )
        value_set_indexes = _build_value_set_indexes(
            job,
            progress_callback=progress_callback,
        )
        rows: list[MaterialCreateCompositionQuoteRow] = []
        issues: list[LoadIssue] = list(header_issues)
        rows_scanned = 0
        error_rows = 0
        if not header_issues:
            for row_number, row_values in _iter_data_rows(worksheet, job.input.header_row):
                if not _include_retry_row(row_values, retry_status_index, retry_statuses):
                    continue
                if limit is not None and rows_scanned >= limit:
                    break
                rows_scanned += 1
                values, row_issues = _row_values(
                    job,
                    row_number=row_number,
                    row_values=row_values,
                    header_map=header_map,
                    reference_indexes=reference_indexes,
                    value_set_indexes=value_set_indexes,
                )
                if not row_issues:
                    row_issues.extend(
                        _resolve_style_supplier_quote_memberships(
                            values,
                            row_number=row_number,
                            refs=supplier_refs,
                        )
                    )
                if row_issues:
                    error_rows += 1
                    issues.extend(row_issues)
                    continue
                rows.append(MaterialCreateCompositionQuoteRow(row=row_number, values=values))
        return {
            "sheet": worksheet.title,
            "rows_scanned": rows_scanned,
            "error_rows": error_rows,
            "issues": issues,
            "rows": tuple(rows),
        }
    finally:
        workbook.close()


def _require_material_create_composition_columns(job: LoadJob) -> None:
    required = {
        "code",
        "product_type",
        "description",
        "compositions",
    }
    present = {column.key for column in job.columns}
    missing = sorted(required - present)
    if missing:
        raise ConfigError(
            f"load job[{job.name}] workflow material_create_with_composition "
            f"is missing columns: {', '.join(missing)}."
        )


def _require_material_create_composition_quote_columns(job: LoadJob) -> None:
    required = {
        "code",
        "product_type",
        "material_description",
        "compositions",
        "supplier",
        "agent",
        "node_name",
        "quote_description",
        "quote_factory",
        "set_production_quote",
    }
    present = {column.key for column in job.columns}
    missing = sorted(required - present)
    if missing:
        raise ConfigError(
            "load job"
            f"[{job.name}] workflow material_create_with_composition_and_quote "
            f"is missing columns: {', '.join(missing)}."
        )


def _material_create_composition_planned_requests(
    rows: tuple[MaterialCreateCompositionRow, ...],
) -> tuple[LoadRequest, ...]:
    requests: list[LoadRequest] = []
    for row in rows:
        requests.append(_material_create_request(row))
        requests.append(
            _material_composition_request(row, material_id="DRY-RUN-MATERIAL")
        )
    return tuple(requests)


def _material_create_composition_quote_planned_requests(
    rows: tuple[MaterialCreateCompositionQuoteRow, ...],
) -> tuple[LoadRequest, ...]:
    requests: list[LoadRequest] = []
    for row in rows:
        requests.append(_material_create_request(row))
        requests.append(
            _material_composition_request(row, material_id="DRY-RUN-MATERIAL")
        )
        quote_row = _material_supplier_quote_row(row, material_id="DRY-RUN-MATERIAL")
        requests.append(_supplier_quote_product_source_request(quote_row, root="material"))
        requests.append(
            _supplier_quote_item_request(quote_row, product_source_id="DRY-RUN-PRODUCT-SOURCE")
        )
        if not _is_blank(row.values.get("quote_factory")):
            requests.append(
                _supplier_quote_revision_request(quote_row, revision_id="DRY-RUN-REVISION")
            )
        if row.values.get("set_production_quote") is True:
            requests.append(
                _supplier_quote_production_request(
                    quote_row,
                    root="material",
                    supplier_item_id="DRY-RUN-SUPPLIER-ITEM",
                )
            )
    return tuple(requests)


def _execute_material_create_composition_workflow(
    auth_ctx: AuthContext,
    rows: tuple[MaterialCreateCompositionRow, ...],
    *,
    requests: list[LoadRequest],
    responses: list[LoadResponse],
    progress_callback: LoadProgressCallback | None,
) -> list[LoadIssue]:
    issues: list[LoadIssue] = []
    total = len(rows) * 2
    index = 0
    for row in rows:
        create_request = _material_create_request(row)
        index += 1
        create_response = _execute_chained_load_request(
            auth_ctx,
            create_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        material_id = _response_field(create_response, "id")
        if not create_response.ok:
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="material_create_failed",
                    message="Material create request failed.",
                )
            )
            continue
        if _is_blank(material_id):
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="material_id_missing",
                    message="Material create response did not include id.",
                )
            )
            continue

        composition_request = _material_composition_request(
            row,
            material_id=str(material_id),
        )
        index += 1
        composition_response = _execute_chained_load_request(
            auth_ctx,
            composition_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        if not composition_response.ok:
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="material_composition_create_failed",
                    message="Material composition request failed.",
                )
            )
    return issues


def _execute_material_create_composition_quote_workflow(
    auth_ctx: AuthContext,
    rows: tuple[MaterialCreateCompositionQuoteRow, ...],
    *,
    requests: list[LoadRequest],
    responses: list[LoadResponse],
    progress_callback: LoadProgressCallback | None,
) -> list[LoadIssue]:
    issues: list[LoadIssue] = []
    total = sum(
        4
        + (0 if _is_blank(row.values.get("quote_factory")) else 1)
        + (1 if row.values.get("set_production_quote") is True else 0)
        for row in rows
    )
    index = 0
    for row in rows:
        create_request = _material_create_request(row)
        index += 1
        create_response = _execute_chained_load_request(
            auth_ctx,
            create_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        material_id = _response_field(create_response, "id")
        if not create_response.ok:
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="material_create_failed",
                    message="Material create request failed.",
                )
            )
            continue
        if _is_blank(material_id):
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="material_id_missing",
                    message="Material create response did not include id.",
                )
            )
            continue

        material_id = str(material_id)
        composition_request = _material_composition_request(row, material_id=material_id)
        index += 1
        composition_response = _execute_chained_load_request(
            auth_ctx,
            composition_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        if not composition_response.ok:
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="material_composition_create_failed",
                    message="Material composition request failed.",
                )
            )

        quote_row = _material_supplier_quote_row(row, material_id=material_id)
        product_source_request = _supplier_quote_product_source_request(
            quote_row,
            root="material",
        )
        index += 1
        product_source_response = _execute_chained_load_request(
            auth_ctx,
            product_source_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        product_source_id = _response_field(product_source_response, "id")
        if not product_source_response.ok or _is_blank(product_source_id):
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="product_source_create_failed",
                    message="Product source request failed or returned no id.",
                )
            )
            continue

        item_request = _supplier_quote_item_request(
            quote_row,
            product_source_id=str(product_source_id),
        )
        index += 1
        item_response = _execute_chained_load_request(
            auth_ctx,
            item_request,
            index=index,
            total=total,
            requests=requests,
            responses=responses,
            progress_callback=progress_callback,
        )
        supplier_item_id = _response_field(item_response, "id")
        revision_id = _response_field(item_response, "latest_revision") or _response_field(
            item_response,
            "current_revision",
        )
        if not item_response.ok or _is_blank(supplier_item_id) or _is_blank(revision_id):
            issues.append(
                LoadIssue(
                    row=row.row,
                    code="supplier_item_create_failed",
                    message=(
                        "Supplier item request failed or returned no supplier item/revision id."
                    ),
                )
            )
            continue

        if not _is_blank(row.values.get("quote_factory")):
            revision_request = _supplier_quote_revision_request(
                quote_row,
                revision_id=str(revision_id),
            )
            index += 1
            revision_response = _execute_chained_load_request(
                auth_ctx,
                revision_request,
                index=index,
                total=total,
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            if not revision_response.ok:
                issues.append(
                    LoadIssue(
                        row=row.row,
                        code="supplier_item_revision_update_failed",
                        message="Supplier item revision update failed.",
                    )
                )

        if row.values.get("set_production_quote") is True:
            production_request = _supplier_quote_production_request(
                quote_row,
                root="material",
                supplier_item_id=str(supplier_item_id),
            )
            index += 1
            production_response = _execute_chained_load_request(
                auth_ctx,
                production_request,
                index=index,
                total=total,
                requests=requests,
                responses=responses,
                progress_callback=progress_callback,
            )
            if not production_response.ok:
                issues.append(
                    LoadIssue(
                        row=row.row,
                        code="production_quote_update_failed",
                        message="Material default quote update failed.",
                    )
                )
    return issues


def _material_create_request(
    row: MaterialCreateCompositionRow | MaterialCreateCompositionQuoteRow,
) -> LoadRequest:
    values = row.values
    body = {
        "code": values["code"],
        "product_type": values["product_type"],
    }
    description = values.get("description", values.get("material_description"))
    if not _is_blank(description):
        body["description"] = description
    return LoadRequest(
        row=row.row,
        method="POST",
        path="/v2/materials",
        body=body,
    )


def _material_composition_request(
    row: MaterialCreateCompositionRow | MaterialCreateCompositionQuoteRow,
    *,
    material_id: str,
) -> LoadRequest:
    return LoadRequest(
        row=row.row,
        method="POST",
        path=f"/v2/materials/{material_id}/technical_compositions",
        body=row.values["compositions"],
    )


def _material_supplier_quote_row(
    row: MaterialCreateCompositionQuoteRow,
    *,
    material_id: str,
) -> MaterialSupplierQuoteRow:
    values = dict(row.values)
    values["material"] = material_id
    values["description"] = values.get("quote_description")
    return MaterialSupplierQuoteRow(row=row.row, values=values)
