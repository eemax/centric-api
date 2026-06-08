from __future__ import annotations

import json
import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..auth import AuthContext
from ..config import ConfigError
from ..load_config import LoadConfig
from .excel import _select_sheet
from .models import (
    MAX_SAMPLES,
    REVIEW_COLUMN_HEADERS,
    REVIEW_STATUSES,
    LoadIssue,
    LoadMaterialized,
    LoadProgressCallback,
    LoadRequest,
    LoadResponse,
    LoadRunResult,
)
from .utils import _emit_progress, _is_blank, _lookup_key


def _execute_chained_load_request(
    auth_ctx: AuthContext,
    request: LoadRequest,
    *,
    index: int,
    total: int,
    requests: list[LoadRequest],
    responses: list[LoadResponse],
    progress_callback: LoadProgressCallback | None,
) -> LoadResponse:
    requests.append(request)
    started = time.time()
    try:
        response = auth_ctx.request(
            request.method,
            _request_url(auth_ctx, request.path),
            json_body=request.body,
        )
        status_code = response.status_code
        body = _response_body(response)
    except Exception as exc:
        status_code = 0
        body = _exception_body(exc)
    _emit_progress(
        progress_callback,
        {
            "event": "load_send",
            "index": index,
            "total": total,
            "row": request.row,
            "method": request.method,
            "path": request.path,
            "status_code": status_code,
            "elapsed_seconds": time.time() - started,
        },
    )
    load_response = LoadResponse(
        row=request.row,
        status_code=status_code,
        ok=0 < status_code < 400,
        body=body,
    )
    responses.append(load_response)
    return load_response


def _response_field(response: LoadResponse, key: str) -> str | None:
    if not isinstance(response.body, dict):
        return None
    value = response.body.get(key)
    return None if _is_blank(value) else str(value)


def _has_row_issues(issues: tuple[LoadIssue, ...]) -> bool:
    return any(issue.row is not None for issue in issues)


def _execute_requests(
    auth_ctx: AuthContext,
    requests: tuple[LoadRequest, ...],
    *,
    progress_callback: LoadProgressCallback | None = None,
) -> list[LoadResponse]:
    responses: list[LoadResponse] = []
    started = time.time()
    total = len(requests)
    for index, request in enumerate(requests, start=1):
        try:
            response = auth_ctx.request(
                request.method,
                _request_url(auth_ctx, request.path),
                json_body=request.body,
            )
            status_code = response.status_code
            body = _response_body(response)
        except Exception as exc:
            status_code = 0
            body = _exception_body(exc)
        _emit_progress(
            progress_callback,
            {
                "event": "load_send",
                "index": index,
                "total": total,
                "row": request.row,
                "method": request.method,
                "path": request.path,
                "status_code": status_code,
                "elapsed_seconds": time.time() - started,
            },
        )
        responses.append(
            LoadResponse(
                row=request.row,
                status_code=status_code,
                ok=0 < status_code < 400,
                body=body,
            )
        )
    return responses


def _write_review_workbook(
    result: LoadMaterialized,
    *,
    responses: tuple[LoadResponse, ...],
    run_id: str,
    processed_at: str,
    output_path: Path,
) -> Path:
    workbook = load_workbook(result.workbook_path)
    try:
        worksheet = _select_sheet(workbook, result.sheet)
        columns = _review_column_indexes(worksheet, result.header_row)
        _clear_review_rows(worksheet, columns, result.header_row)
        requests_by_row = {request.row: request for request in result.requests}
        issues_by_row: dict[int, list[LoadIssue]] = {}
        for issue in result.issues:
            if issue.row is not None:
                issues_by_row.setdefault(issue.row, []).append(issue)
        for row_number, issues in issues_by_row.items():
            _write_review_row(
                worksheet,
                row_number,
                columns,
                run_id=run_id,
                status="validation_error",
                status_code=None,
                message="; ".join(issue.message for issue in issues),
                request_path=requests_by_row.get(row_number).path
                if row_number in requests_by_row
                else "",
                response_id="",
                processed_at=processed_at,
            )
        for response in responses:
            request = requests_by_row.get(response.row)
            status = "success" if response.ok else "failed"
            _write_review_row(
                worksheet,
                response.row,
                columns,
                run_id=run_id,
                status=status,
                status_code=response.status_code,
                message=_response_message(response),
                request_path=request.path if request else "",
                response_id=_response_id(response),
                processed_at=processed_at,
            )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = _temp_output_path(output_path)
        try:
            workbook.save(temp_path)
            temp_path.replace(output_path)
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise
        return output_path
    finally:
        workbook.close()


def _review_column_indexes(worksheet: Any, header_row: int) -> dict[str, int]:
    existing: dict[str, int] = {}
    for cell in worksheet[header_row]:
        value = str(cell.value or "").strip()
        if value:
            existing[_lookup_key(value)] = int(cell.column)
    next_column = int(worksheet.max_column) + 1
    indexes: dict[str, int] = {}
    for header in REVIEW_COLUMN_HEADERS:
        key = _lookup_key(header)
        column = existing.get(key)
        if column is None:
            column = next_column
            worksheet.cell(row=header_row, column=column, value=header)
            next_column += 1
        indexes[header] = column
    return indexes


def _clear_review_rows(worksheet: Any, columns: dict[str, int], header_row: int) -> None:
    for row_number in range(header_row + 1, int(worksheet.max_row) + 1):
        for column in columns.values():
            worksheet.cell(row=row_number, column=column).value = None


def _write_review_row(
    worksheet: Any,
    row_number: int,
    columns: dict[str, int],
    *,
    run_id: str,
    status: str,
    status_code: int | None,
    message: str,
    request_path: str,
    response_id: str,
    processed_at: str,
) -> None:
    if status not in REVIEW_STATUSES:
        raise ConfigError(f"Unknown load review status: {status}")
    values = {
        "_cent_load_run_id": run_id,
        "_cent_load_status": status,
        "_cent_load_status_code": status_code,
        "_cent_load_message": _safe_review_text(message),
        "_cent_load_request_path": _safe_review_text(request_path),
        "_cent_load_response_id": _safe_review_text(response_id),
        "_cent_load_processed_at": processed_at,
    }
    for header, value in values.items():
        worksheet.cell(row=row_number, column=columns[header], value=value)


def _response_message(response: LoadResponse) -> str:
    if response.ok:
        return "ok"
    body = response.body
    if isinstance(body, dict):
        for key in ("error", "message", "detail"):
            value = body.get(key)
            if not _is_blank(value):
                return str(value)
    if _is_blank(body):
        return f"HTTP {response.status_code}"
    return str(body)


def _response_id(response: LoadResponse) -> str:
    if not isinstance(response.body, dict):
        return ""
    value = response.body.get("id")
    return "" if _is_blank(value) else str(value)


def _request_url(auth_ctx: AuthContext, path: str) -> str:
    normalized = path.strip().strip("/")
    return f"{auth_ctx.base_url}/api/{normalized}"


def _response_body(response: Any) -> Any:
    if not response.text:
        return None
    try:
        return response.json()
    except ValueError:
        return response.text


def _exception_body(exc: Exception) -> dict[str, str]:
    return {
        "error": str(exc),
        "type": exc.__class__.__name__,
    }


def _write_requests(path: Path, requests: tuple[LoadRequest, ...]) -> None:
    _write_jsonl(path, [_request_record(request) for request in requests])


def _write_responses(path: Path, responses: tuple[LoadResponse, ...]) -> None:
    _write_jsonl(path, [_response_record(response) for response in responses])


def _write_summary(path: Path, result: LoadRunResult, config: LoadConfig) -> None:
    payload = run_record(result)
    payload["config"] = [str(path) for path in config.paths]
    _write_text_atomic(path, json.dumps(payload, default=str, indent=2, sort_keys=True) + "\n")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    temp_path = _temp_output_path(path)
    try:
        with temp_path.open("w", encoding="utf-8") as fh:
            for row in rows:
                fh.write(json.dumps(row, default=str, sort_keys=True))
                fh.write("\n")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _write_text_atomic(path: Path, text: str) -> None:
    temp_path = _temp_output_path(path)
    try:
        temp_path.write_text(text, encoding="utf-8")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _temp_output_path(path: Path) -> Path:
    return path.parent / f".{path.name}.tmp"


def _safe_review_text(value: str) -> str:
    if not value:
        return value
    return f"'{value}" if value[0] in {"=", "+", "-", "@"} else value


def materialized_record(result: LoadMaterialized) -> dict[str, Any]:
    return {
        "job": result.job_name,
        "title": result.title,
        "workbook": str(result.workbook_path),
        "sheet": result.sheet,
        "header_row": result.header_row,
        "rows_scanned": result.rows_scanned,
        "valid_rows": result.valid_rows,
        "error_rows": result.error_rows,
        "issues": [_issue_record(issue) for issue in result.issues],
        "request_samples": [_request_record(request) for request in result.requests[:MAX_SAMPLES]],
    }


def run_record(result: LoadRunResult) -> dict[str, Any]:
    return {
        "run_id": result.run_id,
        "job": result.job_name,
        "title": result.title,
        "mode": result.mode,
        "dry_run": result.dry_run,
        "workbook": str(result.workbook_path),
        "sheet": result.sheet,
        "started_at": result.started_at,
        "finished_at": result.finished_at,
        "rows_scanned": result.rows_scanned,
        "valid_rows": result.valid_rows,
        "error_rows": result.error_rows,
        "requests": result.request_count,
        "successes": result.success_count,
        "failures": result.failure_count,
        "run_dir": str(result.run_dir),
        "review_workbook": str(result.review_path) if result.review_path else None,
        "issues": [_issue_record(issue) for issue in result.issues],
        "request_samples": [_request_record(request) for request in result.requests[:MAX_SAMPLES]],
        "response_samples": [
            _response_record(response) for response in result.responses[:MAX_SAMPLES]
        ],
    }


def _issue_record(issue: LoadIssue) -> dict[str, Any]:
    return {
        "row": issue.row,
        "code": issue.code,
        "message": issue.message,
        "column": issue.column,
        "sample": issue.sample,
    }


def _request_record(request: LoadRequest) -> dict[str, Any]:
    return {
        "row": request.row,
        "method": request.method,
        "path": request.path,
        "body": request.body,
    }


def _response_record(response: LoadResponse) -> dict[str, Any]:
    return {
        "row": response.row,
        "status_code": response.status_code,
        "ok": response.ok,
        "body": response.body,
    }
