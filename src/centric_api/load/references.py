from __future__ import annotations

import json
import re
import sqlite3
import unicodedata
from collections.abc import Callable
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import ConfigError, runtime_path
from ..load_config import LoadColumn, LoadJob, LoadResolve, LoadScope
from ..store import connect_readonly, endpoint_has_cache_evidence, table_exists
from .composition import _parse_composition_entries
from .excel import _cell_value
from .models import (
    LOAD_VALUE_SETS_DIR,
    MAX_SAMPLES,
    LoadIssue,
    LoadProgressCallback,
    LoadValueSetIndex,
)
from .utils import _emit_progress, _extract_path, _is_blank, _json_dict, _lookup_key


def _row_values(
    job: LoadJob,
    *,
    row_number: int,
    row_values: tuple[Any, ...],
    header_map: dict[str, int],
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
    value_set_indexes: dict[str, LoadValueSetIndex],
) -> tuple[dict[str, Any], list[LoadIssue]]:
    values: dict[str, Any] = {}
    issues: list[LoadIssue] = []
    for column in job.columns:
        raw_value = _cell_value(row_values, header_map.get(column.key))
        if _is_blank(raw_value):
            if column.required:
                issues.append(
                    LoadIssue(
                        row=row_number,
                        code="missing_required_value",
                        column=column.key,
                        message=f"Missing required value for {column.header}.",
                    )
                )
            values[column.key] = None
            continue
        parsed = _parse_value(column, raw_value, row_number)
        if isinstance(parsed, LoadIssue):
            issues.append(parsed)
            continue
        values[column.key] = parsed

    for column in job.columns:
        parsed = values.get(column.key)
        if _is_blank(parsed):
            continue
        if column.type == "ref":
            resolved = _resolve_value(
                column,
                parsed,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.type == "ref_or_id":
            resolved = _resolve_ref_or_id(
                column,
                parsed,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.type == "scoped_ref":
            resolved = _resolve_scoped_ref(
                column,
                parsed,
                values=values,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.type == "composition_list":
            resolved = _resolve_composition_list(
                column,
                parsed,
                row_number=row_number,
                reference_indexes=reference_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
        elif column.value_set is not None:
            resolved = _resolve_value_set(
                column,
                parsed,
                row_number=row_number,
                value_set_indexes=value_set_indexes,
            )
            if isinstance(resolved, LoadIssue):
                issues.append(resolved)
                continue
            values[column.key] = resolved
    return values, issues


def _parse_value(column: LoadColumn, raw_value: Any, row_number: int) -> Any | LoadIssue:
    if column.type in {"text", "ref", "ref_or_id", "scoped_ref"}:
        return str(raw_value).strip()
    if column.type == "number":
        try:
            value = Decimal(_normalize_number_text(raw_value))
        except (InvalidOperation, ValueError, AttributeError):
            return LoadIssue(
                row=row_number,
                code="invalid_number",
                column=column.key,
                message=f"Value for {column.header} must be numeric.",
                sample=raw_value,
            )
        return int(value) if value == value.to_integral_value() else float(value)
    if column.type == "boolean":
        if isinstance(raw_value, bool):
            return raw_value
        text = str(raw_value).strip().casefold()
        if text in {"true", "yes", "y", "1"}:
            return True
        if text in {"false", "no", "n", "0"}:
            return False
        return LoadIssue(
            row=row_number,
            code="invalid_boolean",
            column=column.key,
            message=f"Value for {column.header} must be boolean.",
            sample=raw_value,
        )
    if column.type == "composition_list":
        return _parse_composition_entries(column, raw_value, row_number)
    return raw_value


def _normalize_number_text(value: Any) -> str:
    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if "," not in text:
        return text
    if "." not in text:
        return text.replace(",", ".")
    if text.rfind(",") > text.rfind("."):
        return text.replace(".", "").replace(",", ".")
    return text.replace(",", "")


def _resolve_value(
    column: LoadColumn,
    value: Any,
    *,
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> str | LoadIssue:
    resolve = column.resolve
    if resolve is None:
        raise ConfigError(f"Column {column.key} is missing resolve config.")
    matches = reference_indexes.get(_resolve_key(resolve), {}).get(_lookup_key(str(value)), [])
    if not matches:
        return LoadIssue(
            row=row_number,
            code="ref_not_found",
            column=column.key,
            message=(
                f"{column.header} {value!r} was not found in {resolve.endpoint}.{resolve.match}."
            ),
        )
    if len(matches) > 1:
        return LoadIssue(
            row=row_number,
            code="ref_ambiguous",
            column=column.key,
            message=(
                f"{column.header} {value!r} matched {len(matches)} records in "
                f"{resolve.endpoint}.{resolve.match}."
            ),
            sample=[match.get("id") for match in matches[:MAX_SAMPLES]],
        )
    resolved = matches[0].get(resolve.output)
    if _is_blank(resolved):
        return LoadIssue(
            row=row_number,
            code="ref_output_blank",
            column=column.key,
            message=f"Resolved {resolve.endpoint} record has blank {resolve.output!r}.",
            sample=matches[0].get("id"),
        )
    return str(resolved).strip()


def _resolve_ref_or_id(
    column: LoadColumn,
    value: Any,
    *,
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> str | LoadIssue:
    resolve = column.resolve
    if resolve is None:
        raise ConfigError(f"Column {column.key} is missing resolve config.")
    text = str(value).strip()
    direct_matches = reference_indexes.get(_resolve_direct_key(resolve), {}).get(
        _lookup_key(text),
        [],
    )
    if len(direct_matches) == 1:
        resolved = direct_matches[0].get(resolve.output)
        return str(resolved).strip()
    if len(direct_matches) > 1:
        return LoadIssue(
            row=row_number,
            code="ref_id_ambiguous",
            column=column.key,
            message=(
                f"{column.header} {text!r} matched {len(direct_matches)} records in "
                f"{resolve.endpoint}.{resolve.output}."
            ),
            sample=[match.get("id") for match in direct_matches[:MAX_SAMPLES]],
        )
    resolved = _resolve_value(
        column,
        value,
        row_number=row_number,
        reference_indexes=reference_indexes,
    )
    if isinstance(resolved, LoadIssue) and resolved.code == "ref_not_found":
        return LoadIssue(
            row=row_number,
            code="ref_or_id_not_found",
            column=column.key,
            message=(
                f"{column.header} {text!r} was not found in {resolve.endpoint} by "
                f"{resolve.output} or {resolve.match}."
            ),
        )
    return resolved


def _resolve_scoped_ref(
    column: LoadColumn,
    value: Any,
    *,
    values: dict[str, Any],
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> str | LoadIssue:
    resolve = column.resolve
    if resolve is None or resolve.scope is None:
        raise ConfigError(f"Column {column.key} is missing scoped resolve config.")
    scope = resolve.scope
    scope_value = values.get(scope.column)
    if _is_blank(scope_value):
        return LoadIssue(
            row=row_number,
            code="scope_value_missing",
            column=column.key,
            message=f"{column.header} requires a value for scope column {scope.column!r}.",
        )

    candidates = reference_indexes.get(_resolve_key(resolve), {}).get(_lookup_key(str(value)), [])
    if not candidates:
        return LoadIssue(
            row=row_number,
            code="scoped_ref_not_found",
            column=column.key,
            message=(
                f"{column.header} {value!r} was not found in {resolve.endpoint}.{resolve.match}."
            ),
        )

    scope_index = reference_indexes.get(_scope_index_key(scope), {})
    scoped_matches: list[dict[str, Any]] = []
    missing_scope_refs: list[str] = []
    for candidate in candidates:
        scope_ref = _extract_path(candidate, scope.via)
        if _is_blank(scope_ref):
            continue
        scope_payloads = scope_index.get(_lookup_key(str(scope_ref)), [])
        if not scope_payloads:
            missing_scope_refs.append(str(scope_ref).strip())
            continue
        scope_lookup = _lookup_key(str(scope_value))
        if any(
            _lookup_key(str(_extract_path(scope_payload, scope.match))) == scope_lookup
            for scope_payload in scope_payloads
        ):
            scoped_matches.append(candidate)

    if not scoped_matches and missing_scope_refs:
        return LoadIssue(
            row=row_number,
            code="scope_ref_missing",
            column=column.key,
            message=(
                f"{column.header} {value!r} matched {resolve.endpoint} records, but "
                f"referenced {scope.endpoint} records were missing for "
                f"{resolve.endpoint}.{scope.via}."
            ),
            sample=sorted(set(missing_scope_refs))[:MAX_SAMPLES],
        )
    if not scoped_matches:
        return LoadIssue(
            row=row_number,
            code="scoped_ref_not_in_scope",
            column=column.key,
            message=(
                f"{column.header} {value!r} was not found under {scope.column} "
                f"{scope_value!r} via {resolve.endpoint}.{scope.via} -> "
                f"{scope.endpoint}.{scope.match}."
            ),
        )
    if len(scoped_matches) > 1:
        return LoadIssue(
            row=row_number,
            code="scoped_ref_ambiguous",
            column=column.key,
            message=(
                f"{column.header} {value!r} under {scope.column} {scope_value!r} "
                f"matched {len(scoped_matches)} records."
            ),
            sample=[match.get("id") for match in scoped_matches[:MAX_SAMPLES]],
        )

    resolved = scoped_matches[0].get(resolve.output)
    if _is_blank(resolved):
        return LoadIssue(
            row=row_number,
            code="scoped_ref_output_blank",
            column=column.key,
            message=f"Resolved {resolve.endpoint} record has blank {resolve.output!r}.",
            sample=scoped_matches[0].get("id"),
        )
    return str(resolved).strip()


def _resolve_composition_list(
    column: LoadColumn,
    entries: list[tuple[Decimal, str]],
    *,
    row_number: int,
    reference_indexes: dict[str, dict[str, list[dict[str, Any]]]],
) -> list[dict[str, Any]] | LoadIssue:
    resolve = column.resolve
    if resolve is None:
        raise ConfigError(f"Column {column.key} is missing resolve config.")
    resolved_entries: dict[str, dict[str, Any]] = {}
    for percentage, name in entries:
        matches = _composition_reference_matches(
            reference_indexes.get(_resolve_key(resolve), {}),
            resolve,
            name,
        )
        if not matches:
            return LoadIssue(
                row=row_number,
                code="composition_not_found",
                column=column.key,
                message=(
                    f"Composition {name!r} was not found in {resolve.endpoint}.{resolve.match}."
                ),
            )
        if len(matches) > 1:
            return LoadIssue(
                row=row_number,
                code="composition_ambiguous",
                column=column.key,
                message=(
                    f"Composition {name!r} matched {len(matches)} records in "
                    f"{resolve.endpoint}.{resolve.match}."
                ),
                sample=[match.get("id") for match in matches[:MAX_SAMPLES]],
            )
        resolved = matches[0].get(resolve.output)
        if _is_blank(resolved):
            return LoadIssue(
                row=row_number,
                code="composition_output_blank",
                column=column.key,
                message=f"Resolved {resolve.endpoint} record has blank {resolve.output!r}.",
                sample=matches[0].get("id"),
            )
        resolved_id = str(resolved).strip()
        existing = resolved_entries.get(resolved_id)
        if existing is None:
            resolved_entries[resolved_id] = {
                "percentage": _number_value(percentage),
                "composition": resolved_id,
            }
        else:
            existing["percentage"] = _number_value(
                Decimal(str(existing["percentage"])) + percentage
            )
    return list(resolved_entries.values())


def _composition_reference_matches(
    reference_index: dict[str, list[dict[str, Any]]],
    resolve: LoadResolve,
    name: str,
) -> list[dict[str, Any]]:
    matches = reference_index.get(_lookup_key(name), [])
    if matches:
        return matches

    canonical_name = _composition_lookup_key(name)
    if not canonical_name:
        return []

    canonical_matches: list[dict[str, Any]] = []
    seen: set[str] = set()
    for payloads in reference_index.values():
        for payload in payloads:
            value = _extract_path(payload, resolve.match)
            if _is_blank(value) or _composition_lookup_key(str(value)) != canonical_name:
                continue
            dedupe_key = str(payload.get(resolve.output) or payload.get("id") or id(payload))
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)
            canonical_matches.append(payload)
    return canonical_matches


def _composition_lookup_key(value: str) -> str:
    tokens = re.findall(r"[a-z0-9]+", value.casefold())
    return " ".join(sorted(tokens))


def _number_value(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def _resolve_value_set(
    column: LoadColumn,
    value: Any,
    *,
    row_number: int,
    value_set_indexes: dict[str, LoadValueSetIndex],
) -> str | LoadIssue:
    value_set = column.value_set
    if value_set is None:
        raise ConfigError(f"Column {column.key} is missing value_set config.")
    index = value_set_indexes[value_set.name]
    text = str(value).strip()
    resolved = (
        index.exact.get(text)
        or index.normalized.get(_normalize_value_set_key(text))
        or index.loose.get(_loose_value_set_key(text))
    )
    if resolved is None:
        return LoadIssue(
            row=row_number,
            code="value_set_not_found",
            column=column.key,
            message=f"{column.header} {text!r} was not found in value set {value_set.name}.",
            sample=index.values[:MAX_SAMPLES],
        )
    return resolved


def _build_value_set_indexes(
    job: LoadJob,
    *,
    progress_callback: LoadProgressCallback | None = None,
) -> dict[str, LoadValueSetIndex]:
    value_set_names = {
        column.value_set.name for column in job.columns if column.value_set is not None
    }
    indexes: dict[str, LoadValueSetIndex] = {}
    for name in sorted(value_set_names):
        path = runtime_path(LOAD_VALUE_SETS_DIR / f"{name}.xlsx")
        indexes[name] = _load_value_set_index(name, path)
        _emit_progress(
            progress_callback,
            {
                "event": "load_values",
                "name": name,
                "path": str(path),
                "values": len(indexes[name].values),
            },
        )
    return indexes


def _load_value_set_index(name: str, path: Path) -> LoadValueSetIndex:
    if not path.is_file():
        raise ConfigError(f"Load value set {name!r} not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        values: list[str] = []
        seen_exact: set[str] = set()
        for (raw_value,) in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
            if _is_blank(raw_value):
                continue
            value = str(raw_value).strip()
            if value in seen_exact:
                continue
            seen_exact.add(value)
            values.append(value)
    finally:
        workbook.close()

    if not values:
        raise ConfigError(f"Load value set {name!r} has no values: {path}")
    exact = {value: value for value in values}
    normalized = _value_set_lookup(
        name,
        path,
        values,
        key_func=_normalize_value_set_key,
        label="normalized",
    )
    loose = _value_set_lookup(
        name,
        path,
        values,
        key_func=_loose_value_set_key,
        label="loose",
    )
    return LoadValueSetIndex(
        name=name,
        path=path,
        values=tuple(values),
        exact=exact,
        normalized=normalized,
        loose=loose,
    )


def _value_set_lookup(
    name: str,
    path: Path,
    values: list[str],
    *,
    key_func: Callable[[str], str],
    label: str,
) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for value in values:
        key = key_func(value)
        if not key:
            continue
        existing = lookup.get(key)
        if existing is not None and existing != value:
            raise ConfigError(
                f"Load value set {name!r} has ambiguous {label} values "
                f"{existing!r} and {value!r}: {path}"
            )
        lookup[key] = value
    return lookup


def _normalize_value_set_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).replace("\xa0", " ")
    return " ".join(normalized.strip().casefold().split())


def _loose_value_set_key(value: str) -> str:
    normalized = _normalize_value_set_key(value)
    tokens = re.findall(r"[a-z0-9]+", normalized)
    return "".join(_singular_value_set_token(token) for token in tokens)


def _singular_value_set_token(token: str) -> str:
    if len(token) > 3 and token.endswith("ies"):
        return f"{token[:-3]}y"
    if len(token) > 4 and token.endswith(("ches", "shes")):
        return token[:-2]
    if len(token) > 3 and token.endswith(("ses", "xes", "zes")):
        return token[:-2]
    if len(token) > 3 and token.endswith("s") and not token.endswith("ss"):
        return token[:-1]
    return token


def _build_reference_indexes(
    db_path: Path,
    job: LoadJob,
    *,
    progress_callback: LoadProgressCallback | None = None,
) -> dict[str, dict[str, list[dict[str, Any]]]]:
    ref_columns = [column for column in job.columns if column.resolve is not None]
    refs = [column.resolve for column in ref_columns]
    if not refs:
        return {}
    with connect_readonly(db_path) as conn:
        if not table_exists(conn, "endpoint_records"):
            raise ConfigError(
                "Load reference resolution requires endpoint_records. Run fetch first."
            )
        missing_endpoints = sorted(
            {
                endpoint
                for resolve in refs
                for endpoint in _resolve_required_endpoints(resolve)
                if not endpoint_has_cache_evidence(conn, endpoint)
            }
        )
        if missing_endpoints:
            raise ConfigError(
                "Load reference resolution requires cached endpoint records for: "
                f"{', '.join(missing_endpoints)}. Run centric-api fetch for those endpoints first."
            )
        indexes: dict[str, dict[str, list[dict[str, Any]]]] = {}
        for resolve in refs:
            key = _resolve_key(resolve)
            if key not in indexes:
                indexes[key] = _reference_index(conn, resolve)
                _emit_progress(
                    progress_callback,
                    {
                        "event": "load_refs",
                        "endpoint": resolve.endpoint,
                        "match": resolve.match,
                        "output": resolve.output,
                        "filters": resolve.filters or {},
                        "matched": sum(len(matches) for matches in indexes[key].values()),
                        "values": len(indexes[key]),
                    },
                )
            if resolve.scope is not None:
                scope_key = _scope_index_key(resolve.scope)
                if scope_key not in indexes:
                    indexes[scope_key] = _reference_index(
                        conn,
                        LoadResolve(
                            endpoint=resolve.scope.endpoint,
                            match=resolve.scope.output,
                            output=resolve.scope.output,
                        ),
                    )
        for column in ref_columns:
            if column.type != "ref_or_id" or column.resolve is None:
                continue
            direct_key = _resolve_direct_key(column.resolve)
            if direct_key not in indexes:
                indexes[direct_key] = _reference_index(
                    conn,
                    column.resolve,
                    match_path=column.resolve.output,
                )
        return indexes


def _resolve_required_endpoints(resolve: LoadResolve) -> tuple[str, ...]:
    if resolve.scope is None:
        return (resolve.endpoint,)
    return (resolve.endpoint, resolve.scope.endpoint)


def _reference_index(
    conn: sqlite3.Connection,
    resolve: LoadResolve,
    *,
    match_path: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    rows = conn.execute(
        """
        SELECT payload_json
        FROM endpoint_records
        WHERE endpoint = ?
        ORDER BY record_id
        """,
        [resolve.endpoint],
    ).fetchall()
    index: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        payload = _json_dict(row["payload_json"])
        if not _matches_resolve_filters(payload, resolve):
            continue
        value = _extract_path(payload, match_path or resolve.match)
        if _is_blank(value):
            continue
        index.setdefault(_lookup_key(str(value)), []).append(payload)
    return index


def _matches_resolve_filters(payload: dict[str, Any], resolve: LoadResolve) -> bool:
    if not resolve.filters:
        return True
    return all(
        _extract_path(payload, path) == expected for path, expected in resolve.filters.items()
    )


def _resolve_key(resolve: LoadResolve) -> str:
    filters = json.dumps(resolve.filters or {}, default=str, sort_keys=True)
    return f"{resolve.endpoint}:{resolve.match}:{resolve.output}:{filters}"


def _resolve_direct_key(resolve: LoadResolve) -> str:
    filters = json.dumps(resolve.filters or {}, default=str, sort_keys=True)
    return f"{resolve.endpoint}:{resolve.output}:{resolve.output}:{filters}"


def _scope_index_key(scope: LoadScope) -> str:
    return f"{scope.endpoint}:{scope.output}:{scope.output}:scope"
